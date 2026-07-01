# ======================================================
# appREGIONAL.py
# PARTE 1 DE 5
# CONFIGURACIÓN GENERAL, LIBRERÍAS, CONSTANTES,
# ENCABEZADOS, ESTILOS, LOGOS Y FUNCIONES BASE
# ======================================================

import streamlit as st
import pandas as pd
import plotly.express as px
import folium
import os
import base64
import unicodedata
import re
import textwrap
import glob
import zipfile

from io import BytesIO
from datetime import datetime, date
from streamlit_folium import st_folium

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.drawing.image import Image as XLImage

# ======================================================
# CONFIGURACIÓN GENERAL STREAMLIT
# ======================================================

st.set_page_config(
    page_title="PUMI 2026 - Verificación Regional",
    page_icon="🔎",
    layout="wide",
    initial_sidebar_state="expanded"
)


# ======================================================
# ARCHIVOS Y LOGOS
# Deben estar al mismo nivel que appREGIONAL.py
# ======================================================

LOGO_MINISTERIO = "Logo2.jpeg"
LOGO_PUMI = "logo_pumi.jpeg"
LOGO_FUERZA_PUBLICA = "Logo1.jpeg"
MARCA_AGUA_EXCEL = "marca_agua.png"
ARCHIVO_DATOS_IMPORTANTES = "Datos Importantes.xlsx"
ARCHIVO_MEP = "BASE DE DATOS MEP 2025.xlsx"
ARCHIVO_USUARIOS_CLAVES = "Usuarios y Claves PUMI.xlsx"
CARPETA_METAS_REGIONALES = "."
CARPETA_METAS_ALTERNATIVA = "metas"

NOMBRE_HOJA_EXCEL = "REGISTRO_PUMI_2026"


# ======================================================
# ENCABEZADOS BASE DE LA APP PUMI
# ======================================================

ENCABEZADOS_PUMI = [
    "ID",
    "Fecha Actividad",
    "Hora Actividad",
    "Dirección Regional",
    "Delegación",
    "Responde a",
    "Programa",
    "Actividad",
    "Provincia",
    "Cantón",
    "Distrito",
    "Tipo Lugar",
    "Lugar",
    "Centro Educativo",
    "Código Presupuestario",
    "Dirección Mapa",
    "Latitud",
    "Longitud",
    "Responsable",
    "Cantidad Participantes",
    "Cantidad Hombres",
    "Cantidad Mujeres",
    "Edad 10 a 18",
    "Edad 19 a 30",
    "Edad 31 a 45",
    "Edad 46 en adelante",
    "Instituciones Participantes",
    "Número de Referencia",
    "Número de Expediente Referencia",
    "Observaciones",
    "Usuario Registra"
]


# ======================================================
# ENCABEZADOS ADMINISTRATIVOS
# Estos campos se agregan para verificación.
# ======================================================

ENCABEZADOS_VALIDACION = [
    "Estado Verificación Regional",
    "Observaciones Verificación Regional",
    "Coordinador Regional",
    "Fecha Verificación Regional"
]


ENCABEZADOS_COMPLETOS = ENCABEZADOS_PUMI + ["Archivo Origen"] + ENCABEZADOS_VALIDACION


# ======================================================
# COLORES INSTITUCIONALES
# ======================================================

COLOR_AZUL = "#002B7F"
COLOR_AZUL_MEDIO = "#1E4FA3"
COLOR_AZUL_CLARO = "#DCE8FF"

COLOR_DORADO = "#B88A2A"
COLOR_DORADO_OSCURO = "#8A6418"
COLOR_DORADO_CLARO = "#F4E6C1"

COLOR_BLANCO = "#FFFFFF"
COLOR_GRIS = "#F4F6F8"
COLOR_GRIS_OSCURO = "#2F3542"

COLOR_VERDE = "#1F8A4C"
COLOR_ROJO = "#B42318"
COLOR_AMARILLO = "#F2C94C"


COLORES_VALIDACION = {
    "Pendiente de verificación": "#F2C94C",
    "Verificada para envío": "#1F8A4C",
    "Devuelta a delegación": "#B42318",
    "Con observaciones regionales": "#B88A2A"
}


COLORES_PROGRAMA = {
    "DARE": "purple",
    "GREAT": "orange",
    "GREAT CAMP": "orange",
    "MPAS": "darkblue",
    "PSCC": "blue",
    "VIF": "orange",
    "Política Pública": "green"
}


CENTROS_PROVINCIA = {
    "San José": [9.9281, -84.0907],
    "Alajuela": [10.0162, -84.2116],
    "Cartago": [9.8644, -83.9194],
    "Heredia": [10.0024, -84.1165],
    "Guanacaste": [10.6267, -85.4437],
    "Puntarenas": [9.9763, -84.8384],
    "Limón": [9.9917, -83.0360]
}


PROGRAMAS_BASE = [
    "DARE",
    "GREAT",
    "MPAS",
    "PSCC",
    "VIF",
    "Política Pública"
]

REGIONES_BASE = [
    "R1 San José Central",
    "R2 San José Norte",
    "R3 San José Sur",
    "R4 Alajuela",
    "R5 Cartago",
    "R6 Heredia",
    "R7 Chorotega",
    "R8 Puntarenas",
    "R9 Limón",
    "R10 Brunca",
    "R11 Chorotega Norte",
    "R12"
]

PROVINCIAS_BASE = [
    "San José",
    "Alajuela",
    "Cartago",
    "Heredia",
    "Guanacaste",
    "Puntarenas",
    "Limón"
]


# ======================================================
# SESSION STATE
# ======================================================

def inicializar_session_state_admin():
    if "df_admin" not in st.session_state:
        st.session_state.df_admin = pd.DataFrame(columns=ENCABEZADOS_COMPLETOS)

    if "archivo_admin_nombre" not in st.session_state:
        st.session_state.archivo_admin_nombre = ""

    if "filtros_admin_aplicados" not in st.session_state:
        st.session_state.filtros_admin_aplicados = {}

    if "mapa_imagen_referencia" not in st.session_state:
        st.session_state.mapa_imagen_referencia = None

    if "grafico_imagen_referencia" not in st.session_state:
        st.session_state.grafico_imagen_referencia = None


inicializar_session_state_admin()


# ======================================================
# FUNCIONES BASE
# ======================================================

def imagen_a_base64(ruta):
    if not os.path.exists(ruta):
        return ""

    try:
        with open(ruta, "rb") as archivo:
            return base64.b64encode(archivo.read()).decode()
    except Exception:
        return ""


def normalizar_texto(valor):
    if pd.isna(valor):
        return ""

    texto = str(valor).strip().upper()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(
        caracter for caracter in texto
        if not unicodedata.combining(caracter)
    )
    texto = " ".join(texto.split())

    return texto


def limpiar_nombre_archivo(valor):
    texto = str(valor).strip()

    if texto == "" or texto.lower() == "nan":
        return "SIN_DATO"

    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(
        caracter for caracter in texto
        if not unicodedata.combining(caracter)
    )

    texto = texto.upper()
    texto = re.sub(r"[^\w\s-]", "", texto)
    texto = re.sub(r"\s+", "_", texto)
    texto = texto.strip("_")

    if texto == "":
        return "SIN_DATO"

    return texto


def limpiar_coordenada(valor):
    try:
        if valor is None or str(valor).strip() == "":
            return None

        return float(str(valor).replace(",", ".").strip())

    except Exception:
        return None


def obtener_color_programa(programa):
    return COLORES_PROGRAMA.get(str(programa), "gray")


def generar_nombre_reporte(df, tipo="XLSX"):
    """
    Genera el nombre de descarga para la app regional.
    En esta etapa el archivo representa una Dirección Regional completa,
    por eso NO se incluye la delegación en el nombre.
    """
    extension = tipo.lower()

    if df is None or df.empty:
        return f"VERIFICACION_REGIONAL_PUMI_2026.{extension}"

    region = "VARIAS_REGIONES"

    if "Dirección Regional" in df.columns:
        regiones = (
            df["Dirección Regional"]
            .replace("", pd.NA)
            .dropna()
            .astype(str)
            .str.strip()
            .unique()
            .tolist()
        )

        regiones = [r for r in regiones if r != ""]

        if len(regiones) == 1:
            region = regiones[0]

    region_limpia = limpiar_nombre_archivo(region)

    return f"VERIFICACION_REGIONAL_PUMI_2026_{region_limpia}.{extension}"


# ======================================================
# ESTILOS CSS
# Misma línea visual de la app PUMI principal.
# ======================================================

st.markdown(
    f"""
    <style>

    .stApp {{
        background:
            radial-gradient(circle at top left, #FFFFFF 0, transparent 28%),
            radial-gradient(circle at top right, #DCE8FF 0, transparent 30%),
            linear-gradient(180deg, #FFFFFF 0%, #F4F7FB 48%, #FFFFFF 100%);
    }}

    section[data-testid="stSidebar"] {{
        background: linear-gradient(180deg, #FFFFFF 0%, #EAF1FF 45%, #FFFFFF 100%);
        border-right: 5px solid {COLOR_DORADO};
        box-shadow: 4px 0px 12px rgba(0,0,0,0.10);
    }}

    .sidebar-logo-pumi {{
        background: #FFFFFF;
        border-radius: 16px;
        padding: 12px;
        margin: 8px 0px 18px 0px;
        box-shadow: 0px 4px 14px rgba(0,0,0,0.10);
        text-align: center;
    }}

    .sidebar-logo-pumi img {{
        width: 100%;
        max-height: 145px;
        object-fit: contain;
    }}

    .encabezado-institucional {{
        background: #FFFFFF;
        border-radius: 22px;
        padding: 22px 34px;
        margin-bottom: 24px;
        box-shadow: 0px 8px 22px rgba(0,0,0,0.11);
        border-bottom: 7px solid {COLOR_DORADO};
        display: grid;
        grid-template-columns: 1fr 1fr 1fr;
        align-items: center;
        gap: 22px;
        min-height: 155px;
    }}

    .encabezado-logo-izq,
    .encabezado-logo-centro,
    .encabezado-logo-der {{
        height: 125px;
        display: flex;
        align-items: center;
    }}

    .encabezado-logo-izq {{
        justify-content: flex-start;
    }}

    .encabezado-logo-centro {{
        justify-content: center;
    }}

    .encabezado-logo-der {{
        justify-content: flex-end;
    }}

    .logo-ministerio-app {{
        width: 350px;
        height: 100px;
        object-fit: contain;
    }}

    .logo-pumi-app {{
        width: 300px;
        height: 100px;
        object-fit: contain;
    }}

    .logo-fp-app {{
        width: 400px;
        height: 150px;
        object-fit: contain;
    }}

    .titulo-principal {{
        background: linear-gradient(135deg, {COLOR_AZUL}, #243B63, {COLOR_DORADO});
        padding: 34px;
        border-radius: 22px;
        text-align: center;
        color: white;
        margin-bottom: 30px;
        box-shadow: 0px 8px 22px rgba(0,0,0,0.28);
        border: 3px solid {COLOR_BLANCO};
    }}

    .titulo-principal h1 {{
        font-size: 46px;
        margin-bottom: 8px;
        font-weight: 900;
        letter-spacing: 1.5px;
        color: white;
        text-shadow: 1px 2px 5px rgba(0,0,0,0.35);
    }}

    .titulo-principal h3 {{
        font-size: 25px;
        margin-top: 8px;
        font-weight: 600;
        color: white;
    }}

    .card-admin {{
        padding: 24px;
        border-radius: 20px;
        box-shadow: 0px 6px 16px rgba(0,0,0,0.13);
        margin-bottom: 20px;
        background: linear-gradient(135deg, #FFFFFF 0%, #F8FAFF 100%);
        border-left: 9px solid {COLOR_AZUL};
    }}

    .card-validacion {{
        padding: 22px;
        border-radius: 18px;
        box-shadow: 0px 5px 14px rgba(0,0,0,0.12);
        margin-bottom: 18px;
        background: linear-gradient(135deg, #FFFFFF 0%, #FFF8E8 100%);
        border-left: 8px solid {COLOR_DORADO};
    }}

    .subtitulo-admin {{
        color: {COLOR_AZUL};
        font-weight: 900;
        font-size: 29px;
        margin-bottom: 12px;
        text-align: center;
    }}

    .texto-admin {{
        color: {COLOR_GRIS_OSCURO};
        font-size: 17px;
        line-height: 1.7;
        text-align: justify;
    }}

    div[data-testid="stMetric"] {{
        background: linear-gradient(145deg, #FFFFFF 0%, #F3F6FF 100%);
        padding: 20px;
        border-radius: 20px;
        box-shadow: 0px 5px 15px rgba(0,0,0,0.13);
        border-bottom: 6px solid {COLOR_DORADO};
    }}

    div[data-baseweb="select"] > div {{
        background-color: #FFFFFF !important;
        border-radius: 12px !important;
        border: 1.8px solid #AFC3EE !important;
    }}

    input {{
        background-color: #FFFFFF !important;
        border-radius: 12px !important;
        border: 1.8px solid #AFC3EE !important;
    }}

    textarea {{
        background-color: #FFFFFF !important;
        border-radius: 12px !important;
        border: 1.8px solid {COLOR_DORADO_CLARO} !important;
    }}

    .stButton > button {{
        background: linear-gradient(90deg, {COLOR_AZUL}, {COLOR_DORADO});
        color: white;
        border-radius: 12px;
        border: none;
        font-weight: 800;
        padding: 0.7rem 1.2rem;
        box-shadow: 0px 4px 10px rgba(0,0,0,0.18);
    }}

    .stDownloadButton > button {{
        background: linear-gradient(90deg, {COLOR_DORADO}, {COLOR_AZUL});
        color: white;
        border-radius: 12px;
        border: none;
        font-weight: 800;
        padding: 0.7rem 1.2rem;
        box-shadow: 0px 4px 10px rgba(0,0,0,0.18);
    }}

    div[data-testid="stDataFrame"] {{
        border-radius: 18px;
        overflow: hidden;
        box-shadow: 0px 5px 16px rgba(0,0,0,0.13);
        background-color: white;
    }}

    iframe {{
        border-radius: 18px !important;
        box-shadow: 0px 5px 16px rgba(0,0,0,0.13);
        width: 100% !important;
    }}

    hr {{
        border: none;
        height: 3px;
        background: linear-gradient(90deg, {COLOR_AZUL}, {COLOR_BLANCO}, {COLOR_DORADO});
        margin-top: 25px;
        margin-bottom: 25px;
    }}

    h1, h2, h3 {{
        color: {COLOR_AZUL};
        font-weight: 900;
    }}

    </style>
    """,
    unsafe_allow_html=True
)


# ======================================================
# LOGOS Y ENCABEZADOS
# ======================================================

def mostrar_logo_sidebar():
    logo_pumi_b64 = imagen_a_base64(LOGO_PUMI)

    if logo_pumi_b64:
        st.sidebar.markdown(
            f"""
            <div class="sidebar-logo-pumi">
                <img src="data:image/jpeg;base64,{logo_pumi_b64}">
            </div>
            """,
            unsafe_allow_html=True
        )
    else:
        st.sidebar.warning("Logo PUMI no encontrado.")


def mostrar_encabezado_institucional():
    logo_min_b64 = imagen_a_base64(LOGO_MINISTERIO)
    logo_pumi_b64 = imagen_a_base64(LOGO_PUMI)
    logo_fp_b64 = imagen_a_base64(LOGO_FUERZA_PUBLICA)

    img_min = (
        f'<img class="logo-ministerio-app" src="data:image/jpeg;base64,{logo_min_b64}">'
        if logo_min_b64 else ""
    )

    img_pumi = (
        f'<img class="logo-pumi-app" src="data:image/jpeg;base64,{logo_pumi_b64}">'
        if logo_pumi_b64 else ""
    )

    img_fp = (
        f'<img class="logo-fp-app" src="data:image/jpeg;base64,{logo_fp_b64}">'
        if logo_fp_b64 else ""
    )

    st.markdown(
        f"""
        <div class="encabezado-institucional">
            <div class="encabezado-logo-izq">{img_min}</div>
            <div class="encabezado-logo-centro">{img_pumi}</div>
            <div class="encabezado-logo-der">{img_fp}</div>
        </div>
        """,
        unsafe_allow_html=True
    )


def mostrar_titulo_admin():
    st.markdown(
        """
        <div class="titulo-principal">
            <h1>🔎 P.U.M.I. 2026 - Panel de Verificación Regional</h1>
            <h3>Verificación regional previa al envío para validación nacional</h3>
        </div>
        """,
        unsafe_allow_html=True
    )
# ======================================================
# appREGIONAL.py
# PARTE 2 DE 5 CORREGIDA
# CARGA DE EXCEL PUMI, PREPARACIÓN DE DATOS,
# FILTROS GENERALES, MAPA CON TIPO DE MAPA
# Y DASHBOARD CON COLORES INSTITUCIONALES
# ======================================================


# ======================================================
# PREPARAR DATAFRAME ADMINISTRATIVO
# ======================================================

def preparar_dataframe_admin(df):
    df = df.copy()

    # Limpia nombres de columnas por si vienen con espacios al inicio/final
    df.columns = [str(col).strip() for col in df.columns]

    # Evita error cuando el Excel viene de la app original, regional o admin.
    # Agrega cualquier columna oficial que falte, incluyendo "Archivo Origen".
    for col in ENCABEZADOS_COMPLETOS:
        if col not in df.columns:
            df[col] = ""

    # Orden oficial de columnas para la app regional
    df = df.reindex(columns=ENCABEZADOS_COMPLETOS, fill_value="")

    for col in ENCABEZADOS_COMPLETOS:
        df[col] = df[col].fillna("").astype(str)

    columnas_numericas = [
        "ID",
        "Cantidad Participantes",
        "Cantidad Hombres",
        "Cantidad Mujeres",
        "Edad 10 a 18",
        "Edad 19 a 30",
        "Edad 31 a 45",
        "Edad 46 en adelante"
    ]

    for col in columnas_numericas:
        if col in df.columns:
            df[col] = pd.to_numeric(
                df[col],
                errors="coerce"
            ).fillna(0).astype(int)

    df["Estado Verificación Regional"] = df["Estado Verificación Regional"].replace(
        "",
        "Pendiente de verificación"
    )

    return df


# ======================================================
# CARGAR EXCEL GENERADO POR APP PUMI
# ======================================================

def cargar_excel_pumi_admin(archivo_excel):
    try:
        df = pd.read_excel(
            archivo_excel,
            sheet_name=NOMBRE_HOJA_EXCEL
        )
    except Exception:
        try:
            df = pd.read_excel(archivo_excel)
        except Exception as e:
            st.error("No se pudo leer el Excel cargado.")
            st.exception(e)
            return pd.DataFrame(columns=ENCABEZADOS_COMPLETOS)

    return preparar_dataframe_admin(df)


# ======================================================
# EXPORTAR EXCEL ADMINISTRATIVO
# CON COLORES EN COLUMNAS DE VALIDACIÓN
# ======================================================


def aplicar_marca_agua_como_fondo_excel(xlsx_bytes, ruta_imagen, hoja_xml="xl/worksheets/sheet1.xml"):
    """
    Inserta la marca de agua como fondo de hoja de Excel.
    A diferencia de worksheet.add_image(), este método coloca la imagen detrás
    de los datos y no tapa el contenido de las celdas.
    """
    if not os.path.exists(ruta_imagen):
        return xlsx_bytes

    try:
        entrada = BytesIO(xlsx_bytes)
        salida = BytesIO()
        nombre_media = "marca_agua_fondo.png"
        ruta_media = f"xl/media/{nombre_media}"
        ruta_rels = "xl/worksheets/_rels/sheet1.xml.rels"
        tipo_rel_imagen = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/image"

        with zipfile.ZipFile(entrada, "r") as zin, zipfile.ZipFile(salida, "w", zipfile.ZIP_DEFLATED) as zout:
            nombres = set(zin.namelist())

            for item in zin.infolist():
                nombre = item.filename

                if nombre in {hoja_xml, ruta_rels, "[Content_Types].xml", ruta_media}:
                    continue

                zout.writestr(item, zin.read(nombre))

            # --------------------------------------------------
            # Worksheet XML: agrega referencia de fondo <picture>
            # --------------------------------------------------
            hoja_contenido = zin.read(hoja_xml).decode("utf-8")

            if "xmlns:r=" not in hoja_contenido.split(">", 1)[0]:
                hoja_contenido = hoja_contenido.replace(
                    "<worksheet ",
                    '<worksheet xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" ',
                    1
                )

            if "<picture " not in hoja_contenido:
                hoja_contenido = hoja_contenido.replace(
                    "</worksheet>",
                    '<picture r:id="rIdMarcaAguaFondo"/></worksheet>'
                )

            zout.writestr(hoja_xml, hoja_contenido.encode("utf-8"))

            # --------------------------------------------------
            # Relaciones del worksheet
            # --------------------------------------------------
            if ruta_rels in nombres:
                rels_contenido = zin.read(ruta_rels).decode("utf-8")
                if "rIdMarcaAguaFondo" not in rels_contenido:
                    rels_contenido = rels_contenido.replace(
                        "</Relationships>",
                        f'<Relationship Id="rIdMarcaAguaFondo" Type="{tipo_rel_imagen}" Target="../media/{nombre_media}"/></Relationships>'
                    )
            else:
                rels_contenido = (
                    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                    '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                    f'<Relationship Id="rIdMarcaAguaFondo" Type="{tipo_rel_imagen}" Target="../media/{nombre_media}"/>'
                    '</Relationships>'
                )

            zout.writestr(ruta_rels, rels_contenido.encode("utf-8"))

            # --------------------------------------------------
            # Content Types: asegura soporte PNG
            # --------------------------------------------------
            content_types = zin.read("[Content_Types].xml").decode("utf-8")
            if 'Extension="png"' not in content_types:
                content_types = content_types.replace(
                    "</Types>",
                    '<Default Extension="png" ContentType="image/png"/></Types>'
                )
            zout.writestr("[Content_Types].xml", content_types.encode("utf-8"))

            # --------------------------------------------------
            # Imagen de fondo
            # --------------------------------------------------
            with open(ruta_imagen, "rb") as img:
                zout.writestr(ruta_media, img.read())

        salida.seek(0)
        return salida.getvalue()

    except Exception:
        return xlsx_bytes


def convertir_excel_admin(df):
    output = BytesIO()

    df_exportar = preparar_dataframe_admin(df)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_exportar.to_excel(
            writer,
            index=False,
            sheet_name=NOMBRE_HOJA_EXCEL
        )

    output.seek(0)

    workbook = load_workbook(output)
    worksheet = workbook[NOMBRE_HOJA_EXCEL]

    fill_header = PatternFill(
        start_color="002B7F",
        end_color="002B7F",
        fill_type="solid"
    )

    fill_header_admin = PatternFill(
        start_color="B88A2A",
        end_color="B88A2A",
        fill_type="solid"
    )

    font_header = Font(
        color="FFFFFF",
        bold=True
    )

    border = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF")
    )

    columnas_validacion = {
        "Estado Verificación Regional",
        "Observaciones Verificación Regional",
        "Coordinador Regional",
        "Fecha Verificación Regional"
    }

    for cell in worksheet[1]:
        if cell.value in columnas_validacion:
            cell.fill = fill_header_admin
        else:
            cell.fill = fill_header

        cell.font = font_header
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        cell.border = border

    colores_estado = {
        "Pendiente de verificación": "F2C94C",
        "Verificada para envío": "1F8A4C",
        "Devuelta a delegación": "B42318",
        "Con observaciones regionales": "B88A2A"
    }

    col_estado = None

    for idx, cell in enumerate(worksheet[1], start=1):
        if cell.value == "Estado Verificación Regional":
            col_estado = idx
            break

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )
            cell.border = border

        if col_estado:
            celda_estado = row[col_estado - 1]
            estado = str(celda_estado.value).strip()

            if estado in colores_estado:
                celda_estado.fill = PatternFill(
                    start_color=colores_estado[estado],
                    end_color=colores_estado[estado],
                    fill_type="solid"
                )
                celda_estado.font = Font(
                    color="FFFFFF",
                    bold=True
                )
                celda_estado.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            valor = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(valor))

        worksheet.column_dimensions[column_letter].width = min(
            max(max_length + 2, 14),
            45
        )

    worksheet.freeze_panes = "A2"

    # ==================================================
    # MARCA DE AGUA
    # Se aplica al final como fondo de hoja para que siempre
    # quede detrás de los datos y no tape el contenido.
    # ==================================================

    # ==================================================
    # MARCA DE AGUA PARA IMPRESIÓN
    # ==================================================

    worksheet.oddHeader.center.text = "Dirección de Programas Preventivos Policiales"
    worksheet.oddHeader.center.size = 24
    worksheet.oddHeader.center.font = "Arial,Bold"
    worksheet.oddHeader.center.color = "D9D9D9"

    worksheet.evenHeader.center.text = "Dirección de Programas Preventivos Policiales"
    worksheet.evenHeader.center.size = 24
    worksheet.evenHeader.center.font = "Arial,Bold"
    worksheet.evenHeader.center.color = "D9D9D9"

    # ==================================================
    # CONFIGURACIÓN DE IMPRESIÓN
    # ==================================================

    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True

    worksheet.page_margins.left = 0.25
    worksheet.page_margins.right = 0.25
    worksheet.page_margins.top = 0.75
    worksheet.page_margins.bottom = 0.75

    # ==================================================
    # BLOQUEAR TODAS LAS CELDAS Y PROTEGER HOJA/LIBRO
    # Contraseña: DPPP23
    # ==================================================

    for fila in worksheet.iter_rows():
        for celda in fila:
            celda.protection = Protection(locked=True)

    worksheet.protection.sheet = True
    worksheet.protection.password = "DPPP23"
    worksheet.protection.objects = True
    worksheet.protection.scenarios = True
    worksheet.protection.formatCells = False
    worksheet.protection.formatColumns = False
    worksheet.protection.formatRows = False
    worksheet.protection.insertColumns = False
    worksheet.protection.insertRows = False
    worksheet.protection.insertHyperlinks = False
    worksheet.protection.deleteColumns = False
    worksheet.protection.deleteRows = False
    worksheet.protection.sort = False
    worksheet.protection.autoFilter = False
    worksheet.protection.pivotTables = False
    worksheet.protection.selectLockedCells = True
    worksheet.protection.selectUnlockedCells = False

    workbook.security.lockStructure = True
    workbook.security.workbookPassword = "DPPP23"

    final_output = BytesIO()
    workbook.save(final_output)
    final_output.seek(0)

    excel_bytes = final_output.getvalue()
    excel_bytes = aplicar_marca_agua_como_fondo_excel(
        excel_bytes,
        MARCA_AGUA_EXCEL
    )

    return excel_bytes


# ======================================================
# LIMPIEZA PARA MÉTRICAS Y FILTROS
# ======================================================

def limpiar_dataframe_metricas_admin(df):
    df = df.copy()

    columnas_numericas = [
        "Cantidad Participantes",
        "Cantidad Hombres",
        "Cantidad Mujeres",
        "Edad 10 a 18",
        "Edad 19 a 30",
        "Edad 31 a 45",
        "Edad 46 en adelante"
    ]

    for col in columnas_numericas:
        if col in df.columns:
            df[col] = pd.to_numeric(
                df[col],
                errors="coerce"
            ).fillna(0)

    if "Fecha Actividad" in df.columns:
        df["Fecha Actividad"] = pd.to_datetime(
            df["Fecha Actividad"],
            errors="coerce",
            dayfirst=True
        )

    return df


# ======================================================
# FILTROS GENERALES DEL PANEL ADMIN
# ======================================================

def aplicar_filtros_admin(df):
    if df is None or df.empty:
        return df

    df_filtrado = df.copy()

    st.markdown("### Filtros de consulta")

    col1, col2, col3 = st.columns(3)

    with col1:
        filtro_region = st.multiselect(
            "Dirección Regional",
            sorted(
                df_filtrado["Dirección Regional"]
                .replace("", pd.NA)
                .dropna()
                .astype(str)
                .unique()
                .tolist()
            )
        )

        filtro_delegacion = st.multiselect(
            "Delegación",
            sorted(
                df_filtrado["Delegación"]
                .replace("", pd.NA)
                .dropna()
                .astype(str)
                .unique()
                .tolist()
            )
        )

    with col2:
        filtro_programa = st.multiselect(
            "Programa",
            sorted(
                df_filtrado["Programa"]
                .replace("", pd.NA)
                .dropna()
                .astype(str)
                .unique()
                .tolist()
            )
        )

        filtro_actividad = st.multiselect(
            "Actividad",
            sorted(
                df_filtrado["Actividad"]
                .replace("", pd.NA)
                .dropna()
                .astype(str)
                .unique()
                .tolist()
            )
        )

    with col3:
        filtro_estado = st.multiselect(
            "Estado de verificación",
            [
                "Pendiente de verificación",
                "Verificada para envío",
                "Devuelta a delegación",
                "Con observaciones regionales"
            ]
        )

        filtro_provincia = st.multiselect(
            "Provincia",
            sorted(
                df_filtrado["Provincia"]
                .replace("", pd.NA)
                .dropna()
                .astype(str)
                .unique()
                .tolist()
            )
        )

    colf1, colf2 = st.columns(2)

    with colf1:
        fecha_inicio = st.date_input(
            "Fecha inicio",
            value=None,
            format="DD/MM/YYYY"
        )

    with colf2:
        fecha_fin = st.date_input(
            "Fecha final",
            value=None,
            format="DD/MM/YYYY"
        )

    if filtro_region:
        df_filtrado = df_filtrado[
            df_filtrado["Dirección Regional"].isin(filtro_region)
        ]

    if filtro_delegacion:
        df_filtrado = df_filtrado[
            df_filtrado["Delegación"].isin(filtro_delegacion)
        ]

    if filtro_programa:
        df_filtrado = df_filtrado[
            df_filtrado["Programa"].isin(filtro_programa)
        ]

    if filtro_actividad:
        df_filtrado = df_filtrado[
            df_filtrado["Actividad"].isin(filtro_actividad)
        ]

    if filtro_estado:
        df_filtrado = df_filtrado[
            df_filtrado["Estado Verificación Regional"].isin(filtro_estado)
        ]

    if filtro_provincia:
        df_filtrado = df_filtrado[
            df_filtrado["Provincia"].isin(filtro_provincia)
        ]

    if fecha_inicio or fecha_fin:
        df_fecha = df_filtrado.copy()

        df_fecha["Fecha_Actividad_DT"] = pd.to_datetime(
            df_fecha["Fecha Actividad"],
            errors="coerce",
            dayfirst=True
        )

        if fecha_inicio:
            df_fecha = df_fecha[
                df_fecha["Fecha_Actividad_DT"].dt.date >= fecha_inicio
            ]

        if fecha_fin:
            df_fecha = df_fecha[
                df_fecha["Fecha_Actividad_DT"].dt.date <= fecha_fin
            ]

        df_filtrado = df_fecha.drop(columns=["Fecha_Actividad_DT"])

    st.session_state.filtros_admin_aplicados = {
        "Dirección Regional": filtro_region,
        "Delegación": filtro_delegacion,
        "Programa": filtro_programa,
        "Actividad": filtro_actividad,
        "Estado Verificación Regional": filtro_estado,
        "Provincia": filtro_provincia,
        "Fecha Inicio": str(fecha_inicio) if fecha_inicio else "",
        "Fecha Final": str(fecha_fin) if fecha_fin else ""
    }

    return df_filtrado


# ======================================================
# CREAR MAPA BASE ADMIN CON TIPO DE MAPA
# ======================================================

def crear_mapa_base_admin(centro, zoom, tipo_mapa):
    mapa = folium.Map(
        location=centro,
        zoom_start=zoom,
        tiles=None
    )

    if tipo_mapa == "OpenStreetMap":
        folium.TileLayer(
            "OpenStreetMap",
            name="OpenStreetMap"
        ).add_to(mapa)

    elif tipo_mapa == "Mapa claro":
        folium.TileLayer(
            "CartoDB positron",
            name="Mapa claro"
        ).add_to(mapa)

    elif tipo_mapa == "Mapa oscuro":
        folium.TileLayer(
            "CartoDB dark_matter",
            name="Mapa oscuro"
        ).add_to(mapa)

    elif tipo_mapa == "Topográfico":
        folium.TileLayer(
            "OpenTopoMap",
            name="Topográfico"
        ).add_to(mapa)

    elif tipo_mapa == "Satélite":
        folium.TileLayer(
            tiles=(
                "https://server.arcgisonline.com/ArcGIS/rest/services/"
                "World_Imagery/MapServer/tile/{z}/{y}/{x}"
            ),
            attr="Esri World Imagery",
            name="Satélite"
        ).add_to(mapa)

    folium.LayerControl().add_to(mapa)

    return mapa


# ======================================================
# DATAFRAME PARA MAPA
# ======================================================

def preparar_dataframe_mapa_admin(df):
    df_mapa = df.copy()

    if "Latitud" not in df_mapa.columns:
        df_mapa["Latitud"] = ""

    if "Longitud" not in df_mapa.columns:
        df_mapa["Longitud"] = ""

    df_mapa["Latitud_Num"] = df_mapa["Latitud"].apply(limpiar_coordenada)
    df_mapa["Longitud_Num"] = df_mapa["Longitud"].apply(limpiar_coordenada)

    df_mapa = df_mapa.dropna(
        subset=[
            "Latitud_Num",
            "Longitud_Num"
        ]
    )

    return df_mapa


# ======================================================
# MAPA GENERAL DE REGISTROS ADMIN
# ======================================================

def crear_mapa_admin(df, tipo_mapa="OpenStreetMap"):
    df_mapa = preparar_dataframe_mapa_admin(df)

    if df_mapa.empty:
        return crear_mapa_base_admin(
            centro=[9.7489, -83.7534],
            zoom=7,
            tipo_mapa=tipo_mapa
        )

    if len(df_mapa) == 1:
        centro = [
            float(df_mapa.iloc[0]["Latitud_Num"]),
            float(df_mapa.iloc[0]["Longitud_Num"])
        ]
        zoom = 16
    else:
        centro = [
            float(df_mapa["Latitud_Num"].mean()),
            float(df_mapa["Longitud_Num"].mean())
        ]
        zoom = 8

    mapa = crear_mapa_base_admin(
        centro=centro,
        zoom=zoom,
        tipo_mapa=tipo_mapa
    )

    for _, row in df_mapa.iterrows():
        estado = str(row.get("Estado Verificación Regional", "Pendiente de verificación"))
        programa = str(row.get("Programa", ""))

        color = obtener_color_programa(programa)

        popup_html = f"""
        <div style="font-family:Arial; width:300px;">
            <h4 style="color:#002B7F;">Registro PUMI #{row.get("ID", "")}</h4>
            <b>Estado:</b> {estado}<br>
            <b>Dirección Regional:</b> {row.get("Dirección Regional", "")}<br>
            <b>Delegación:</b> {row.get("Delegación", "")}<br>
            <b>Programa:</b> {row.get("Programa", "")}<br>
            <b>Actividad:</b> {row.get("Actividad", "")}<br>
            <b>Fecha:</b> {row.get("Fecha Actividad", "")}<br>
            <b>Lugar:</b> {row.get("Lugar", "")}<br>
            <b>Participantes:</b> {row.get("Cantidad Participantes", "")}<br>
            <b>Observación verificación:</b> {row.get("Observaciones Verificación Regional", "")}<br>
        </div>
        """

        folium.Marker(
            location=[
                float(row["Latitud_Num"]),
                float(row["Longitud_Num"])
            ],
            popup=folium.Popup(popup_html, max_width=350),
            tooltip=f'{row.get("Delegación", "")} - {estado}',
            icon=folium.Icon(
                color=color,
                icon="info-sign"
            )
        ).add_to(mapa)

    if len(df_mapa) > 1:
        bounds = [
            [
                df_mapa["Latitud_Num"].min(),
                df_mapa["Longitud_Num"].min()
            ],
            [
                df_mapa["Latitud_Num"].max(),
                df_mapa["Longitud_Num"].max()
            ]
        ]

        mapa.fit_bounds(bounds)

    return mapa


def mostrar_mapa_admin(df, key="mapa_admin"):
    st.markdown("### Mapa de actividades filtradas")

    tipo_mapa_admin = st.selectbox(
        "Tipo de mapa",
        [
            "OpenStreetMap",
            "Mapa claro",
            "Mapa oscuro",
            "Topográfico",
            "Satélite"
        ],
        key=f"tipo_mapa_{key}"
    )

    st.session_state["tipo_mapa_admin_actual"] = tipo_mapa_admin

    df_mapa = preparar_dataframe_mapa_admin(df)

    if df_mapa.empty:
        st.info("No hay registros con coordenadas para mostrar en el mapa.")

    mapa = crear_mapa_admin(
        df,
        tipo_mapa=tipo_mapa_admin
    )

    st_folium(
        mapa,
        height=560,
        use_container_width=True,
        key=key
    )


# ======================================================
# MÉTRICAS ADMIN
# ======================================================

def mostrar_metricas_admin(df):
    df_metricas = limpiar_dataframe_metricas_admin(df)

    total_registros = len(df_metricas)

    total_participantes = (
        int(df_metricas["Cantidad Participantes"].sum())
        if not df_metricas.empty
        else 0
    )

    total_aprobadas = (
        len(df_metricas[df_metricas["Estado Verificación Regional"] == "Verificada para envío"])
        if not df_metricas.empty
        else 0
    )

    total_pendientes = (
        len(df_metricas[df_metricas["Estado Verificación Regional"] == "Pendiente de verificación"])
        if not df_metricas.empty
        else 0
    )

    total_rechazadas = (
        len(df_metricas[df_metricas["Estado Verificación Regional"] == "Devuelta a delegación"])
        if not df_metricas.empty
        else 0
    )

    total_observadas = (
        len(df_metricas[df_metricas["Estado Verificación Regional"] == "Con observaciones regionales"])
        if not df_metricas.empty
        else 0
    )

    col1, col2, col3, col4, col5 = st.columns(5)

    with col1:
        st.metric("Registros", total_registros)

    with col2:
        st.metric("Participantes", total_participantes)

    with col3:
        st.metric("Aprobadas", total_aprobadas)

    with col4:
        st.metric("Pendientes", total_pendientes)

    with col5:
        st.metric("Con observaciones regionales", total_observadas)

    col6, col7 = st.columns(2)

    with col6:
        st.metric("Rechazadas", total_rechazadas)

    with col7:
        porcentaje_aprobacion = 0

        if total_registros > 0:
            porcentaje_aprobacion = (total_aprobadas / total_registros) * 100

        st.metric("Porcentaje aprobación", f"{porcentaje_aprobacion:.1f}%")


# ======================================================
# APLICAR DISEÑO INSTITUCIONAL A GRÁFICOS
# Evita gráficos negros y mantiene azul/dorado.
# ======================================================

def aplicar_estilo_grafico_institucional(fig):
    fig.update_layout(
        title_x=0.5,
        plot_bgcolor="white",
        paper_bgcolor="white",
        font=dict(
            color=COLOR_GRIS_OSCURO,
            size=14
        ),
        title_font=dict(
            color=COLOR_AZUL,
            size=20
        ),
        xaxis=dict(
            title_font=dict(color=COLOR_AZUL),
            tickfont=dict(color=COLOR_GRIS_OSCURO),
            gridcolor="#E6EAF2",
            zerolinecolor="#D7DCE5"
        ),
        yaxis=dict(
            title_font=dict(color=COLOR_AZUL),
            tickfont=dict(color=COLOR_GRIS_OSCURO),
            gridcolor="#E6EAF2",
            zerolinecolor="#D7DCE5"
        ),
        legend=dict(
            font=dict(color=COLOR_GRIS_OSCURO),
            bgcolor="rgba(255,255,255,0.85)"
        )
    )

    return fig


# ======================================================
# GRÁFICOS ADMIN
# ======================================================

def mostrar_graficos_admin(df):
    df_metricas = limpiar_dataframe_metricas_admin(df)

    if df_metricas.empty:
        st.info("No hay datos para graficar.")
        return

    st.markdown("### Gráficos de análisis")

    # ==================================================
    # GRÁFICO POR ESTADO DE VALIDACIÓN
    # ==================================================

    if "Estado Verificación Regional" in df_metricas.columns:
        conteo_estado = (
            df_metricas
            .groupby("Estado Verificación Regional")
            .size()
            .reset_index(name="Cantidad")
            .sort_values("Cantidad", ascending=False)
        )

        if not conteo_estado.empty:
            fig_estado = px.pie(
                conteo_estado,
                names="Estado Verificación Regional",
                values="Cantidad",
                title="Distribución por estado de verificación",
                hole=0.35,
                color="Estado Verificación Regional",
                color_discrete_map={
                    "Pendiente de verificación": COLOR_AMARILLO,
                    "Verificada para envío": COLOR_VERDE,
                    "Devuelta a delegación": COLOR_ROJO,
                    "Con observaciones regionales": COLOR_DORADO
                }
            )

            fig_estado.update_traces(
                textinfo="percent+label+value",
                marker=dict(
                    line=dict(
                        color="white",
                        width=2
                    )
                )
            )

            fig_estado = aplicar_estilo_grafico_institucional(fig_estado)

            st.plotly_chart(
                fig_estado,
                use_container_width=True
            )

    # ==================================================
    # GRÁFICO POR PROGRAMA
    # ==================================================

    if "Programa" in df_metricas.columns:
        conteo_programa = (
            df_metricas
            .groupby("Programa")
            .size()
            .reset_index(name="Cantidad")
            .sort_values("Cantidad", ascending=False)
        )

        if not conteo_programa.empty:
            fig_programa = px.bar(
                conteo_programa,
                x="Programa",
                y="Cantidad",
                title="Cantidad de registros por programa",
                text="Cantidad",
                color_discrete_sequence=[COLOR_DORADO]
            )

            fig_programa.update_traces(
                marker_line_color=COLOR_AZUL,
                marker_line_width=1.5,
                textposition="outside",
                textfont=dict(
                    color=COLOR_AZUL,
                    size=14
                )
            )

            fig_programa.update_yaxes(
                rangemode="tozero",
                dtick=1
            )

            fig_programa = aplicar_estilo_grafico_institucional(fig_programa)

            st.plotly_chart(
                fig_programa,
                use_container_width=True
            )

    # ==================================================
    # GRÁFICO POR DIRECCIÓN REGIONAL
    # ==================================================

    if "Dirección Regional" in df_metricas.columns:
        conteo_region = (
            df_metricas
            .groupby("Dirección Regional")
            .size()
            .reset_index(name="Cantidad")
            .sort_values("Cantidad", ascending=False)
        )

        if not conteo_region.empty:
            fig_region = px.bar(
                conteo_region,
                x="Dirección Regional",
                y="Cantidad",
                title="Cantidad de registros por Dirección Regional",
                text="Cantidad",
                color_discrete_sequence=[COLOR_AZUL]
            )

            fig_region.update_traces(
                marker_line_color=COLOR_DORADO,
                marker_line_width=1.5,
                textposition="outside",
                textfont=dict(
                    color=COLOR_AZUL,
                    size=14
                )
            )

            fig_region.update_yaxes(
                rangemode="tozero",
                dtick=1
            )

            fig_region.update_xaxes(
                tickangle=-30
            )

            fig_region = aplicar_estilo_grafico_institucional(fig_region)

            st.plotly_chart(
                fig_region,
                use_container_width=True
            )

    # ==================================================
    # GRÁFICO POR DELEGACIÓN
    # ==================================================

    if "Delegación" in df_metricas.columns:
        conteo_delegacion = (
            df_metricas
            .groupby("Delegación")
            .size()
            .reset_index(name="Cantidad")
            .sort_values("Cantidad", ascending=False)
            .head(15)
        )

        if not conteo_delegacion.empty:
            fig_delegacion = px.bar(
                conteo_delegacion,
                x="Delegación",
                y="Cantidad",
                title="Top delegaciones por cantidad de registros",
                text="Cantidad",
                color_discrete_sequence=[COLOR_DORADO]
            )

            fig_delegacion.update_traces(
                marker_line_color=COLOR_AZUL,
                marker_line_width=1.5,
                textposition="outside",
                textfont=dict(
                    color=COLOR_AZUL,
                    size=14
                )
            )

            fig_delegacion.update_yaxes(
                rangemode="tozero",
                dtick=1
            )

            fig_delegacion.update_xaxes(
                tickangle=-30
            )

            fig_delegacion = aplicar_estilo_grafico_institucional(fig_delegacion)

            st.plotly_chart(
                fig_delegacion,
                use_container_width=True
            )
# ======================================================
# appREGIONAL.py
# PARTE 3 DE 5 CORREGIDA
# MÓDULO DE VALIDACIÓN INDIVIDUAL DE ACTIVIDADES,
# ACTUALIZACIÓN DE ESTADOS Y DESCARGA DE EXCEL VALIDADO
# FECHA DE VALIDACIÓN SIN HORA
# ======================================================


# ======================================================
# ACTUALIZAR VERIFICACIÓN POR ÍNDICE ÚNICO
# ======================================================

def actualizar_validacion_registro(
    indice_registro,
    estado_validacion,
    observaciones_validacion,
    funcionario_verificador
):
    df = st.session_state.df_admin.copy()

    try:
        indice_registro = int(indice_registro)
    except Exception:
        return False

    if indice_registro not in df.index:
        return False

    df.loc[indice_registro, "Estado Verificación Regional"] = estado_validacion
    df.loc[indice_registro, "Observaciones Verificación Regional"] = observaciones_validacion
    df.loc[indice_registro, "Coordinador Regional"] = funcionario_verificador
    df.loc[indice_registro, "Fecha Verificación Regional"] = datetime.now().strftime("%d/%m/%Y")

    st.session_state.df_admin = df.reset_index(drop=True)

    return True


def etiqueta_registro_regional(df, indice):
    try:
        fila = df.loc[indice]
    except Exception:
        return str(indice)

    delegacion = str(fila.get("Delegación", "")).strip()
    id_registro = str(fila.get("ID", "")).strip()
    archivo = str(fila.get("Archivo Origen", "")).strip()
    actividad = str(fila.get("Actividad", "")).strip()

    if not delegacion:
        delegacion = "Delegación sin dato"

    if not id_registro:
        id_registro = "Sin ID"

    etiqueta = f"{delegacion} | ID {id_registro}"

    if archivo:
        etiqueta += f" | Archivo: {archivo}"

    if actividad:
        etiqueta += f" | {actividad}"

    return etiqueta


# ======================================================
# TARJETAS RESUMEN DE VALIDACIÓN
# ======================================================

def mostrar_resumen_validacion(df):
    if df.empty:
        return

    total = len(df)
    aprobadas = len(df[df["Estado Verificación Regional"] == "Verificada para envío"])
    rechazadas = len(df[df["Estado Verificación Regional"] == "Devuelta a delegación"])
    pendientes = len(df[df["Estado Verificación Regional"] == "Pendiente de verificación"])
    observadas = len(df[df["Estado Verificación Regional"] == "Con observaciones regionales"])

    col1, col2, col3, col4, col5 = st.columns(5)

    with col1:
        st.metric("Total filtrado", total)

    with col2:
        st.metric("Aprobadas", aprobadas)

    with col3:
        st.metric("Pendientes", pendientes)

    with col4:
        st.metric("Rechazadas", rechazadas)

    with col5:
        st.metric("Con observaciones regionales", observadas)


# ======================================================
# MOSTRAR BADGE DE ESTADO
# ======================================================

def mostrar_badge_estado(estado):
    color = COLORES_VALIDACION.get(estado, COLOR_AMARILLO)

    st.markdown(
        f"""
        <div style="
            background:{color};
            color:white;
            padding:9px 16px;
            border-radius:14px;
            text-align:center;
            font-weight:900;
            margin-bottom:12px;
            box-shadow:0px 4px 10px rgba(0,0,0,0.16);
        ">
            Estado actual: {estado}
        </div>
        """,
        unsafe_allow_html=True
    )


# ======================================================
# MOSTRAR TARJETA DEL REGISTRO
# ======================================================

def mostrar_tarjeta_registro_validacion(fila):
    st.markdown(
        """
        <div class="card-validacion">
            <div class="texto-admin">
                Detalle del registro seleccionado para revisión regional.
            </div>
        </div>
        """,
        unsafe_allow_html=True
    )

    col1, col2 = st.columns(2)

    with col1:
        st.markdown("#### Datos generales")
        st.write(f"**ID:** {fila.get('ID', '')}")
        st.write(f"**Fecha:** {fila.get('Fecha Actividad', '')}")
        st.write(f"**Hora:** {fila.get('Hora Actividad', '')}")
        st.write(f"**Dirección Regional:** {fila.get('Dirección Regional', '')}")
        st.write(f"**Delegación:** {fila.get('Delegación', '')}")
        st.write(f"**Responsable:** {fila.get('Responsable', '')}")
        st.write(f"**Usuario Registra:** {fila.get('Usuario Registra', '')}")

    with col2:
        st.markdown("#### Actividad")
        st.write(f"**Programa:** {fila.get('Programa', '')}")
        st.write(f"**Actividad:** {fila.get('Actividad', '')}")
        st.write(f"**Lugar:** {fila.get('Lugar', '')}")
        st.write(f"**Provincia:** {fila.get('Provincia', '')}")
        st.write(f"**Cantón:** {fila.get('Cantón', '')}")
        st.write(f"**Distrito:** {fila.get('Distrito', '')}")

    st.markdown("#### Participantes")

    colp1, colp2, colp3, colp4 = st.columns(4)

    with colp1:
        st.metric("Total", fila.get("Cantidad Participantes", 0))

    with colp2:
        st.metric("Hombres", fila.get("Cantidad Hombres", 0))

    with colp3:
        st.metric("Mujeres", fila.get("Cantidad Mujeres", 0))

    with colp4:
        st.metric("Referencia", fila.get("Número de Referencia", ""))

    colr1, colr2, colr3, colr4 = st.columns(4)

    with colr1:
        st.metric("Edad 10 a 18", fila.get("Edad 10 a 18", 0))

    with colr2:
        st.metric("Edad 19 a 30", fila.get("Edad 19 a 30", 0))

    with colr3:
        st.metric("Edad 31 a 45", fila.get("Edad 31 a 45", 0))

    with colr4:
        st.metric("Edad 46+", fila.get("Edad 46 en adelante", 0))

    st.markdown("#### Observaciones del registro original")

    observacion_original = str(fila.get("Observaciones", "")).strip()

    if observacion_original:
        st.info(observacion_original)
    else:
        st.info("Sin observaciones registradas.")

    st.markdown("#### Datos de referencia")

    colref1, colref2 = st.columns(2)

    with colref1:
        st.write(f"**Número de referencia:** {fila.get('Número de Referencia', '')}")
        st.write(f"**Expediente referencia:** {fila.get('Número de Expediente Referencia', '')}")

    with colref2:
        st.write(f"**Instituciones participantes:** {fila.get('Instituciones Participantes', '')}")
        st.write(f"**Dirección mapa:** {fila.get('Dirección Mapa', '')}")


# ======================================================
# MÓDULO PRINCIPAL DE VALIDACIÓN INDIVIDUAL
# ======================================================

def modulo_validacion_actividades(df_filtrado):
    st.markdown("## Verificación individual de actividades")

    if df_filtrado.empty:
        st.info("No hay registros disponibles con los filtros aplicados.")
        return

    mostrar_resumen_validacion(df_filtrado)

    st.markdown(
        """
        <div class="card-validacion">
            <div class="texto-admin">
                En este apartado puede revisar cada actividad registrada en PUMI
                y asignar su estado regional: aprobada, rechazada, pendiente
                o con observaciones.
            </div>
        </div>
        """,
        unsafe_allow_html=True
    )

    st.markdown("### Selección del registro")

    indices_filtrados = df_filtrado.index.tolist()

    indice_verificar = st.selectbox(
        "Seleccione el registro a verificar",
        indices_filtrados,
        format_func=lambda idx: etiqueta_registro_regional(df_filtrado, idx),
        key="indice_verificar_individual"
    )

    if indice_verificar not in df_filtrado.index:
        st.warning("No se encontró el registro seleccionado.")
        return

    fila = df_filtrado.loc[indice_verificar].to_dict()
    id_verificar = str(fila.get("ID", ""))

    estado_actual = fila.get("Estado Verificación Regional", "Pendiente de verificación")

    mostrar_badge_estado(estado_actual)

    with st.expander("Ver detalle completo del registro", expanded=True):
        mostrar_tarjeta_registro_validacion(fila)

        lat = str(fila.get("Latitud", "")).strip()
        lon = str(fila.get("Longitud", "")).strip()

        if lat and lon:
            st.markdown("#### Ubicación del registro")

            df_unico = pd.DataFrame([fila])

            mostrar_mapa_admin(
                df_unico,
                key=f"mapa_validacion_{id_verificar}"
            )
        else:
            st.info("Este registro no tiene coordenadas disponibles.")

    st.markdown("---")
    st.markdown("### Actualizar verificación del registro")

    estados = [
        "Pendiente de verificación",
        "Verificada para envío",
        "Devuelta a delegación",
        "Con observaciones regionales"
    ]

    indice_estado = 0

    if estado_actual in estados:
        indice_estado = estados.index(estado_actual)

    colv1, colv2 = st.columns(2)

    with colv1:
        nuevo_estado = st.selectbox(
            "Estado de verificación",
            estados,
            index=indice_estado,
            key="nuevo_estado_individual"
        )

    with colv2:
        funcionario_verificador = st.text_input(
            "Funcionario verificador",
            value=str(fila.get("Coordinador Regional", "")),
            key="funcionario_verificador_individual"
        )

    observaciones_validacion = st.text_area(
        "Observaciones de verificación",
        value=str(fila.get("Observaciones Verificación Regional", "")),
        key="observaciones_validacion_individual"
    )

    if nuevo_estado in ["Devuelta a delegación", "Con observaciones regionales"]:
        if observaciones_validacion.strip() == "":
            st.warning(
                "Para registros rechazados o con observaciones, se recomienda "
                "indicar claramente el motivo en el campo de observaciones."
            )

    if st.button("💾 Guardar verificación del registro"):
        if funcionario_verificador.strip() == "":
            st.warning("Debe indicar el funcionario verificador.")
        elif nuevo_estado in ["Devuelta a delegación", "Con observaciones regionales"] and observaciones_validacion.strip() == "":
            st.warning(
                "Debe agregar una observación cuando el estado sea Rechazada "
                "o Con observaciones."
            )
        else:
            actualizado = actualizar_validacion_registro(
                indice_registro=indice_verificar,
                estado_validacion=nuevo_estado,
                observaciones_validacion=observaciones_validacion,
                funcionario_verificador=funcionario_verificador
            )

            if actualizado:
                st.success("Verificación guardada correctamente.")
                st.rerun()
            else:
                st.error("No se pudo guardar la verificación.")




# ======================================================
# CATÁLOGOS BASE PARA EDICIÓN REGIONAL
# Estas funciones usan las mismas bases de la app donde
# las delegaciones registran: Datos Importantes.xlsx y
# BASE DE DATOS MEP 2025.xlsx.
# ======================================================

def obtener_columna_por_nombre(df, posibles_nombres):
    columnas = list(df.columns)
    for posible in posibles_nombres:
        posible_norm = normalizar_texto(posible)
        for col in columnas:
            if normalizar_texto(col) == posible_norm:
                return col
    return None


def ordenar_regiones_numericamente(lista_regiones):
    def extraer_numero_region(region):
        texto = str(region)
        numero = ""
        for caracter in texto:
            if caracter.isdigit():
                numero += caracter
            elif numero:
                break
        return int(numero) if numero else 999
    return sorted(lista_regiones, key=extraer_numero_region)


def separar_responde_a(valor):
    if pd.isna(valor):
        return []
    texto = str(valor).strip()
    if texto == "" or texto.lower() == "nan":
        return []
    return [parte.strip() for parte in texto.split("/") if parte.strip()]


@st.cache_data
def cargar_datos_importantes():
    columnas_base = [
        "Provincia", "Cantón", "Distrito", "Dirección Regional",
        "Delegación", "Responde a", "Actividad Realizada", "Programa"
    ]

    if not os.path.exists(ARCHIVO_DATOS_IMPORTANTES):
        return pd.DataFrame(columns=columnas_base)

    try:
        df_original = pd.read_excel(ARCHIVO_DATOS_IMPORTANTES)

        col_provincia = obtener_columna_por_nombre(df_original, ["Provincia"])
        col_canton = obtener_columna_por_nombre(df_original, ["Cantón", "Canton"])
        col_distrito = obtener_columna_por_nombre(df_original, ["Distrito", "Distritos"])
        col_region = obtener_columna_por_nombre(df_original, ["Dirección Regional", "Direccion Regional"])
        col_delegacion = obtener_columna_por_nombre(df_original, ["Delegación", "Delegacion"])
        col_responde_a = obtener_columna_por_nombre(df_original, ["Responde a", "Responde a:", "Responde", "Responda a", "Responda a:"])
        col_actividad = obtener_columna_por_nombre(df_original, ["Actividad Realizada", "Actividad"])
        col_programa = obtener_columna_por_nombre(df_original, ["Programa"])

        columnas_requeridas = [
            col_provincia, col_canton, col_distrito, col_region,
            col_delegacion, col_responde_a, col_actividad, col_programa
        ]

        if not all(columnas_requeridas):
            st.warning(
                "El archivo Datos Importantes.xlsx no tiene todas las columnas requeridas. "
                "Debe incluir: Provincia, Cantón, Distrito, Dirección Regional, Delegación, "
                "Responde a, Actividad Realizada y Programa."
            )
            return pd.DataFrame(columns=columnas_base)

        df = df_original[[
            col_provincia, col_canton, col_distrito, col_region,
            col_delegacion, col_responde_a, col_actividad, col_programa
        ]].copy()
        df.columns = columnas_base

        for col in columnas_base:
            df[col] = df[col].fillna("").astype(str).str.strip()

        df = df.replace("nan", "")
        df = df.drop_duplicates()

        df["Provincia_Normalizada"] = df["Provincia"].apply(normalizar_texto)
        df["Cantón_Normalizado"] = df["Cantón"].apply(normalizar_texto)
        df["Distrito_Normalizado"] = df["Distrito"].apply(normalizar_texto)
        df["Región_Normalizada"] = df["Dirección Regional"].apply(normalizar_texto)
        df["Delegación_Normalizada"] = df["Delegación"].apply(normalizar_texto)
        df["Responde_a_Normalizado"] = df["Responde a"].apply(normalizar_texto)
        df["Actividad_Normalizada"] = df["Actividad Realizada"].apply(normalizar_texto)
        df["Programa_Normalizado"] = df["Programa"].apply(normalizar_texto)

        filas_expandidas = []
        for _, fila in df.iterrows():
            opciones_responde = separar_responde_a(fila["Responde a"])
            if not opciones_responde:
                opciones_responde = [fila["Responde a"]]
            for opcion in opciones_responde:
                nueva_fila = fila.copy()
                nueva_fila["Responde a"] = opcion
                nueva_fila["Responde_a_Normalizado"] = normalizar_texto(opcion)
                filas_expandidas.append(nueva_fila)

        if filas_expandidas:
            df = pd.DataFrame(filas_expandidas).drop_duplicates()

        return df

    except Exception as e:
        st.error(f"Error leyendo {ARCHIVO_DATOS_IMPORTANTES}")
        st.exception(e)
        return pd.DataFrame(columns=columnas_base)


def agregar_valor_actual(opciones, valor_actual):
    opciones = [str(x).strip() for x in opciones if str(x).strip() and str(x).strip().lower() != "nan"]
    opciones = list(dict.fromkeys(opciones))
    valor_actual = str(valor_actual).strip()
    if valor_actual and valor_actual.lower() != "nan" and valor_actual not in opciones:
        opciones.insert(0, valor_actual)
    if not opciones:
        opciones = [valor_actual] if valor_actual else [""]
    return opciones


def obtener_regiones_datos(valor_actual=""):
    df = cargar_datos_importantes()
    if df.empty:
        return agregar_valor_actual(ordenar_regiones_numericamente(REGIONES_BASE), valor_actual)
    regiones = df["Dirección Regional"].dropna().unique().tolist()
    regiones = [x for x in regiones if str(x).strip()]
    return agregar_valor_actual(ordenar_regiones_numericamente(regiones), valor_actual)


def obtener_delegaciones_por_region(region, valor_actual=""):
    df = cargar_datos_importantes()
    if df.empty or not region:
        return agregar_valor_actual([], valor_actual)
    region_norm = normalizar_texto(region)
    delegaciones = df[df["Región_Normalizada"] == region_norm]["Delegación"].dropna().unique().tolist()
    return agregar_valor_actual(sorted(delegaciones), valor_actual)


def obtener_responde_a_datos(valor_actual=""):
    df = cargar_datos_importantes()
    if df.empty or "Responde a" not in df.columns:
        return agregar_valor_actual([], valor_actual)
    responde_a = df["Responde a"].dropna().unique().tolist()
    return agregar_valor_actual(sorted(responde_a), valor_actual)


def obtener_programas_por_responde_a(responde_a, valor_actual=""):
    df = cargar_datos_importantes()
    if df.empty or not responde_a:
        return agregar_valor_actual(PROGRAMAS_BASE, valor_actual)
    responde_norm = normalizar_texto(responde_a)
    programas = df[df["Responde_a_Normalizado"] == responde_norm]["Programa"].dropna().unique().tolist()

    programas_limpios = []
    for programa in programas:
        if normalizar_texto(programa) == "GREAT CAMP":
            if "GREAT" not in programas_limpios:
                programas_limpios.append("GREAT")
        elif programa not in programas_limpios:
            programas_limpios.append(programa)

    orden_oficial = [p for p in PROGRAMAS_BASE if p in programas_limpios]
    extras = sorted([p for p in programas_limpios if p not in orden_oficial])
    return agregar_valor_actual(orden_oficial + extras, valor_actual)


def obtener_actividades_por_responde_a_programa(responde_a, programa, valor_actual=""):
    df = cargar_datos_importantes()
    if df.empty or not responde_a or not programa:
        return agregar_valor_actual([], valor_actual)

    responde_norm = normalizar_texto(responde_a)
    programa_norm = normalizar_texto(programa)

    if programa_norm == "GREAT":
        actividades = df[
            (df["Responde_a_Normalizado"] == responde_norm) &
            (df["Programa_Normalizado"].isin(["GREAT", "GREAT CAMP"]))
        ]["Actividad Realizada"].dropna().unique().tolist()
    else:
        actividades = df[
            (df["Responde_a_Normalizado"] == responde_norm) &
            (df["Programa_Normalizado"] == programa_norm)
        ]["Actividad Realizada"].dropna().unique().tolist()

    return agregar_valor_actual(sorted(actividades), valor_actual)


def obtener_provincias_datos(valor_actual=""):
    df = cargar_datos_importantes()
    if df.empty:
        return agregar_valor_actual(PROVINCIAS_BASE, valor_actual)
    provincias = df["Provincia"].dropna().unique().tolist()
    return agregar_valor_actual(sorted(provincias), valor_actual)


def obtener_cantones_por_provincia(provincia, valor_actual=""):
    df = cargar_datos_importantes()
    if df.empty or not provincia:
        return agregar_valor_actual([], valor_actual)
    provincia_norm = normalizar_texto(provincia)
    cantones = df[df["Provincia_Normalizada"] == provincia_norm]["Cantón"].dropna().unique().tolist()
    return agregar_valor_actual(sorted(cantones), valor_actual)


def obtener_distritos_por_provincia_canton(provincia, canton, valor_actual=""):
    df = cargar_datos_importantes()
    if df.empty or not provincia or not canton:
        return agregar_valor_actual([], valor_actual)
    provincia_norm = normalizar_texto(provincia)
    canton_norm = normalizar_texto(canton)
    distritos = df[
        (df["Provincia_Normalizada"] == provincia_norm) &
        (df["Cantón_Normalizado"] == canton_norm)
    ]["Distrito"].dropna().unique().tolist()
    return agregar_valor_actual(sorted(distritos), valor_actual)


@st.cache_data
def cargar_base_mep():
    columnas_base = ["PROVINCIA", "NOMBRE", "CODIGO_PRESUPUESTARIO"]
    if not os.path.exists(ARCHIVO_MEP):
        return pd.DataFrame(columns=columnas_base)
    try:
        df_original = pd.read_excel(ARCHIVO_MEP)
        col_provincia = obtener_columna_por_nombre(df_original, ["PROVINCIA", "Provincia"])
        col_nombre = obtener_columna_por_nombre(df_original, ["NOMBRE", "Nombre", "CENTRO EDUCATIVO", "Centro Educativo", "INSTITUCION", "Institución"])
        col_codigo = obtener_columna_por_nombre(df_original, ["CODIGO PRESUPUESTARIO", "CÓDIGO PRESUPUESTARIO", "Codigo Presupuestario", "Código Presupuestario", "CODIGO", "CÓDIGO"])

        if not col_provincia or not col_nombre:
            return pd.DataFrame(columns=columnas_base)
        if not col_codigo:
            df_original["CODIGO_PRESUPUESTARIO_TMP"] = ""
            col_codigo = "CODIGO_PRESUPUESTARIO_TMP"

        df = df_original[[col_provincia, col_nombre, col_codigo]].copy()
        df.columns = columnas_base
        for col in columnas_base:
            df[col] = df[col].fillna("").astype(str).str.strip()

        df["PROVINCIA_NORMALIZADA"] = df["PROVINCIA"].apply(normalizar_texto)
        df["NOMBRE_NORMALIZADO"] = df["NOMBRE"].apply(normalizar_texto)
        df["CENTRO_MOSTRAR"] = df.apply(
            lambda row: f"{row['NOMBRE']} - Código: {row['CODIGO_PRESUPUESTARIO']}"
            if row["CODIGO_PRESUPUESTARIO"].strip() != "" else row["NOMBRE"],
            axis=1
        )
        return df.drop_duplicates(subset=["PROVINCIA_NORMALIZADA", "NOMBRE_NORMALIZADO", "CODIGO_PRESUPUESTARIO"])
    except Exception as e:
        st.error(f"Error leyendo {ARCHIVO_MEP}")
        st.exception(e)
        return pd.DataFrame(columns=columnas_base)


def obtener_centros_por_provincia(provincia, valor_actual=""):
    df = cargar_base_mep()
    if df.empty or not provincia:
        return agregar_valor_actual([], valor_actual)
    provincia_norm = normalizar_texto(provincia)
    centros = df[df["PROVINCIA_NORMALIZADA"] == provincia_norm]["CENTRO_MOSTRAR"].dropna().astype(str).drop_duplicates().sort_values().tolist()
    return agregar_valor_actual(centros, valor_actual)


def obtener_datos_centro_educativo(centro_mostrar):
    df = cargar_base_mep()
    if df.empty or not centro_mostrar:
        return centro_mostrar, ""
    fila = df[df["CENTRO_MOSTRAR"] == centro_mostrar]
    if fila.empty:
        return centro_mostrar, ""
    nombre = fila.iloc[0].get("NOMBRE", "")
    codigo = fila.iloc[0].get("CODIGO_PRESUPUESTARIO", "")
    return nombre, codigo


def selectbox_con_valor_actual(label, opciones, valor_actual, key):
    opciones = agregar_valor_actual(opciones, valor_actual)
    return st.selectbox(
        label,
        opciones,
        index=indice_opcion_seguro(opciones, valor_actual),
        key=key
    )

# ======================================================
# EDITAR Y ELIMINAR REGISTROS REGIONALES
# ======================================================

def eliminar_registro_regional(indice_registro):
    df = st.session_state.df_admin.copy()

    try:
        indice_registro = int(indice_registro)
    except Exception:
        return False

    if indice_registro not in df.index:
        return False

    df = df.drop(index=indice_registro).reset_index(drop=True)
    st.session_state.df_admin = preparar_dataframe_admin(df)
    return True


def actualizar_registro_regional_completo(indice_registro, nuevos_datos):
    df = st.session_state.df_admin.copy()

    try:
        indice_registro = int(indice_registro)
    except Exception:
        return False

    if indice_registro not in df.index:
        return False

    for col, valor in nuevos_datos.items():
        if col in df.columns:
            df.loc[indice_registro, col] = valor

    st.session_state.df_admin = preparar_dataframe_admin(df.reset_index(drop=True))
    return True


def convertir_entero_seguro(valor):
    try:
        return int(pd.to_numeric(valor, errors="coerce"))
    except Exception:
        return 0



def obtener_opciones_columna_regional(columna, df_base=None, filtros=None, valor_actual=""):
    """Devuelve opciones limpias para selectbox usando los datos cargados."""
    if df_base is None:
        df_base = st.session_state.df_admin.copy()
    else:
        df_base = df_base.copy()

    if filtros:
        for col_filtro, valor_filtro in filtros.items():
            if col_filtro in df_base.columns and str(valor_filtro).strip() != "":
                df_base = df_base[
                    df_base[col_filtro].astype(str).apply(normalizar_texto)
                    == normalizar_texto(valor_filtro)
                ]

    opciones = []
    if columna in df_base.columns:
        opciones = (
            df_base[columna]
            .replace("", pd.NA)
            .dropna()
            .astype(str)
            .map(lambda x: x.strip())
            .drop_duplicates()
            .tolist()
        )

    opciones = [x for x in opciones if x and x.lower() != "nan"]
    opciones = sorted(opciones)

    valor_actual = str(valor_actual).strip()
    if valor_actual and valor_actual.lower() != "nan" and valor_actual not in opciones:
        opciones.insert(0, valor_actual)

    if not opciones:
        opciones = [valor_actual] if valor_actual else [""]

    return opciones


def indice_opcion_seguro(opciones, valor_actual):
    valor_actual = str(valor_actual).strip()
    try:
        return opciones.index(valor_actual)
    except Exception:
        return 0


def selectbox_edicion_regional(label, columna, fila, key, df_base=None, filtros=None):
    valor_actual = str(fila.get(columna, "")).strip()
    opciones = obtener_opciones_columna_regional(
        columna=columna,
        df_base=df_base,
        filtros=filtros,
        valor_actual=valor_actual
    )
    return st.selectbox(
        label,
        opciones,
        index=indice_opcion_seguro(opciones, valor_actual),
        key=key
    )

def modulo_editar_eliminar_registros(df_filtrado):
    st.markdown("## Editar o eliminar registros")

    if df_filtrado.empty:
        st.info("No hay registros disponibles con los filtros aplicados.")
        return

    st.markdown(
        """
        <div class="card-validacion">
            <div class="texto-admin">
                En este apartado el coordinador regional puede corregir datos puntuales
                antes de enviar el Excel unificado al nivel nacional. También puede eliminar
                registros que no correspondan a la Dirección Regional o que estén duplicados.
            </div>
        </div>
        """,
        unsafe_allow_html=True
    )

    indices_filtrados = df_filtrado.index.tolist()

    indice_editar = st.selectbox(
        "Seleccione el registro a editar o eliminar",
        indices_filtrados,
        format_func=lambda idx: etiqueta_registro_regional(df_filtrado, idx),
        key="indice_editar_eliminar_regional"
    )

    if indice_editar not in df_filtrado.index:
        st.warning("No se encontró el registro seleccionado.")
        return

    fila = df_filtrado.loc[indice_editar].to_dict()

    mostrar_badge_estado(fila.get("Estado Verificación Regional", "Pendiente de verificación"))

    with st.expander("Ver detalle actual del registro", expanded=False):
        mostrar_tarjeta_registro_validacion(fila)

    st.markdown("### Editar registro")
    # IMPORTANTE:
    # No se usa st.form en esta sección porque dentro de un formulario Streamlit
    # no actualiza los selectbox dependientes hasta presionar guardar.
    # Al dejar los widgets fuera del form, cada cambio refresca la pantalla y
    # los catálogos se actualizan inmediatamente, igual que en la app de delegaciones.

    col1, col2 = st.columns(2)

    with col1:
        fecha_actividad = st.text_input(
            "Fecha Actividad",
            value=str(fila.get("Fecha Actividad", "")),
            key=f"edit_fecha_{indice_editar}"
        )

        hora_actividad = st.text_input(
            "Hora Actividad",
            value=str(fila.get("Hora Actividad", "")),
            key=f"edit_hora_{indice_editar}"
        )

        direccion_regional = selectbox_con_valor_actual(
            "Dirección Regional",
            obtener_regiones_datos(fila.get("Dirección Regional", "")),
            fila.get("Dirección Regional", ""),
            key=f"edit_region_{indice_editar}"
        )

        valor_delegacion_actual = (
            fila.get("Delegación", "")
            if normalizar_texto(direccion_regional) == normalizar_texto(fila.get("Dirección Regional", ""))
            else ""
        )

        delegacion = selectbox_con_valor_actual(
            "Delegación",
            obtener_delegaciones_por_region(direccion_regional, valor_delegacion_actual),
            valor_delegacion_actual,
            key=f"edit_delegacion_{indice_editar}"
        )

        responde_a = selectbox_con_valor_actual(
            "Responde a",
            obtener_responde_a_datos(fila.get("Responde a", "")),
            fila.get("Responde a", ""),
            key=f"edit_responde_{indice_editar}"
        )

        valor_programa_actual = (
            fila.get("Programa", "")
            if normalizar_texto(responde_a) == normalizar_texto(fila.get("Responde a", ""))
            else ""
        )

        programa = selectbox_con_valor_actual(
            "Programa",
            obtener_programas_por_responde_a(responde_a, valor_programa_actual),
            valor_programa_actual,
            key=f"edit_programa_{indice_editar}"
        )

        valor_actividad_actual = (
            fila.get("Actividad", "")
            if (
                normalizar_texto(responde_a) == normalizar_texto(fila.get("Responde a", "")) and
                normalizar_texto(programa) == normalizar_texto(fila.get("Programa", ""))
            )
            else ""
        )

        actividad = selectbox_con_valor_actual(
            "Actividad",
            obtener_actividades_por_responde_a_programa(responde_a, programa, valor_actividad_actual),
            valor_actividad_actual,
            key=f"edit_actividad_{indice_editar}"
        )

        provincia = selectbox_con_valor_actual(
            "Provincia",
            obtener_provincias_datos(fila.get("Provincia", "")),
            fila.get("Provincia", ""),
            key=f"edit_provincia_{indice_editar}"
        )

        valor_canton_actual = (
            fila.get("Cantón", "")
            if normalizar_texto(provincia) == normalizar_texto(fila.get("Provincia", ""))
            else ""
        )

        canton = selectbox_con_valor_actual(
            "Cantón",
            obtener_cantones_por_provincia(provincia, valor_canton_actual),
            valor_canton_actual,
            key=f"edit_canton_{indice_editar}"
        )

        valor_distrito_actual = (
            fila.get("Distrito", "")
            if (
                normalizar_texto(provincia) == normalizar_texto(fila.get("Provincia", "")) and
                normalizar_texto(canton) == normalizar_texto(fila.get("Cantón", ""))
            )
            else ""
        )

        distrito = selectbox_con_valor_actual(
            "Distrito",
            obtener_distritos_por_provincia_canton(provincia, canton, valor_distrito_actual),
            valor_distrito_actual,
            key=f"edit_distrito_{indice_editar}"
        )

        tipo_lugar_opciones = ["Centro educativo", "Otro lugar"]
        tipo_lugar_actual = str(fila.get("Tipo Lugar", "")).strip()
        if tipo_lugar_actual and tipo_lugar_actual not in tipo_lugar_opciones:
            tipo_lugar_opciones.insert(0, tipo_lugar_actual)

        tipo_lugar = st.selectbox(
            "Tipo Lugar",
            tipo_lugar_opciones,
            index=indice_opcion_seguro(tipo_lugar_opciones, tipo_lugar_actual),
            key=f"edit_tipo_lugar_{indice_editar}"
        )

        if tipo_lugar == "Centro educativo":
            valor_centro_actual = (
                fila.get("Centro Educativo", "")
                if normalizar_texto(provincia) == normalizar_texto(fila.get("Provincia", ""))
                else ""
            )

            centro_opciones = obtener_centros_por_provincia(provincia, valor_centro_actual)

            centro_mostrar = selectbox_con_valor_actual(
                "Centro educativo según base MEP 2025",
                centro_opciones,
                valor_centro_actual,
                key=f"edit_centro_mep_{indice_editar}"
            )

            centro_educativo, codigo_presupuestario_auto = obtener_datos_centro_educativo(centro_mostrar)
            codigo_actual = str(fila.get("Código Presupuestario", "")).strip()

            codigo_presupuestario = (
                codigo_presupuestario_auto
                if codigo_presupuestario_auto
                else codigo_actual
            )

            lugar = centro_educativo

            st.info(f"Centro educativo seleccionado: {centro_educativo}")
            st.info(f"Código presupuestario: {codigo_presupuestario if codigo_presupuestario else 'No disponible'}")
        else:
            lugar = st.text_input(
                "Lugar",
                value=str(fila.get("Lugar", "")),
                key=f"edit_lugar_{indice_editar}"
            )
            centro_educativo = ""
            codigo_presupuestario = ""

        # ==============================================
        # Corrección de ubicación en mapa para edición
        # ==============================================
        st.markdown("#### Corregir ubicación en mapa")

        lat_key = f"edit_latitud_mapa_regional_{indice_editar}"
        lon_key = f"edit_longitud_mapa_regional_{indice_editar}"

        if lat_key not in st.session_state:
            st.session_state[lat_key] = str(fila.get("Latitud", ""))
        if lon_key not in st.session_state:
            st.session_state[lon_key] = str(fila.get("Longitud", ""))

        tipo_mapa_edit = st.selectbox(
            "Tipo de mapa para corregir marca",
            ["OpenStreetMap", "Mapa claro", "Mapa oscuro", "Topográfico", "Satélite"],
            key=f"edit_tipo_mapa_regional_{indice_editar}"
        )

        lat_actual_edit = limpiar_coordenada(st.session_state.get(lat_key, ""))
        lon_actual_edit = limpiar_coordenada(st.session_state.get(lon_key, ""))

        if lat_actual_edit is not None and lon_actual_edit is not None:
            centro_edit = [lat_actual_edit, lon_actual_edit]
            zoom_edit = 16
        else:
            centro_edit = CENTROS_PROVINCIA.get(provincia, [9.7489, -83.7534])
            zoom_edit = 10

        mapa_edit = crear_mapa_base_admin(
            centro=centro_edit,
            zoom=zoom_edit,
            tipo_mapa=tipo_mapa_edit
        )

        if lat_actual_edit is not None and lon_actual_edit is not None:
            folium.Marker(
                location=[lat_actual_edit, lon_actual_edit],
                popup="Ubicación actual del registro",
                tooltip="Marca actual",
                icon=folium.Icon(color="red", icon="map-marker")
            ).add_to(mapa_edit)

            folium.Circle(
                location=[lat_actual_edit, lon_actual_edit],
                radius=180,
                color=COLOR_AZUL,
                fill=True,
                fill_color=COLOR_DORADO,
                fill_opacity=0.18,
                weight=2
            ).add_to(mapa_edit)

        resultado_mapa_edit = st_folium(
            mapa_edit,
            height=420,
            use_container_width=True,
            key=f"mapa_edicion_regional_{indice_editar}"
        )

        if resultado_mapa_edit and resultado_mapa_edit.get("last_clicked"):
            st.session_state[lat_key] = str(resultado_mapa_edit["last_clicked"]["lat"])
            st.session_state[lon_key] = str(resultado_mapa_edit["last_clicked"]["lng"])
            st.success("Marca actualizada en el mapa. Revise las coordenadas antes de guardar.")
            st.rerun()

        latitud = st.text_input(
            "Latitud",
            value=st.session_state.get(lat_key, ""),
            key=f"input_latitud_regional_{indice_editar}"
        )

        longitud = st.text_input(
            "Longitud",
            value=st.session_state.get(lon_key, ""),
            key=f"input_longitud_regional_{indice_editar}"
        )

        st.session_state[lat_key] = latitud
        st.session_state[lon_key] = longitud

    with col2:
        responsable = st.text_input(
            "Responsable",
            value=str(fila.get("Responsable", "")),
            key=f"edit_responsable_{indice_editar}"
        )

        cantidad_participantes = st.number_input(
            "Cantidad Participantes",
            min_value=0,
            step=1,
            value=convertir_entero_seguro(fila.get("Cantidad Participantes", 0)),
            key=f"edit_cantidad_{indice_editar}"
        )

        cantidad_hombres = st.number_input(
            "Cantidad Hombres",
            min_value=0,
            step=1,
            value=convertir_entero_seguro(fila.get("Cantidad Hombres", 0)),
            key=f"edit_hombres_{indice_editar}"
        )

        cantidad_mujeres = st.number_input(
            "Cantidad Mujeres",
            min_value=0,
            step=1,
            value=convertir_entero_seguro(fila.get("Cantidad Mujeres", 0)),
            key=f"edit_mujeres_{indice_editar}"
        )

        edad_10_18 = st.number_input(
            "Edad 10 a 18",
            min_value=0,
            step=1,
            value=convertir_entero_seguro(fila.get("Edad 10 a 18", 0)),
            key=f"edit_edad_10_18_{indice_editar}"
        )

        edad_19_30 = st.number_input(
            "Edad 19 a 30",
            min_value=0,
            step=1,
            value=convertir_entero_seguro(fila.get("Edad 19 a 30", 0)),
            key=f"edit_edad_19_30_{indice_editar}"
        )

        edad_31_45 = st.number_input(
            "Edad 31 a 45",
            min_value=0,
            step=1,
            value=convertir_entero_seguro(fila.get("Edad 31 a 45", 0)),
            key=f"edit_edad_31_45_{indice_editar}"
        )

        edad_46_mas = st.number_input(
            "Edad 46 en adelante",
            min_value=0,
            step=1,
            value=convertir_entero_seguro(fila.get("Edad 46 en adelante", 0)),
            key=f"edit_edad_46_{indice_editar}"
        )

        instituciones = st.text_area(
            "Instituciones Participantes",
            value=str(fila.get("Instituciones Participantes", "")),
            height=80,
            key=f"edit_instituciones_{indice_editar}"
        )

        numero_referencia = st.text_input(
            "Número de Referencia",
            value=str(fila.get("Número de Referencia", "")),
            key=f"edit_referencia_{indice_editar}"
        )

        numero_expediente = st.text_input(
            "Número de Expediente Referencia",
            value=str(fila.get("Número de Expediente Referencia", "")),
            key=f"edit_expediente_{indice_editar}"
        )

        observaciones = st.text_area(
            "Observaciones del registro",
            value=str(fila.get("Observaciones", "")),
            height=90,
            key=f"edit_observaciones_{indice_editar}"
        )

        usuario_registra = st.text_input(
            "Usuario Registra",
            value=str(fila.get("Usuario Registra", "")),
            key=f"edit_usuario_{indice_editar}"
        )

        archivo_origen = st.text_input(
            "Archivo Origen",
            value=str(fila.get("Archivo Origen", "")),
            disabled=True,
            key=f"edit_archivo_origen_{indice_editar}"
        )

    suma_sexo = cantidad_hombres + cantidad_mujeres
    suma_edades = edad_10_18 + edad_19_30 + edad_31_45 + edad_46_mas

    if cantidad_participantes > 0:
        if suma_sexo != cantidad_participantes:
            st.warning(
                f"La suma de hombres y mujeres ({suma_sexo}) no coincide con el total de participantes ({cantidad_participantes})."
            )
        if suma_edades != cantidad_participantes:
            st.warning(
                f"La suma de rangos de edad ({suma_edades}) no coincide con el total de participantes ({cantidad_participantes})."
            )

    st.markdown("#### Verificación regional")

    estados = [
        "Pendiente de verificación",
        "Verificada para envío",
        "Devuelta a delegación",
        "Con observaciones regionales"
    ]

    estado_actual = str(fila.get("Estado Verificación Regional", "Pendiente de verificación"))
    indice_estado = estados.index(estado_actual) if estado_actual in estados else 0

    colv1, colv2 = st.columns(2)

    with colv1:
        estado_verificacion = st.selectbox(
            "Estado Verificación Regional",
            estados,
            index=indice_estado,
            key=f"edit_estado_verificacion_{indice_editar}"
        )

        coordinador_regional = st.text_input(
            "Coordinador Regional",
            value=str(fila.get("Coordinador Regional", "")),
            key=f"edit_coordinador_{indice_editar}"
        )

    with colv2:
        fecha_verificacion = st.text_input(
            "Fecha Verificación Regional",
            value=str(fila.get("Fecha Verificación Regional", "")),
            key=f"edit_fecha_verificacion_{indice_editar}"
        )

        observaciones_verificacion = st.text_area(
            "Observaciones Verificación Regional",
            value=str(fila.get("Observaciones Verificación Regional", "")),
            height=95,
            key=f"edit_obs_verificacion_{indice_editar}"
        )

    if st.button("💾 Guardar cambios del registro", key=f"btn_guardar_edicion_{indice_editar}"):
        if cantidad_participantes > 0 and suma_sexo != cantidad_participantes:
            st.error("No se puede guardar: hombres + mujeres no coincide con el total de participantes.")
            return

        if cantidad_participantes > 0 and suma_edades != cantidad_participantes:
            st.error("No se puede guardar: los rangos de edad no coinciden con el total de participantes.")
            return

        nuevos_datos = {
            "Fecha Actividad": fecha_actividad,
            "Hora Actividad": hora_actividad,
            "Dirección Regional": direccion_regional,
            "Delegación": delegacion,
            "Responde a": responde_a,
            "Programa": programa,
            "Actividad": actividad,
            "Provincia": provincia,
            "Cantón": canton,
            "Distrito": distrito,
            "Tipo Lugar": tipo_lugar,
            "Lugar": lugar,
            "Centro Educativo": centro_educativo,
            "Código Presupuestario": codigo_presupuestario,
            "Dirección Mapa": "",
            "Latitud": latitud,
            "Longitud": longitud,
            "Responsable": responsable,
            "Cantidad Participantes": cantidad_participantes,
            "Cantidad Hombres": cantidad_hombres,
            "Cantidad Mujeres": cantidad_mujeres,
            "Edad 10 a 18": edad_10_18,
            "Edad 19 a 30": edad_19_30,
            "Edad 31 a 45": edad_31_45,
            "Edad 46 en adelante": edad_46_mas,
            "Instituciones Participantes": instituciones,
            "Número de Referencia": numero_referencia,
            "Número de Expediente Referencia": numero_expediente,
            "Observaciones": observaciones,
            "Usuario Registra": usuario_registra,
            "Estado Verificación Regional": estado_verificacion,
            "Observaciones Verificación Regional": observaciones_verificacion,
            "Coordinador Regional": coordinador_regional,
            "Fecha Verificación Regional": fecha_verificacion,
        }

        actualizado = actualizar_registro_regional_completo(indice_editar, nuevos_datos)

        if actualizado:
            st.success("Registro actualizado correctamente.")
            st.rerun()
        else:
            st.error("No se pudo actualizar el registro.")

    st.markdown("---")
    st.markdown("### Eliminar registro")

    st.warning(
        "Esta acción elimina el registro de la sesión regional actual. "
        "El cambio se reflejará en el Excel unificado que descargue después."
    )

    confirmar_eliminacion = st.checkbox(
        "Confirmo que deseo eliminar este registro",
        key=f"confirmar_eliminacion_regional_{indice_editar}"
    )

    if st.button("🗑️ Eliminar registro seleccionado", key=f"btn_eliminar_regional_{indice_editar}"):
        if not confirmar_eliminacion:
            st.warning("Debe marcar la confirmación antes de eliminar.")
        else:
            eliminado = eliminar_registro_regional(indice_editar)
            if eliminado:
                st.success("Registro eliminado correctamente.")
                st.rerun()
            else:
                st.error("No se pudo eliminar el registro.")

# ======================================================
# DESCARGA DE EXCEL ADMIN VALIDADO
# ======================================================

def boton_descargar_excel_admin(df, key="descarga_excel_admin"):
    if df is None or df.empty:
        st.info("No hay registros para descargar.")
        return

    nombre_archivo = generar_nombre_reporte(df, tipo="XLSX")

    st.download_button(
        "⬇️ Descargar Excel regional verificado",
        data=convertir_excel_admin(df),
        file_name=nombre_archivo,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=key
    )


# ======================================================
# TABLA RESUMIDA PARA REVISIÓN
# ======================================================

def mostrar_tabla_resumen_validacion(df):
    if df.empty:
        st.info("No hay registros para mostrar.")
        return

    columnas_resumen = [
        "Archivo Origen",
        "ID",
        "Fecha Actividad",
        "Dirección Regional",
        "Delegación",
        "Programa",
        "Actividad",
        "Cantidad Participantes",
        "Estado Verificación Regional",
        "Observaciones Verificación Regional",
        "Coordinador Regional",
        "Fecha Verificación Regional"
    ]

    columnas_existentes = [
        col for col in columnas_resumen
        if col in df.columns
    ]

    st.dataframe(
        df[columnas_existentes],
        use_container_width=True,
        hide_index=True
    )


# ======================================================
# TABLA DETALLADA ADMINISTRATIVA
# ======================================================

def mostrar_tabla_detallada_admin(df):
    if df.empty:
        st.info("No hay registros para mostrar.")
        return

    st.dataframe(
        df,
        use_container_width=True,
        hide_index=True
    )
# ======================================================
# appREGIONAL.py
# PARTE 4 DE 5
# MÓDULO DE INFORME ELIMINADO
# La app regional solo genera Excel unificado, protegido
# y con marca de agua.
# ======================================================



# ======================================================
# ACCESO REGIONAL DESDE EXCEL DE USUARIOS Y CLAVES
# ======================================================

MESES_OFICIALES = [
    "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
    "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"
]


def limpiar_numero_meta(valor):
    if pd.isna(valor) or valor == "":
        return 0
    if isinstance(valor, str):
        valor = valor.replace("%", "").replace(",", ".").strip()
    try:
        return float(valor)
    except Exception:
        return 0


def obtener_codigo_region_desde_texto(valor):
    texto = normalizar_texto(valor)
    match = re.search(r"\bDR\s*([0-9]+)\s*([A-Z])?\b", texto)
    if match:
        numero = match.group(1)
        letra = match.group(2) or ""
        return f"DR{numero}{letra}"

    match = re.search(r"DIRECCION REGIONAL\s*([0-9]+)", texto)
    if match:
        return f"DR{match.group(1)}"

    match = re.search(r"REGION\s*([0-9]+)", texto)
    if match:
        return f"DR{match.group(1)}"

    return ""


def simplificar_usuario_region(valor):
    texto = normalizar_texto(valor)
    texto = texto.replace("DIRECCION REGIONAL", "REGION")
    texto = texto.replace("SAN JOSE", "")
    texto = texto.replace("HUETAR", "")
    texto = re.sub(r"[^A-Z0-9 ]", " ", texto)
    texto = " ".join(texto.split())
    return texto


@st.cache_data(show_spinner=False)
def cargar_catalogo_usuarios_claves_regional():
    columnas_base = ["Regiones", "Claves.1"]
    if not os.path.exists(ARCHIVO_USUARIOS_CLAVES):
        return pd.DataFrame(columns=columnas_base)

    try:
        df = pd.read_excel(ARCHIVO_USUARIOS_CLAVES)
    except Exception:
        return pd.DataFrame(columns=columnas_base)

    for col in columnas_base:
        if col not in df.columns:
            df[col] = ""

    df = df[columnas_base].copy()
    df.columns = ["Region", "Clave"]

    for col in ["Region", "Clave"]:
        df[col] = df[col].fillna("").astype(str).str.strip()

    df = df[(df["Region"] != "") & (df["Clave"] != "")].copy()
    df["Region_Normalizada"] = df["Region"].apply(normalizar_texto)
    df["Region_Simple"] = df["Region"].apply(simplificar_usuario_region)
    df["Clave_Normalizada"] = df["Clave"].apply(normalizar_texto)
    df["Codigo Region"] = df["Clave"].apply(obtener_codigo_region_desde_texto)

    return df.reset_index(drop=True)


def validar_login_regional(usuario, clave):
    df = cargar_catalogo_usuarios_claves_regional()
    if df.empty:
        return None

    usuario_norm = normalizar_texto(usuario)
    usuario_simple = simplificar_usuario_region(usuario)
    clave_norm = normalizar_texto(clave)

    coincidencia = df[
        (
            (df["Region_Normalizada"] == usuario_norm) |
            (df["Region_Simple"] == usuario_simple)
        ) &
        (df["Clave_Normalizada"] == clave_norm)
    ]

    if coincidencia.empty:
        return None

    fila = coincidencia.iloc[0].to_dict()
    return {
        "region": fila.get("Region", ""),
        "clave": fila.get("Clave", ""),
        "codigo": fila.get("Codigo Region", "")
    }


def inicializar_session_state_regional_login():
    if "regional_autenticado" not in st.session_state:
        st.session_state.regional_autenticado = False
    if "regional_region" not in st.session_state:
        st.session_state.regional_region = ""
    if "regional_codigo" not in st.session_state:
        st.session_state.regional_codigo = ""
    if "regional_usuario" not in st.session_state:
        st.session_state.regional_usuario = ""
    if "regional_menu" not in st.session_state:
        st.session_state.regional_menu = "Inicio"


inicializar_session_state_regional_login()


def mostrar_login_regional():
    mostrar_encabezado_institucional()
    mostrar_titulo_admin()

    st.markdown(
        """
        <div style="max-width:650px;margin:0 auto;text-align:center;">
            <h1 style="text-align:center;">Ingreso regional</h1>
        </div>
        """,
        unsafe_allow_html=True
    )

    col_izq, col_centro, col_der = st.columns([1, 1.2, 1])

    with col_centro:
        usuario = st.text_input("Usuario", key="login_regional_usuario")
        clave = st.text_input("Clave", type="password", key="login_regional_clave")

        if st.button("Ingresar", key="btn_login_regional"):
            datos = validar_login_regional(usuario, clave)
            if datos:
                st.session_state.regional_autenticado = True
                st.session_state.regional_region = datos["region"]
                st.session_state.regional_codigo = datos["codigo"]
                st.session_state.regional_usuario = usuario.strip()
                st.success("Acceso regional correcto.")
                st.rerun()
            else:
                st.error("Usuario o clave incorrectos.")


def region_usuario_actual():
    return st.session_state.get("regional_region", "")


def codigo_region_actual():
    return st.session_state.get("regional_codigo", "")


def normalizar_codigo_region_base(valor):
    """
    Convierte variantes de región a un código base comparable.

    Ejemplos:
    - DR1, DR1N, DR1 NORTE -> DR1
    - Dirección Regional 1 - San José Norte -> DR1
    - Región 12 - Caribe -> DR12

    Se usa código base para evitar que un Excel con DR1 sea rechazado
    cuando el usuario ingresó como Dirección Regional 1 - San José Norte.
    """
    texto = normalizar_texto(valor)
    if not texto:
        return ""

    match = re.search(r"\bDR\s*([0-9]+)", texto)
    if match:
        return f"DR{int(match.group(1))}"

    match = re.search(r"DIRECCION REGIONAL\s*([0-9]+)", texto)
    if match:
        return f"DR{int(match.group(1))}"

    match = re.search(r"REGION\s*([0-9]+)", texto)
    if match:
        return f"DR{int(match.group(1))}"

    return texto


def regiones_compatibles(valor_excel, region_usuario, codigo_usuario=""):
    """
    Valida si el valor regional del Excel pertenece a la región autenticada.
    Acepta códigos cortos y nombres completos sin exigir coincidencia literal.
    """
    valor_excel_norm = normalizar_texto(valor_excel)
    region_usuario_norm = normalizar_texto(region_usuario)
    codigo_usuario_norm = normalizar_texto(codigo_usuario)

    if not valor_excel_norm:
        return False

    # Coincidencia literal cuando ambos vienen con el mismo texto.
    if valor_excel_norm == region_usuario_norm:
        return True

    # Coincidencia por código base: DR1, DR2, DR12, etc.
    codigo_excel_base = normalizar_codigo_region_base(valor_excel)
    codigo_region_base = normalizar_codigo_region_base(region_usuario)
    codigo_usuario_base = normalizar_codigo_region_base(codigo_usuario)

    codigos_validos_usuario = {
        c for c in [codigo_region_base, codigo_usuario_base] if c
    }

    if codigo_excel_base and codigo_excel_base in codigos_validos_usuario:
        return True

    # Respaldo: si el código de usuario completo aparece en el texto del Excel
    # o viceversa, también se acepta.
    if codigo_usuario_norm and (codigo_usuario_norm in valor_excel_norm or valor_excel_norm in codigo_usuario_norm):
        return True

    return False


def filtrar_df_por_region_autenticada(df):
    if df is None or df.empty:
        return df

    region = region_usuario_actual()
    codigo = codigo_region_actual()

    if not region and not codigo:
        return df

    df = df.copy()

    if "Dirección Regional" not in df.columns:
        return df.iloc[0:0]

    mascara_region = df["Dirección Regional"].astype(str).apply(
        lambda valor: regiones_compatibles(
            valor_excel=valor,
            region_usuario=region,
            codigo_usuario=codigo
        )
    )

    return df[mascara_region].copy()


def mapa_codigos_a_regiones():
    df = cargar_catalogo_usuarios_claves_regional()
    mapa = {}
    if df.empty:
        return mapa
    for _, row in df.iterrows():
        codigo = str(row.get("Codigo Region", "")).strip()
        region = str(row.get("Region", "")).strip()
        if codigo and region:
            mapa[codigo] = region
    return mapa


def obtener_region_desde_archivo_meta(ruta_archivo):
    nombre = os.path.basename(ruta_archivo)
    nombre_sin_ext = os.path.splitext(nombre)[0]
    codigo = obtener_codigo_region_desde_texto(nombre_sin_ext)
    mapa = mapa_codigos_a_regiones()

    if codigo and codigo in mapa:
        return mapa[codigo]

    # Casos de archivos como DR1 C, DR1 N, DR1 S.
    texto_nombre = normalizar_texto(nombre_sin_ext)
    for cod, region in mapa.items():
        if normalizar_texto(cod) in texto_nombre.replace(" ", ""):
            return region

    return codigo if codigo else nombre_sin_ext


def buscar_archivos_metas_regionales():
    rutas = []
    patrones = [
        os.path.join(CARPETA_METAS_REGIONALES, "DR*.xlsx"),
        os.path.join(CARPETA_METAS_REGIONALES, "DR*.xlsm"),
        os.path.join(CARPETA_METAS_REGIONALES, "DR*.xls"),
    ]

    if os.path.isdir(CARPETA_METAS_ALTERNATIVA):
        patrones.extend([
            os.path.join(CARPETA_METAS_ALTERNATIVA, "DR*.xlsx"),
            os.path.join(CARPETA_METAS_ALTERNATIVA, "DR*.xlsm"),
            os.path.join(CARPETA_METAS_ALTERNATIVA, "DR*.xls"),
        ])

    for patron in patrones:
        rutas.extend(glob.glob(patron))

    rutas_limpias = []
    for ruta in rutas:
        nombre = os.path.basename(ruta).upper()
        if nombre.startswith("~$"):
            continue
        if "REGISTRO_PUMI" in nombre or "INFORME_VALIDACION" in nombre:
            continue
        if ruta not in rutas_limpias:
            rutas_limpias.append(ruta)

    return sorted(rutas_limpias)


@st.cache_data(show_spinner=False)
def cargar_metas_regionales():
    registros = []
    archivos = buscar_archivos_metas_regionales()

    for ruta in archivos:
        try:
            xl = pd.ExcelFile(ruta)
        except Exception:
            continue

        region_archivo = obtener_region_desde_archivo_meta(ruta)

        for hoja in xl.sheet_names:
            if str(hoja).strip().upper().startswith("TOTAL"):
                continue

            try:
                raw = pd.read_excel(ruta, sheet_name=hoja, header=None)
            except Exception:
                continue

            if raw.empty or raw.shape[1] < 4:
                continue

            df = raw.iloc[2:].copy()
            if df.empty:
                continue

            df = df.rename(columns={0: "Programa", 1: "Actividad", 2: "Meta 2026", 3: "Avance base"})

            for idx_mes, mes in enumerate(MESES_OFICIALES, start=4):
                if idx_mes in df.columns:
                    df[mes.title()] = df[idx_mes]
                else:
                    df[mes.title()] = 0

            df["Programa"] = df["Programa"].ffill()

            for _, fila in df.iterrows():
                programa = str(fila.get("Programa", "")).strip()
                actividad = str(fila.get("Actividad", "")).strip()

                if not programa or programa.lower() == "nan":
                    continue
                if not actividad or actividad.lower() == "nan":
                    continue

                meta = limpiar_numero_meta(fila.get("Meta 2026", 0))
                avance_base = limpiar_numero_meta(fila.get("Avance base", 0))

                if meta == 0 and avance_base == 0:
                    meses_sum = sum(limpiar_numero_meta(fila.get(mes.title(), 0)) for mes in MESES_OFICIALES)
                    if meses_sum == 0:
                        continue

                item = {
                    "Archivo meta": os.path.basename(ruta),
                    "Dirección Regional": region_archivo,
                    "Delegación": str(hoja).strip(),
                    "Programa": programa,
                    "Actividad": actividad,
                    "Meta 2026": meta,
                    "Avance base": avance_base,
                    "Clave Regional": normalizar_texto(region_archivo),
                    "Clave Delegación": normalizar_texto(hoja),
                    "Clave Programa": normalizar_texto(programa),
                    "Clave Actividad": normalizar_texto(actividad),
                }

                for mes in MESES_OFICIALES:
                    item[mes.title()] = limpiar_numero_meta(fila.get(mes.title(), 0))

                registros.append(item)

    if not registros:
        return pd.DataFrame(columns=[
            "Archivo meta", "Dirección Regional", "Delegación", "Programa", "Actividad",
            "Meta 2026", "Avance base", "Clave Regional", "Clave Delegación",
            "Clave Programa", "Clave Actividad"
        ])

    df_metas = pd.DataFrame(registros).drop_duplicates(
        subset=["Archivo meta", "Delegación", "Programa", "Actividad"],
        keep="first"
    ).reset_index(drop=True)

    return df_metas


def filtrar_metas_region_actual(df_metas=None):
    if df_metas is None:
        df_metas = cargar_metas_regionales()
    if df_metas.empty:
        return df_metas

    region = region_usuario_actual()
    if not region:
        return df_metas

    region_norm = normalizar_texto(region)
    return df_metas[df_metas["Clave Regional"] == region_norm].copy()


def clasificar_avance_meta(porc):
    if porc >= 1:
        return "Completa"
    if porc > 0:
        return "En avance"
    return "Pendiente"


def generar_avance_contra_metas_regional(df_registros):
    metas = filtrar_metas_region_actual().copy()

    if metas.empty:
        return pd.DataFrame()

    if df_registros is None or df_registros.empty:
        metas["Registrado PUMI"] = 0
    else:
        registros = filtrar_df_por_region_autenticada(df_registros).copy()

        for col in ["Delegación", "Programa", "Actividad"]:
            if col not in registros.columns:
                registros[col] = ""

        registros["Clave Delegación"] = registros["Delegación"].apply(normalizar_texto)
        registros["Clave Programa"] = registros["Programa"].apply(normalizar_texto)
        registros["Clave Actividad"] = registros["Actividad"].apply(normalizar_texto)

        if "Avance Realizado" not in registros.columns:
            registros["Avance Realizado"] = 1

        registros["Avance Realizado"] = pd.to_numeric(
            registros["Avance Realizado"],
            errors="coerce"
        ).fillna(0)

        conteo = (
            registros
            .groupby(["Clave Delegación", "Clave Programa", "Clave Actividad"], as_index=False)["Avance Realizado"]
            .sum()
            .rename(columns={"Avance Realizado": "Registrado PUMI"})
        )

        metas = metas.merge(
            conteo,
            on=["Clave Delegación", "Clave Programa", "Clave Actividad"],
            how="left"
        )
        metas["Registrado PUMI"] = metas["Registrado PUMI"].fillna(0)

    metas["Avance acumulado"] = metas["Avance base"] + metas["Registrado PUMI"]
    metas["Pendiente"] = (metas["Meta 2026"] - metas["Avance acumulado"]).clip(lower=0)
    metas["% Cumplimiento"] = metas.apply(
        lambda x: x["Avance acumulado"] / x["Meta 2026"] if x["Meta 2026"] else 0,
        axis=1
    )
    metas["Estado cumplimiento"] = metas["% Cumplimiento"].apply(clasificar_avance_meta)

    columnas = [
        "Dirección Regional", "Delegación", "Programa", "Actividad", "Meta 2026",
        "Avance base", "Registrado PUMI", "Avance acumulado", "Pendiente",
        "% Cumplimiento", "Estado cumplimiento", "Archivo meta"
    ]

    return metas[columnas].copy()


def mostrar_dashboard_avances_regional(df_registros):
    st.markdown("## Dashboard regional de metas y avances")

    df_avance = generar_avance_contra_metas_regional(df_registros)

    if df_avance.empty:
        st.info("No se encontraron metas oficiales para esta región.")
        return

    st.markdown("### Filtros del avance")

    col1, col2, col3 = st.columns(3)

    with col1:
        delegaciones = sorted(df_avance["Delegación"].dropna().astype(str).unique().tolist())
        filtro_delegacion = st.multiselect("Delegación", delegaciones, key="dash_regional_delegacion")

    with col2:
        programas = sorted(df_avance["Programa"].dropna().astype(str).unique().tolist())
        filtro_programa = st.multiselect("Programa", programas, key="dash_regional_programa")

    with col3:
        filtro_estado = st.multiselect(
            "Estado",
            ["Completa", "En avance", "Pendiente"],
            key="dash_regional_estado"
        )

    df_filtrado = df_avance.copy()

    if filtro_delegacion:
        df_filtrado = df_filtrado[df_filtrado["Delegación"].isin(filtro_delegacion)]
    if filtro_programa:
        df_filtrado = df_filtrado[df_filtrado["Programa"].isin(filtro_programa)]
    if filtro_estado:
        df_filtrado = df_filtrado[df_filtrado["Estado cumplimiento"].isin(filtro_estado)]

    total_meta = df_filtrado["Meta 2026"].sum()
    total_base = df_filtrado["Avance base"].sum()
    total_pumi = df_filtrado["Registrado PUMI"].sum()
    total_acum = df_filtrado["Avance acumulado"].sum()
    total_pend = df_filtrado["Pendiente"].sum()
    porc = total_acum / total_meta if total_meta else 0

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Meta oficial", f"{total_meta:,.0f}")
    c2.metric("Avance base", f"{total_base:,.0f}")
    c3.metric("Registrado PUMI", f"{total_pumi:,.0f}")
    c4.metric("Pendiente", f"{total_pend:,.0f}")
    c5.metric("% avance", f"{porc:.1%}")

    # Gráfico de cumplimiento por estado eliminado por solicitud del usuario.

    st.markdown("### Avance por delegación")

    resumen_delegacion = (
        df_filtrado
        .groupby("Delegación", as_index=False)[["Meta 2026", "Avance acumulado", "Pendiente"]]
        .sum()
    )

    if not resumen_delegacion.empty:
        grafico_delegacion = resumen_delegacion.melt(
            id_vars="Delegación",
            value_vars=["Meta 2026", "Avance acumulado", "Pendiente"],
            var_name="Indicador",
            value_name="Cantidad"
        )
        fig_delegacion = px.bar(
            grafico_delegacion,
            x="Delegación",
            y="Cantidad",
            color="Indicador",
            barmode="group",
            title="Meta, avance y pendiente por delegación",
            text="Cantidad"
        )
        fig_delegacion.update_xaxes(tickangle=-30)
        fig_delegacion = aplicar_estilo_grafico_institucional(fig_delegacion)
        st.plotly_chart(fig_delegacion, use_container_width=True)

    st.markdown("### Detalle de metas y avances")

    vista = df_filtrado.copy()
    vista["% Cumplimiento"] = vista["% Cumplimiento"].apply(lambda x: f"{x:.1%}")
    for col in ["Meta 2026", "Avance base", "Registrado PUMI", "Avance acumulado", "Pendiente"]:
        vista[col] = vista[col].apply(lambda x: f"{x:,.0f}")

    st.dataframe(
        vista,
        use_container_width=True,
        hide_index=True
    )

    st.markdown("### Descargar avance regional")
    salida_avance = BytesIO()
    with pd.ExcelWriter(salida_avance, engine="openpyxl") as writer:
        df_filtrado.to_excel(writer, index=False, sheet_name="AVANCE_REGIONAL")
    salida_avance.seek(0)
    st.download_button(
        "⬇️ Descargar avance regional filtrado",
        data=salida_avance.getvalue(),
        file_name=f"AVANCE_REGIONAL_PUMI_2026_{limpiar_nombre_archivo(region_usuario_actual())}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="descarga_avance_regional_dashboard"
    )


def mostrar_inicio_regional():
    st.markdown(
        f"""
        <div class="card-admin">
            <div class="subtitulo-admin">Panel Regional PUMI 2026</div>
            <div class="texto-admin" style="text-align:center;">
                Región activa: <b>{region_usuario_actual()}</b><br>
                Desde este panel puede cargar los Excel remitidos por las delegaciones de su región,
                revisar registros, validar actividades, editar o eliminar datos cuando corresponda,
                consultar avances contra metas oficiales y descargar el Excel regional verificado.
            </div>
        </div>
        """,
        unsafe_allow_html=True
    )

    st.markdown("### Accesos rápidos")
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        if st.button("🔎 Consulta y filtros", key="inicio_ir_consulta"):
            st.session_state.regional_menu = "Consulta y filtros"
            st.rerun()
    with col2:
        if st.button("✅ Verificación", key="inicio_ir_validacion"):
            st.session_state.regional_menu = "Verificación de actividades"
            st.rerun()
    with col3:
        if st.button("✏️ Editar o eliminar", key="inicio_ir_editar"):
            st.session_state.regional_menu = "Editar o eliminar registros"
            st.rerun()
    with col4:
        if st.button("📊 Dashboard", key="inicio_ir_dashboard"):
            st.session_state.regional_menu = "Dashboard"
            st.rerun()

    df_admin_region = filtrar_df_por_region_autenticada(st.session_state.df_admin)
    if df_admin_region.empty:
        st.info("Cargue uno o varios Excel de las delegaciones desde el panel lateral.")
    else:
        mostrar_metricas_admin(df_admin_region)
        st.markdown("### Vista rápida de registros cargados")
        mostrar_tabla_resumen_validacion(df_admin_region)
        st.markdown("### Descargar Excel regional")
        boton_descargar_excel_admin(df_admin_region, key="descarga_inicio_regional_login")


# ======================================================
# FLUJO PRINCIPAL DE LA APP REGIONAL CON ACCESO POR REGIÓN
# ======================================================

if not st.session_state.regional_autenticado:
    mostrar_login_regional()
    st.stop()

# ======================================================
# SIDEBAR
# ======================================================

mostrar_logo_sidebar()

st.sidebar.markdown("## Panel Regional PUMI 2026")
st.sidebar.success(f"Región: {region_usuario_actual()}")

if st.sidebar.button("Cerrar sesión", key="cerrar_sesion_regional"):
    st.session_state.regional_autenticado = False
    st.session_state.regional_region = ""
    st.session_state.regional_codigo = ""
    st.session_state.regional_usuario = ""
    st.session_state.df_admin = pd.DataFrame(columns=ENCABEZADOS_COMPLETOS)
    st.session_state.archivo_admin_nombre = ""
    st.rerun()

st.sidebar.markdown(
    """
    <div style="
        background:#FFFFFF;
        padding:14px;
        border-radius:14px;
        border-left:6px solid #B88A2A;
        box-shadow:0px 4px 12px rgba(0,0,0,0.10);
        margin-bottom:15px;
        font-size:14px;
    ">
    Cargue únicamente los Excel generados por las delegaciones de su región.
    </div>
    """,
    unsafe_allow_html=True
)

# ======================================================
# CARGA DEL EXCEL PUMI
# ======================================================

archivos_admin = st.sidebar.file_uploader(
    "Subir uno o varios Excel generados por las delegaciones",
    type=["xlsx"],
    accept_multiple_files=True
)

if archivos_admin:
    nombres_archivos = " | ".join([archivo.name for archivo in archivos_admin])

    if st.session_state.archivo_admin_nombre != nombres_archivos:
        dataframes_cargados = []
        archivos_con_error = []
        registros_excluidos = 0

        for archivo_excel in archivos_admin:
            df_cargado = cargar_excel_pumi_admin(archivo_excel)

            if not df_cargado.empty:
                df_cargado["Archivo Origen"] = archivo_excel.name
                total_antes = len(df_cargado)
                df_cargado = filtrar_df_por_region_autenticada(df_cargado)
                registros_excluidos += max(total_antes - len(df_cargado), 0)

                if not df_cargado.empty:
                    dataframes_cargados.append(df_cargado)
            else:
                archivos_con_error.append(archivo_excel.name)

        if dataframes_cargados:
            df_unificado = pd.concat(dataframes_cargados, ignore_index=True)
            df_unificado = preparar_dataframe_admin(df_unificado)

            st.session_state.df_admin = df_unificado
            st.session_state.archivo_admin_nombre = nombres_archivos
            st.sidebar.success(f"{len(dataframes_cargados)} Excel cargado(s) correctamente.")

            if registros_excluidos:
                st.sidebar.warning(f"Se excluyeron {registros_excluidos} registro(s) que no pertenecen a esta región.")

            if archivos_con_error:
                st.sidebar.warning("No se pudieron cargar estos archivos: " + ", ".join(archivos_con_error))
        else:
            st.sidebar.warning("Ningún Excel contiene registros válidos para esta región.")

# ======================================================
# MENÚ PRINCIPAL
# ======================================================

opciones_menu = [
    "Inicio",
    "Consulta y filtros",
    "Verificación de actividades",
    "Editar o eliminar registros",
    "Dashboard"
]

if st.session_state.regional_menu not in opciones_menu:
    st.session_state.regional_menu = "Inicio"

menu_admin = st.sidebar.radio(
    "Menú regional",
    opciones_menu,
    index=opciones_menu.index(st.session_state.regional_menu),
    key="radio_menu_regional"
)

st.session_state.regional_menu = menu_admin

# ======================================================
# ENCABEZADO GENERAL
# ======================================================

mostrar_encabezado_institucional()
mostrar_titulo_admin()

# ======================================================
# DATAFRAME BASE FILTRADO POR REGIÓN AUTENTICADA
# ======================================================

df_admin = filtrar_df_por_region_autenticada(st.session_state.df_admin.copy())

# ======================================================
# INICIO
# ======================================================

if menu_admin == "Inicio":
    mostrar_inicio_regional()

# ======================================================
# CONSULTA Y FILTROS
# ======================================================

elif menu_admin == "Consulta y filtros":

    st.markdown("## Consulta general de registros")

    if df_admin.empty:
        st.info("Debe cargar primero el Excel generado por PUMI para esta región.")
    else:
        df_filtrado = aplicar_filtros_admin(df_admin)

        st.markdown("---")
        mostrar_resumen_validacion(df_filtrado)

        st.markdown("### Registros filtrados")
        mostrar_tabla_resumen_validacion(df_filtrado)

        st.markdown("### Descargar Excel filtrado verificado")
        boton_descargar_excel_admin(df_filtrado, key="descarga_consulta_filtrada")

# ======================================================
# VALIDACIÓN DE ACTIVIDADES
# ======================================================

elif menu_admin == "Verificación de actividades":

    st.markdown("## Módulo de verificación regional")

    if df_admin.empty:
        st.info("Debe cargar primero el Excel generado por PUMI para esta región.")
    else:
        df_filtrado = aplicar_filtros_admin(df_admin)
        st.markdown("---")
        modulo_validacion_actividades(df_filtrado)

        st.markdown("---")
        st.markdown("### Descargar Excel con verificaciones")
        boton_descargar_excel_admin(filtrar_df_por_region_autenticada(st.session_state.df_admin), key="descarga_validacion_admin")

# ======================================================
# EDITAR O ELIMINAR REGISTROS
# ======================================================

elif menu_admin == "Editar o eliminar registros":

    st.markdown("## Módulo regional de edición y eliminación")

    if df_admin.empty:
        st.info("Debe cargar primero el Excel generado por PUMI para esta región.")
    else:
        df_filtrado = aplicar_filtros_admin(df_admin)
        st.markdown("---")
        modulo_editar_eliminar_registros(df_filtrado)

        st.markdown("---")
        st.markdown("### Descargar Excel actualizado")
        boton_descargar_excel_admin(filtrar_df_por_region_autenticada(st.session_state.df_admin), key="descarga_edicion_regional")

# ======================================================
# DASHBOARD
# ======================================================

elif menu_admin == "Dashboard":

    mostrar_dashboard_avances_regional(df_admin)

    st.markdown("---")
    st.markdown("## Mapa y registros cargados")

    if df_admin.empty:
        st.info("No hay registros PUMI cargados para mostrar en el mapa. Las metas oficiales se mantienen visibles en el dashboard.")
    else:
        df_filtrado = aplicar_filtros_admin(df_admin)
        st.markdown("---")
        mostrar_mapa_admin(df_filtrado, key="mapa_dashboard_regional_login")
        st.markdown("### Registros incluidos")
        mostrar_tabla_resumen_validacion(df_filtrado)
        st.markdown("### Descargar Excel filtrado")
        boton_descargar_excel_admin(df_filtrado, key="descarga_dashboard_registros_filtrados")

# ======================================================
# DESCARGA GLOBAL Y LIMPIEZA EN SIDEBAR
# ======================================================

st.sidebar.markdown("---")
st.sidebar.markdown("### Opciones regionales")

df_sidebar = filtrar_df_por_region_autenticada(st.session_state.df_admin)

if df_sidebar.empty:
    st.sidebar.info("No hay datos cargados.")
else:
    nombre_excel_admin = generar_nombre_reporte(df_sidebar, tipo="XLSX")

    st.sidebar.download_button(
        "⬇️ Descargar Excel verificado",
        data=convertir_excel_admin(df_sidebar),
        file_name=nombre_excel_admin,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="descarga_sidebar_admin"
    )

    if st.sidebar.button("🧹 Limpiar datos cargados"):
        st.session_state.df_admin = pd.DataFrame(columns=ENCABEZADOS_COMPLETOS)
        st.session_state.archivo_admin_nombre = ""
        st.session_state.filtros_admin_aplicados = {}
        st.session_state.mapa_imagen_referencia = None
        st.session_state.grafico_imagen_referencia = None
        st.sidebar.success("Datos limpiados correctamente.")
        st.rerun()
