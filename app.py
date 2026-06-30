# ======================================================
# PARTE 1 DE 10
# LIBRERÍAS, CONFIGURACIÓN GENERAL, ARCHIVOS BASE,
# ENCABEZADOS, CATÁLOGOS Y SESSION STATE
# ======================================================

import streamlit as st
import pandas as pd
import plotly.express as px
import unicodedata
import os
import base64
import folium
import re
import glob

from streamlit_folium import st_folium
from geopy.geocoders import Nominatim
from geopy.extra.rate_limiter import RateLimiter
from streamlit_js_eval import get_geolocation
from datetime import datetime, date, time
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.drawing.image import Image


# ======================================================
# CONFIGURACIÓN GENERAL DE STREAMLIT
# ======================================================

st.set_page_config(
    page_title="PUMI 2026",
    page_icon="🛡️",
    layout="wide",
    initial_sidebar_state="expanded"
)


# ======================================================
# ARCHIVOS BASE LOCALES
# ======================================================

ARCHIVO_MEP = "BASE DE DATOS MEP 2025.xlsx"
ARCHIVO_DELEGACIONES = "DELEGACIONES Y DISTRITOS.xlsx"
ARCHIVO_DATOS_IMPORTANTES = "Datos Importantes.xlsx"
ARCHIVO_USUARIOS_CLAVES = "Usuarios y Claves PUMI.xlsx"
MARCA_AGUA_EXCEL = "marca_agua.png"

# Excel oficiales de metas regionales.
# Puede colocar archivos como DR2 AVANCE JUNIO.xlsx, DR3 AVANCE JUNIO.xlsx, etc.
# al mismo nivel de app.py o dentro de una carpeta llamada metas.
CARPETA_METAS_REGIONALES = "."
CARPETA_METAS_ALTERNATIVA = "metas"


# ======================================================
# LOGOS INSTITUCIONALES
# ======================================================

LOGO_MINISTERIO = "Logo2.jpeg"
LOGO_PUMI = "logo_pumi.jpeg"
LOGO_FUERZA_PUBLICA = "Logo1.jpeg"


# ======================================================
# NOMBRE DE HOJA PARA EXPORTAR / IMPORTAR EXCEL
# ======================================================

NOMBRE_HOJA_EXCEL = "REGISTRO_PUMI_2026"


# ======================================================
# ENCABEZADOS OFICIALES
# ======================================================

ENCABEZADOS = [
    "ID",
    "Fecha Actividad",
    "Hora Actividad",
    "Dirección Regional",
    "Delegación",
    "Responde a",
    "Programa",
    "Actividad",
    "Provincia",
    "Cantón",
    "Distrito",
    "Tipo Lugar",
    "Lugar",
    "Centro Educativo",
    "Código Presupuestario",
    "Dirección Mapa",
    "Latitud",
    "Longitud",
    "Responsable",
    "Avance Realizado",
    "Cantidad Participantes",
    "Cantidad Hombres",
    "Cantidad Mujeres",
    "Edad 10 a 18",
    "Edad 19 a 30",
    "Edad 31 a 45",
    "Edad 46 en adelante",
    "Instituciones Participantes",
    "Número de Referencia",
    "Número de Expediente Referencia",
    "Observaciones",
    "Usuario Registra"
]


# ======================================================
# CATÁLOGOS BASE
# ======================================================

PROGRAMAS = [
    "DARE",
    "GREAT",
    "MPAS",
    "PSCC",
    "VIF",
    "Política Pública"
]

REGIONES = [
    "R1 San José Central",
    "R2 San José Norte",
    "R3 San José Sur",
    "R4 Alajuela",
    "R5 Cartago",
    "R6 Heredia",
    "R7 Chorotega",
    "R8 Puntarenas",
    "R9 Limón",
    "R10 Brunca",
    "R11 Chorotega Norte",
    "R12"
]

PROVINCIAS = [
    "San José",
    "Alajuela",
    "Cartago",
    "Heredia",
    "Guanacaste",
    "Puntarenas",
    "Limón"
]

RANGOS_EDAD = [
    "Edad 10 a 18",
    "Edad 19 a 30",
    "Edad 31 a 45",
    "Edad 46 en adelante"
]

MESES_OFICIALES = [
    "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
    "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"
]


# ======================================================
# COLORES INSTITUCIONALES
# ======================================================

COLOR_AZUL = "#002B7F"
COLOR_AZUL_MEDIO = "#1E4FA3"
COLOR_AZUL_CLARO = "#DCE8FF"

COLOR_DORADO = "#B88A2A"
COLOR_DORADO_OSCURO = "#8A6418"
COLOR_DORADO_CLARO = "#F4E6C1"

COLOR_BLANCO = "#FFFFFF"
COLOR_GRIS = "#F4F6F8"
COLOR_GRIS_OSCURO = "#2F3542"

COLOR_ROJO = COLOR_DORADO
COLOR_ROJO_OSCURO = COLOR_DORADO_OSCURO
COLOR_ROJO_CLARO = COLOR_DORADO_CLARO

COLOR_VERDE = COLOR_DORADO
COLOR_VERDE_CLARO = COLOR_DORADO_CLARO


COLORES_PROGRAMA = {
    "DARE": "purple",
    "GREAT": "orange",
    "GREAT CAMP": "orange",
    "MPAS": "darkblue",
    "PSCC": "blue",
    "VIF": "orange",
    "Política Pública": "green"
}


# ======================================================
# CENTROS GENERALES PARA MAPA POR PROVINCIA
# ======================================================

CENTROS_PROVINCIA = {
    "San José": [9.9281, -84.0907],
    "Alajuela": [10.0162, -84.2116],
    "Cartago": [9.8644, -83.9194],
    "Heredia": [10.0024, -84.1165],
    "Guanacaste": [10.6267, -85.4437],
    "Puntarenas": [9.9763, -84.8384],
    "Limón": [9.9917, -83.0360]
}


# ======================================================
# SESSION STATE GENERAL
# Aquí se guarda temporalmente la información mientras
# la app está abierta.
# ======================================================

def inicializar_session_state():
    if "registros_pumi" not in st.session_state:
        st.session_state.registros_pumi = pd.DataFrame(columns=ENCABEZADOS)

    if "latitud_registro" not in st.session_state:
        st.session_state.latitud_registro = ""

    if "longitud_registro" not in st.session_state:
        st.session_state.longitud_registro = ""

    if "direccion_encontrada_mapa" not in st.session_state:
        st.session_state.direccion_encontrada_mapa = ""

    if "metodo_ubicacion_actual" not in st.session_state:
        st.session_state.metodo_ubicacion_actual = ""

    if "archivo_cargado_nombre" not in st.session_state:
        st.session_state.archivo_cargado_nombre = ""

    if "ultima_direccion_regional" not in st.session_state:
        st.session_state.ultima_direccion_regional = ""

    if "ultima_delegacion" not in st.session_state:
        st.session_state.ultima_delegacion = ""

    if "usuario_autenticado" not in st.session_state:
        st.session_state.usuario_autenticado = None


inicializar_session_state()
# ======================================================
# PARTE 2 DE 10
# ESTILOS CSS, LOGOS Y ENCABEZADO INSTITUCIONAL
# ======================================================

# ======================================================
# FUNCIÓN PARA CONVERTIR IMÁGENES A BASE64
# Se usa para mostrar logos dentro del HTML de Streamlit.
# ======================================================

def imagen_a_base64(ruta):
    if not os.path.exists(ruta):
        return ""

    try:
        with open(ruta, "rb") as archivo:
            return base64.b64encode(archivo.read()).decode()
    except Exception:
        return ""


# ======================================================
# ESTILOS CSS GENERALES DE LA APP
# ======================================================

st.markdown(
    f"""
    <style>

    .stApp {{
        background:
            radial-gradient(circle at top left, #FFFFFF 0, transparent 28%),
            radial-gradient(circle at top right, #DCE8FF 0, transparent 30%),
            linear-gradient(180deg, #FFFFFF 0%, #F4F7FB 48%, #FFFFFF 100%);
    }}

    section[data-testid="stSidebar"] {{
        background: linear-gradient(180deg, #FFFFFF 0%, #EAF1FF 45%, #FFFFFF 100%);
        border-right: 5px solid {COLOR_DORADO};
        box-shadow: 4px 0px 12px rgba(0,0,0,0.10);
    }}

    .sidebar-logo-pumi {{
        background: #FFFFFF;
        border-radius: 16px;
        padding: 12px;
        margin: 8px 0px 18px 0px;
        box-shadow: 0px 4px 14px rgba(0,0,0,0.10);
        text-align: center;
    }}

    .sidebar-logo-pumi img {{
        width: 100%;
        max-height: 145px;
        object-fit: contain;
    }}

    .encabezado-institucional {{
        background: #FFFFFF;
        border-radius: 22px;
        padding: 22px 34px;
        margin-bottom: 24px;
        box-shadow: 0px 8px 22px rgba(0,0,0,0.11);
        border-bottom: 7px solid {COLOR_DORADO};
        display: grid;
        grid-template-columns: 1fr 1fr 1fr;
        align-items: center;
        gap: 22px;
        min-height: 155px;
    }}

    .encabezado-logo-izq,
    .encabezado-logo-centro,
    .encabezado-logo-der {{
        height: 125px;
        display: flex;
        align-items: center;
    }}

    .encabezado-logo-izq {{
        justify-content: flex-start;
    }}

    .encabezado-logo-centro {{
        justify-content: center;
    }}

    .encabezado-logo-der {{
        justify-content: flex-end;
    }}

    .logo-ministerio-app {{
        width: 350px;
        height: 100px;
        object-fit: contain;
    }}

    .logo-pumi-app {{
        width: 300px;
        height: 100px;
        object-fit: contain;
    }}

    .logo-fp-app {{
        width: 400px;
        height: 150px;
        object-fit: contain;
    }}

    .titulo-principal {{
        background: linear-gradient(135deg, {COLOR_AZUL}, #243B63, {COLOR_DORADO});
        padding: 38px;
        border-radius: 22px;
        text-align: center;
        color: white;
        margin-bottom: 30px;
        box-shadow: 0px 8px 22px rgba(0,0,0,0.28);
        border: 3px solid {COLOR_BLANCO};
    }}

    .titulo-principal h1 {{
        font-size: 52px;
        margin-bottom: 8px;
        font-weight: 900;
        letter-spacing: 1.5px;
        color: white;
        text-shadow: 1px 2px 5px rgba(0,0,0,0.35);
    }}

    .titulo-principal h3 {{
        font-size: 27px;
        margin-top: 8px;
        font-weight: 600;
        color: white;
    }}

    .card-pumi,
    .card-azul,
    .card-dorado {{
        padding: 24px;
        border-radius: 20px;
        box-shadow: 0px 6px 16px rgba(0,0,0,0.13);
        margin-bottom: 20px;
        background: linear-gradient(135deg, #FFFFFF 0%, #F8FAFF 100%);
    }}

    .card-pumi {{
        border-left: 9px solid {COLOR_AZUL};
    }}

    .card-azul {{
        border-left: 9px solid {COLOR_DORADO};
    }}

    .card-dorado {{
        border-left: 9px solid {COLOR_DORADO};
    }}

    .bloque-datos {{
        background: linear-gradient(135deg, #FFFFFF 0%, #EEF4FF 100%);
        padding: 18px;
        border-radius: 18px;
        border-left: 7px solid {COLOR_AZUL};
        margin-bottom: 18px;
    }}

    .bloque-territorio {{
        background: linear-gradient(135deg, #FFFFFF 0%, #FFF8E8 100%);
        padding: 18px;
        border-radius: 18px;
        border-left: 7px solid {COLOR_DORADO};
        margin-bottom: 18px;
    }}

    .bloque-actividad {{
        background: linear-gradient(135deg, #FFFFFF 0%, #EEF4FF 100%);
        padding: 18px;
        border-radius: 18px;
        border-left: 7px solid {COLOR_AZUL};
        margin-bottom: 18px;
    }}

    .bloque-mapa {{
        background: linear-gradient(135deg, #FFFFFF 0%, #FFF8E8 100%);
        padding: 18px;
        border-radius: 18px;
        border-left: 7px solid {COLOR_DORADO};
        margin-bottom: 18px;
    }}

    .subtitulo-pumi {{
        color: {COLOR_AZUL};
        font-weight: 900;
        font-size: 30px;
        margin-bottom: 12px;
        text-align: center;
    }}

    .texto-pumi {{
        color: {COLOR_GRIS_OSCURO};
        font-size: 17px;
        line-height: 1.7;
        text-align: justify;
    }}

    div[data-testid="stMetric"] {{
        background: linear-gradient(145deg, #FFFFFF 0%, #F3F6FF 100%);
        padding: 20px;
        border-radius: 20px;
        box-shadow: 0px 5px 15px rgba(0,0,0,0.13);
        border-bottom: 6px solid {COLOR_DORADO};
    }}

    div[data-baseweb="select"] > div {{
        background-color: #FFFFFF !important;
        border-radius: 12px !important;
        border: 1.8px solid #AFC3EE !important;
    }}

    input {{
        background-color: #FFFFFF !important;
        border-radius: 12px !important;
        border: 1.8px solid #AFC3EE !important;
    }}

    textarea {{
        background-color: #FFFFFF !important;
        border-radius: 12px !important;
        border: 1.8px solid {COLOR_DORADO_CLARO} !important;
    }}

    .stButton > button {{
        background: linear-gradient(90deg, {COLOR_AZUL}, {COLOR_DORADO});
        color: white;
        border-radius: 12px;
        border: none;
        font-weight: 800;
        padding: 0.7rem 1.2rem;
        box-shadow: 0px 4px 10px rgba(0,0,0,0.18);
    }}

    .stDownloadButton > button {{
        background: linear-gradient(90deg, {COLOR_DORADO}, {COLOR_AZUL});
        color: white;
        border-radius: 12px;
        border: none;
        font-weight: 800;
        padding: 0.7rem 1.2rem;
        box-shadow: 0px 4px 10px rgba(0,0,0,0.18);
    }}

    div[data-testid="stDataFrame"] {{
        border-radius: 18px;
        overflow: hidden;
        box-shadow: 0px 5px 16px rgba(0,0,0,0.13);
        background-color: white;
    }}

    iframe {{
        border-radius: 18px !important;
        box-shadow: 0px 5px 16px rgba(0,0,0,0.13);
        width: 100% !important;
    }}

    hr {{
        border: none;
        height: 3px;
        background: linear-gradient(90deg, {COLOR_AZUL}, {COLOR_BLANCO}, {COLOR_DORADO});
        margin-top: 25px;
        margin-bottom: 25px;
    }}

    h1, h2, h3 {{
        color: {COLOR_AZUL};
        font-weight: 900;
    }}


    .meta-oficial-box {{
        width: 100%;
        box-sizing: border-box;
        background: #FFFFFF;
        border-radius: 20px;
        padding: 22px 24px;
        margin: 20px 0 24px 0;
        box-shadow: 0px 6px 18px rgba(0,0,0,0.10);
        border-left: 8px solid #002B7F;
        border-bottom: 5px solid #B88A2A;
        overflow: visible;
    }}

    .meta-titulo {{
        color: #002B7F;
        font-size: 28px;
        line-height: 1.15;
        font-weight: 900;
        margin-bottom: 18px;
    }}

    .meta-grid {{
        width: 100%;
        display: grid;
        grid-template-columns: repeat(4, minmax(150px, 1fr));
        gap: 16px;
        align-items: stretch;
    }}

    .meta-card {{
        background: linear-gradient(145deg, #FFFFFF 0%, #F3F6FF 100%);
        border-radius: 16px;
        padding: 18px 18px;
        border: 1px solid #E2E8F0;
        min-height: 105px;
        box-shadow: 0px 4px 12px rgba(0,0,0,0.08);
        box-sizing: border-box;
    }}

    .meta-label {{
        color: #4B5563;
        font-size: 16px;
        font-weight: 800;
        margin-bottom: 10px;
        white-space: normal;
        line-height: 1.2;
    }}

    .meta-value {{
        color: #111827;
        font-size: 36px;
        font-weight: 900;
        line-height: 1.05;
        word-break: normal;
    }}

    .meta-meses {{
        background: #EAF1FF;
        color: #002B7F;
        padding: 13px 16px;
        border-radius: 12px;
        margin-top: 16px;
        font-size: 16px;
        line-height: 1.35;
        width: 100%;
        box-sizing: border-box;
    }}

    .bloque-avance-registro {{
        background: linear-gradient(135deg, #FFFFFF 0%, #F8FAFF 100%);
        border-radius: 18px;
        padding: 20px 22px;
        margin: 10px 0 22px 0;
        border-left: 7px solid #B88A2A;
        box-shadow: 0px 4px 14px rgba(0,0,0,0.08);
    }}

    @media (max-width: 1200px) {{
        .meta-grid {{
            grid-template-columns: repeat(2, minmax(150px, 1fr));
        }}
    }}

    @media (max-width: 700px) {{
        .meta-grid {{
            grid-template-columns: 1fr;
        }}
    }}

    @media (max-width: 900px) {{
        .encabezado-institucional {{
            grid-template-columns: 1fr;
            text-align: center;
            min-height: auto;
        }}

        .encabezado-logo-izq,
        .encabezado-logo-centro,
        .encabezado-logo-der {{
            justify-content: center;
            height: auto;
        }}

        .logo-ministerio-app {{
            width: 300px;
            height: 80px;
        }}

        .logo-pumi-app {{
            width: 250px;
            height: 80px;
        }}

        .logo-fp-app {{
            width: 350px;
            height: 100px;
        }}
    }}

    </style>
    """,
    unsafe_allow_html=True
)


# ======================================================
# LOGO EN SIDEBAR
# ======================================================

def mostrar_logo():
    logo_pumi_b64 = imagen_a_base64(LOGO_PUMI)

    if logo_pumi_b64:
        st.sidebar.markdown(
            f"""
            <div class="sidebar-logo-pumi">
                <img src="data:image/jpeg;base64,{logo_pumi_b64}">
            </div>
            """,
            unsafe_allow_html=True
        )
    else:
        st.sidebar.warning("Logo PUMI no encontrado.")


# ======================================================
# ENCABEZADO INSTITUCIONAL SUPERIOR
# ======================================================

def mostrar_encabezado_institucional():
    logo_min_b64 = imagen_a_base64(LOGO_MINISTERIO)
    logo_pumi_b64 = imagen_a_base64(LOGO_PUMI)
    logo_fp_b64 = imagen_a_base64(LOGO_FUERZA_PUBLICA)

    img_min = (
        f'<img class="logo-ministerio-app" src="data:image/jpeg;base64,{logo_min_b64}">'
        if logo_min_b64 else ""
    )

    img_pumi = (
        f'<img class="logo-pumi-app" src="data:image/jpeg;base64,{logo_pumi_b64}">'
        if logo_pumi_b64 else ""
    )

    img_fp = (
        f'<img class="logo-fp-app" src="data:image/jpeg;base64,{logo_fp_b64}">'
        if logo_fp_b64 else ""
    )

    st.markdown(
        f"""
        <div class="encabezado-institucional">
            <div class="encabezado-logo-izq">
                {img_min}
            </div>
            <div class="encabezado-logo-centro">
                {img_pumi}
            </div>
            <div class="encabezado-logo-der">
                {img_fp}
            </div>
        </div>
        """,
        unsafe_allow_html=True
    )


def mostrar_logos_encabezado():
    mostrar_encabezado_institucional()


# ======================================================
# TÍTULO PRINCIPAL DE LA APP
# ======================================================

def mostrar_titulo_principal():
    st.markdown(
        """
        <div class="titulo-principal">
            <h1>🛡️ P.U.M.I. 2026</h1>
            <h3>Proceso Unificado para el Manejo de la Información</h3>
        </div>
        """,
        unsafe_allow_html=True
    )
    # ======================================================
# PARTE 3 DE 10
# FUNCIONES GENERALES, NORMALIZACIÓN, ORDENAMIENTO
# Y MANEJO DE REGISTROS EN SESSION STATE
# ======================================================


# ======================================================
# NORMALIZAR TEXTO
# Convierte texto a mayúsculas, elimina tildes y limpia espacios.
# Sirve para comparar columnas aunque tengan diferencias de escritura.
# ======================================================

def normalizar_texto(valor):
    if pd.isna(valor):
        return ""

    texto = str(valor).strip().upper()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(
        [
            caracter
            for caracter in texto
            if not unicodedata.combining(caracter)
        ]
    )
    texto = " ".join(texto.split())

    return texto


# ======================================================
# BUSCAR COLUMNA POR NOMBRE
# Permite encontrar columnas aunque vengan con tildes,
# mayúsculas, minúsculas o pequeñas variaciones.
# ======================================================

def obtener_columna_por_nombre(df, posibles_nombres):
    columnas = list(df.columns)

    for posible in posibles_nombres:
        posible_norm = normalizar_texto(posible)

        for col in columnas:
            if normalizar_texto(col) == posible_norm:
                return col

    return None


# ======================================================
# ORDENAR REGIONES NUMÉRICAMENTE
# Ordena R1, R2, R3... sin que R10 quede antes de R2.
# ======================================================

def ordenar_regiones_numericamente(lista_regiones):
    def extraer_numero_region(region):
        texto = str(region)
        numero = ""

        for caracter in texto:
            if caracter.isdigit():
                numero += caracter
            elif numero:
                break

        if numero:
            return int(numero)

        return 999

    return sorted(lista_regiones, key=extraer_numero_region)


# ======================================================
# SEPARAR CAMPO "RESPONDE A"
# Cuando una actividad responde a dos opciones separadas por "/",
# la función las divide para poder filtrarlas correctamente.
# ======================================================

def separar_responde_a(valor):
    if pd.isna(valor):
        return []

    texto = str(valor).strip()

    if texto == "" or texto.lower() == "nan":
        return []

    partes = texto.split("/")

    return [
        parte.strip()
        for parte in partes
        if parte.strip() != ""
    ]


# ======================================================
# PORCENTAJE EN FORMATO TEXTO
# ======================================================

def porcentaje(valor, total):
    try:
        valor = float(valor)
        total = float(total)

        if total == 0:
            return "0.0%"

        return f"{(valor / total) * 100:.1f}%"

    except Exception:
        return "0.0%"


# ======================================================
# LIMPIAR TEXTO PARA NOMBRE DE ARCHIVO
# Evita caracteres inválidos en Windows o en navegadores.
# ======================================================

def limpiar_nombre_archivo(valor):
    texto = str(valor).strip()

    if texto == "" or texto.lower() == "nan":
        return "SIN_DATO"

    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(
        [
            caracter
            for caracter in texto
            if not unicodedata.combining(caracter)
        ]
    )

    texto = texto.upper()
    texto = re.sub(r"[^\w\s-]", "", texto)
    texto = re.sub(r"\s+", "_", texto)
    texto = texto.replace("__", "_")
    texto = texto.strip("_")

    if texto == "":
        return "SIN_DATO"

    return texto


# ======================================================
# GENERAR NOMBRE AUTOMÁTICO PARA EXCEL
# Usa Dirección Regional + Delegación.
# Ejemplo:
# REGISTRO_PUMI_2026_R1_SAN_JOSE_CENTRAL_DELEGACION_PAVAS.xlsx
# ======================================================

def generar_nombre_excel(direccion_regional="", delegacion="", filtrado=False):
    direccion = limpiar_nombre_archivo(direccion_regional)
    deleg = limpiar_nombre_archivo(delegacion)

    prefijo = "REGISTRO_PUMI_2026"

    if filtrado:
        prefijo = "REGISTRO_PUMI_2026_FILTRADO"

    if direccion == "SIN_DATO" and deleg == "SIN_DATO":
        return f"{prefijo}.xlsx"

    return f"{prefijo}_{direccion}_{deleg}.xlsx"


# ======================================================
# OBTENER NOMBRE DE EXCEL SEGÚN DATOS DEL DATAFRAME
# Si hay una sola Dirección Regional y una sola Delegación,
# usa esos nombres. Si hay varias, usa "VARIAS".
# ======================================================

def generar_nombre_excel_desde_df(df, filtrado=False):
    if df is None or df.empty:
        return generar_nombre_excel(filtrado=filtrado)

    direccion = "VARIAS_DIRECCIONES"
    delegacion = "VARIAS_DELEGACIONES"

    if "Dirección Regional" in df.columns:
        regiones = (
            df["Dirección Regional"]
            .replace("", pd.NA)
            .dropna()
            .astype(str)
            .unique()
            .tolist()
        )

        if len(regiones) == 1:
            direccion = regiones[0]

    if "Delegación" in df.columns:
        delegaciones = (
            df["Delegación"]
            .replace("", pd.NA)
            .dropna()
            .astype(str)
            .unique()
            .tolist()
        )

        if len(delegaciones) == 1:
            delegacion = delegaciones[0]

    return generar_nombre_excel(
        direccion_regional=direccion,
        delegacion=delegacion,
        filtrado=filtrado
    )


# ======================================================
# PREPARAR DATAFRAME IMPORTADO
# Acomoda cualquier Excel cargado para que tenga las columnas oficiales.
# ======================================================

def preparar_dataframe_importado(df):
    df = df.copy()

    columnas_a_eliminar = [
        "Fecha Registro",
        "Plan Estratégico Relacionado"
    ]

    for col in columnas_a_eliminar:
        if col in df.columns:
            df = df.drop(columns=[col])

    for col in ENCABEZADOS:
        if col not in df.columns:
            df[col] = ""

    df = df[ENCABEZADOS]

    for col in ENCABEZADOS:
        df[col] = df[col].fillna("").astype(str)

    columnas_numericas = [
        "ID",
        "Avance Realizado",
        "Cantidad Participantes",
        "Cantidad Hombres",
        "Cantidad Mujeres",
        "Edad 10 a 18",
        "Edad 19 a 30",
        "Edad 31 a 45",
        "Edad 46 en adelante"
    ]

    for col in columnas_numericas:
        df[col] = (
            pd.to_numeric(df[col], errors="coerce")
            .fillna(0)
            .astype(int)
        )

    return df


# ======================================================
# CARGAR EXCEL EXISTENTE
# Permite continuar trabajando sobre un archivo exportado antes.
# ======================================================

def cargar_excel_existente(archivo_excel):
    try:
        df = pd.read_excel(
            archivo_excel,
            sheet_name=NOMBRE_HOJA_EXCEL
        )
    except Exception:
        try:
            df = pd.read_excel(archivo_excel)
        except Exception as e:
            st.error("No se pudo leer el archivo Excel cargado.")
            st.exception(e)
            return pd.DataFrame(columns=ENCABEZADOS)

    return preparar_dataframe_importado(df)


# ======================================================
# GENERAR ID CONSECUTIVO
# ======================================================

def generar_id_consecutivo():
    df = st.session_state.registros_pumi

    if df.empty:
        return 1

    ids = pd.to_numeric(df["ID"], errors="coerce").dropna()

    if ids.empty:
        return 1

    return int(ids.max()) + 1


# ======================================================
# AGREGAR REGISTRO A SESSION STATE
# ======================================================

def agregar_registro_session(registro):
    nuevo_df = pd.DataFrame(
        [registro],
        columns=ENCABEZADOS
    )

    st.session_state.registros_pumi = pd.concat(
        [
            st.session_state.registros_pumi,
            nuevo_df
        ],
        ignore_index=True
    )


# ======================================================
# ELIMINAR REGISTRO POR ID
# ======================================================

def eliminar_registro_session(id_registro):
    df = st.session_state.registros_pumi.copy()

    df = df[df["ID"].astype(str) != str(id_registro)]

    st.session_state.registros_pumi = df.reset_index(drop=True)


# ======================================================
# ACTUALIZAR REGISTRO POR ID
# ======================================================

def actualizar_registro_session(id_registro, nuevos_datos):
    df = st.session_state.registros_pumi.copy()

    mascara = df["ID"].astype(str) == str(id_registro)

    if not mascara.any():
        return False

    for col in ENCABEZADOS:
        df.loc[mascara, col] = nuevos_datos.get(col, "")

    st.session_state.registros_pumi = df.reset_index(drop=True)

    return True


# ======================================================
# LIMPIAR DATAFRAME PARA MÉTRICAS Y DASHBOARD
# Convierte columnas numéricas y fechas para poder graficar.
# ======================================================

def limpiar_dataframe_para_metricas(df):
    df = df.copy()

    columnas_numericas = [
        "Avance Realizado",
        "Cantidad Participantes",
        "Cantidad Hombres",
        "Cantidad Mujeres",
        "Edad 10 a 18",
        "Edad 19 a 30",
        "Edad 31 a 45",
        "Edad 46 en adelante"
    ]

    for col in columnas_numericas:
        if col in df.columns:
            df[col] = pd.to_numeric(
                df[col],
                errors="coerce"
            ).fillna(0)

    if "Fecha Actividad" in df.columns:
        df["Fecha Actividad"] = pd.to_datetime(
            df["Fecha Actividad"],
            errors="coerce",
            dayfirst=True
        )

    if "Latitud" in df.columns:
        df["Latitud"] = df["Latitud"].astype(str)

    if "Longitud" in df.columns:
        df["Longitud"] = df["Longitud"].astype(str)

    return df
    
# ======================================================
# LECTURA DE METAS REGIONALES DESDE EXCEL DR
# Lee archivos como DR2 AVANCE JUNIO.xlsx ubicados al mismo nivel de app.py.
# Estructura esperada por hoja/delegación:
# Columna A: Programa
# Columna B: Actividad
# Columna C: Meta 2026
# Columna D: Avance base del Excel
# Columnas siguientes: meses
# ======================================================

def limpiar_numero_meta(valor):
    if pd.isna(valor) or valor == "":
        return 0

    if isinstance(valor, str):
        valor = valor.replace("%", "").replace(",", ".").strip()

    try:
        return float(valor)
    except Exception:
        return 0


def obtener_codigo_region_desde_texto(valor):
    texto = normalizar_texto(valor)
    match = re.search(r"\bDR\s*([0-9]+)\b", texto)
    if match:
        return f"DR{match.group(1)}"

    match = re.search(r"DIRECCION REGIONAL\s*([0-9]+)", texto)
    if match:
        return f"DR{match.group(1)}"

    return ""


def cargar_catalogo_usuarios_claves():
    columnas = [
        "Delegaciones", "Claves", "Regiones", "Claves.1", "Programas", "CLAVES"
    ]

    if not os.path.exists(ARCHIVO_USUARIOS_CLAVES):
        return pd.DataFrame(columns=columnas)

    try:
        df = pd.read_excel(ARCHIVO_USUARIOS_CLAVES)
    except Exception:
        return pd.DataFrame(columns=columnas)

    for col in columnas:
        if col not in df.columns:
            df[col] = ""

    for col in columnas:
        df[col] = df[col].fillna("").astype(str).str.strip()

    return df


def obtener_mapa_regiones_desde_usuarios():
    df = cargar_catalogo_usuarios_claves()
    mapa = {}

    if df.empty:
        return mapa

    for _, row in df.iterrows():
        region = str(row.get("Regiones", "")).strip()
        clave = str(row.get("Claves.1", "")).strip().upper()

        if not region or not clave:
            continue

        codigo = obtener_codigo_region_desde_texto(clave)
        if codigo:
            mapa[codigo] = region

    return mapa


def obtener_region_desde_archivo_meta(ruta_archivo):
    nombre = os.path.basename(ruta_archivo)
    nombre_sin_ext = os.path.splitext(nombre)[0]
    codigo = obtener_codigo_region_desde_texto(nombre_sin_ext)

    if codigo:
        mapa_regiones = obtener_mapa_regiones_desde_usuarios()
        if codigo in mapa_regiones:
            return mapa_regiones[codigo]
        return codigo

    return nombre_sin_ext


def buscar_archivos_metas_regionales():
    rutas = []

    patrones = [
        os.path.join(CARPETA_METAS_REGIONALES, "DR*.xlsx"),
        os.path.join(CARPETA_METAS_REGIONALES, "DR*.xlsm"),
        os.path.join(CARPETA_METAS_REGIONALES, "DR*.xls"),
    ]

    if os.path.isdir(CARPETA_METAS_ALTERNATIVA):
        patrones.extend([
            os.path.join(CARPETA_METAS_ALTERNATIVA, "DR*.xlsx"),
            os.path.join(CARPETA_METAS_ALTERNATIVA, "DR*.xlsm"),
            os.path.join(CARPETA_METAS_ALTERNATIVA, "DR*.xls"),
        ])

    for patron in patrones:
        rutas.extend(glob.glob(patron))

    rutas_limpias = []
    for ruta in rutas:
        nombre = os.path.basename(ruta).upper()
        if nombre.startswith("~$"):
            continue
        if "REGISTRO_PUMI" in nombre or "INFORME_VALIDACION" in nombre:
            continue
        if ruta not in rutas_limpias:
            rutas_limpias.append(ruta)

    return sorted(rutas_limpias)


@st.cache_data(show_spinner=False)
def cargar_metas_regionales():
    registros = []
    archivos = buscar_archivos_metas_regionales()

    for ruta in archivos:
        try:
            xl = pd.ExcelFile(ruta)
        except Exception:
            continue

        region_archivo = obtener_region_desde_archivo_meta(ruta)

        for hoja in xl.sheet_names:
            if str(hoja).strip().upper().startswith("TOTAL"):
                continue

            try:
                raw = pd.read_excel(ruta, sheet_name=hoja, header=None)
            except Exception:
                continue

            if raw.empty or raw.shape[1] < 4:
                continue

            # En estos archivos las dos primeras filas son encabezados visuales.
            df = raw.iloc[2:].copy()

            if df.empty:
                continue

            df = df.rename(columns={0: "Programa", 1: "Actividad", 2: "Meta 2026", 3: "Avance base"})

            # Agregar meses según posición de columnas.
            for idx_mes, mes in enumerate(MESES_OFICIALES, start=4):
                if idx_mes in df.columns:
                    df[mes.title()] = df[idx_mes]
                else:
                    df[mes.title()] = 0

            df["Programa"] = df["Programa"].ffill()

            for _, fila in df.iterrows():
                programa = str(fila.get("Programa", "")).strip()
                actividad = str(fila.get("Actividad", "")).strip()

                if not programa or programa.lower() == "nan":
                    continue

                if not actividad or actividad.lower() == "nan":
                    continue

                meta = limpiar_numero_meta(fila.get("Meta 2026", 0))
                avance_base = limpiar_numero_meta(fila.get("Avance base", 0))

                # Evita filas decorativas sin meta ni avance.
                if meta == 0 and avance_base == 0:
                    meses_sum = sum(limpiar_numero_meta(fila.get(m.title(), 0)) for m in MESES_OFICIALES)
                    if meses_sum == 0:
                        continue

                item = {
                    "Archivo meta": os.path.basename(ruta),
                    "Dirección Regional": region_archivo,
                    "Delegación": str(hoja).strip(),
                    "Programa": programa,
                    "Actividad": actividad,
                    "Meta 2026": meta,
                    "Avance base": avance_base,
                    "Clave Programa": normalizar_texto(programa),
                    "Clave Actividad": normalizar_texto(actividad),
                    "Clave Delegación": normalizar_texto(hoja),
                    "Clave Regional": normalizar_texto(region_archivo),
                }

                for mes in MESES_OFICIALES:
                    item[mes.title()] = limpiar_numero_meta(fila.get(mes.title(), 0))

                registros.append(item)

    if not registros:
        return pd.DataFrame(columns=[
            "Archivo meta", "Dirección Regional", "Delegación", "Programa", "Actividad",
            "Meta 2026", "Avance base", "Clave Programa", "Clave Actividad",
            "Clave Delegación", "Clave Regional"
        ])

    df_metas = pd.DataFrame(registros)
    df_metas = df_metas.drop_duplicates(
        subset=["Archivo meta", "Delegación", "Programa", "Actividad"],
        keep="first"
    ).reset_index(drop=True)

    return df_metas


def metas_disponibles():
    return not cargar_metas_regionales().empty


def obtener_regiones_metas():
    df = cargar_metas_regionales()
    if df.empty:
        return []
    return sorted(df["Dirección Regional"].dropna().astype(str).unique().tolist())


def obtener_delegaciones_metas_por_region(region):
    df = cargar_metas_regionales()
    if df.empty or not region:
        return []

    region_norm = normalizar_texto(region)
    lista = df[df["Clave Regional"] == region_norm]["Delegación"].dropna().astype(str).unique().tolist()
    return sorted(lista)


def obtener_programas_metas(region, delegacion):
    df = cargar_metas_regionales()
    if df.empty or not region or not delegacion:
        return []

    region_norm = normalizar_texto(region)
    deleg_norm = normalizar_texto(delegacion)

    lista = df[
        (df["Clave Regional"] == region_norm) &
        (df["Clave Delegación"] == deleg_norm)
    ]["Programa"].dropna().astype(str).unique().tolist()

    return sorted(lista)


def obtener_actividades_metas(region, delegacion, programa):
    df = cargar_metas_regionales()
    if df.empty or not region or not delegacion or not programa:
        return []

    region_norm = normalizar_texto(region)
    deleg_norm = normalizar_texto(delegacion)
    programa_norm = normalizar_texto(programa)

    actividades = df[
        (df["Clave Regional"] == region_norm) &
        (df["Clave Delegación"] == deleg_norm) &
        (df["Clave Programa"] == programa_norm)
    ]["Actividad"].dropna().astype(str).unique().tolist()

    return actividades


def obtener_fila_meta(region, delegacion, programa, actividad):
    df = cargar_metas_regionales()
    if df.empty:
        return None

    region_norm = normalizar_texto(region)
    deleg_norm = normalizar_texto(delegacion)
    programa_norm = normalizar_texto(programa)
    actividad_norm = normalizar_texto(actividad)

    fila = df[
        (df["Clave Regional"] == region_norm) &
        (df["Clave Delegación"] == deleg_norm) &
        (df["Clave Programa"] == programa_norm) &
        (df["Clave Actividad"] == actividad_norm)
    ]

    if fila.empty:
        return None

    return fila.iloc[0].to_dict()


def clasificar_avance_meta(porc):
    if porc >= 1:
        return "Completa"
    if porc > 0:
        return "En avance"
    return "Pendiente"


def generar_avance_contra_metas(df_registros):
    df_metas = cargar_metas_regionales()

    if df_metas.empty:
        return pd.DataFrame()

    metas = df_metas.copy()

    if df_registros is None or df_registros.empty:
        metas["Realizado PUMI"] = 0
    else:
        registros = df_registros.copy()
        for col in ["Dirección Regional", "Delegación", "Programa", "Actividad"]:
            if col not in registros.columns:
                registros[col] = ""

        registros["Clave Delegación"] = registros["Delegación"].apply(normalizar_texto)
        registros["Clave Programa"] = registros["Programa"].apply(normalizar_texto)
        registros["Clave Actividad"] = registros["Actividad"].apply(normalizar_texto)

        if "Avance Realizado" not in registros.columns:
            registros["Avance Realizado"] = 1

        registros["Avance Realizado"] = pd.to_numeric(
            registros["Avance Realizado"],
            errors="coerce"
        ).fillna(0)

        conteo = (
            registros
            .groupby(["Clave Delegación", "Clave Programa", "Clave Actividad"], as_index=False)["Avance Realizado"]
            .sum()
            .rename(columns={"Avance Realizado": "Realizado PUMI"})
        )

        metas = metas.merge(
            conteo,
            on=["Clave Delegación", "Clave Programa", "Clave Actividad"],
            how="left"
        )
        metas["Realizado PUMI"] = metas["Realizado PUMI"].fillna(0)

    metas["Realizado total"] = metas["Avance base"] + metas["Realizado PUMI"]
    metas["Pendiente"] = (metas["Meta 2026"] - metas["Realizado total"]).clip(lower=0)
    metas["% Cumplimiento"] = metas.apply(
        lambda x: x["Realizado total"] / x["Meta 2026"] if x["Meta 2026"] else 0,
        axis=1
    )
    metas["Estado cumplimiento"] = metas["% Cumplimiento"].apply(clasificar_avance_meta)

    columnas = [
        "Dirección Regional", "Delegación", "Programa", "Actividad", "Meta 2026",
        "Avance base", "Realizado PUMI", "Realizado total", "Pendiente",
        "% Cumplimiento", "Estado cumplimiento", "Archivo meta"
    ]

    return metas[columnas].copy()


def obtener_programas_metas_por_delegacion(delegacion):
    df = cargar_metas_regionales()
    if df.empty or not delegacion:
        return []

    deleg_norm = normalizar_texto(delegacion)
    lista = df[df["Clave Delegación"] == deleg_norm]["Programa"].dropna().astype(str).unique().tolist()

    # Mantiene el orden institucional cuando aplique.
    programas_limpios = []
    for programa in lista:
        if normalizar_texto(programa) == "GREAT CAMP":
            if "GREAT" not in programas_limpios:
                programas_limpios.append("GREAT")
        elif programa not in programas_limpios:
            programas_limpios.append(programa)

    orden_oficial = [p for p in PROGRAMAS if p in programas_limpios]
    extras = sorted([p for p in programas_limpios if p not in orden_oficial])
    return orden_oficial + extras


def obtener_actividades_metas_por_delegacion_programa(delegacion, programa):
    df = cargar_metas_regionales()
    if df.empty or not delegacion or not programa:
        return []

    deleg_norm = normalizar_texto(delegacion)
    programa_norm = normalizar_texto(programa)

    actividades = df[
        (df["Clave Delegación"] == deleg_norm) &
        (df["Clave Programa"] == programa_norm)
    ]["Actividad"].dropna().astype(str).unique().tolist()

    return actividades


def obtener_fila_meta_por_delegacion(delegacion, programa, actividad):
    df = cargar_metas_regionales()
    if df.empty:
        return None

    deleg_norm = normalizar_texto(delegacion)
    programa_norm = normalizar_texto(programa)
    actividad_norm = normalizar_texto(actividad)

    fila = df[
        (df["Clave Delegación"] == deleg_norm) &
        (df["Clave Programa"] == programa_norm) &
        (df["Clave Actividad"] == actividad_norm)
    ]

    if fila.empty:
        return None

    return fila.iloc[0].to_dict()


def mostrar_resumen_meta_seleccionada(fila_meta):
    if not fila_meta:
        return 1, 0, 0, 0

    meta = float(fila_meta.get("Meta 2026", 0) or 0)
    avance_base = float(fila_meta.get("Avance base", 0) or 0)
    pendiente = max(meta - avance_base, 0)
    porcentaje_meta = avance_base / meta if meta else 0

    st.markdown(
        f"""
        <div class="meta-oficial-box">
            <div class="meta-titulo">Meta oficial</div>
            <div class="meta-grid">
                <div class="meta-card">
                    <div class="meta-label">Meta 2026</div>
                    <div class="meta-value">{meta:,.0f}</div>
                </div>
                <div class="meta-card">
                    <div class="meta-label">Avance base</div>
                    <div class="meta-value">{avance_base:,.0f}</div>
                </div>
                <div class="meta-card">
                    <div class="meta-label">Pendiente</div>
                    <div class="meta-value">{pendiente:,.0f}</div>
                </div>
                <div class="meta-card">
                    <div class="meta-label">% base</div>
                    <div class="meta-value">{porcentaje_meta:.1%}</div>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True
    )

    valor_sugerido = 1 if pendiente >= 1 else 0
    return valor_sugerido, int(meta), int(avance_base), int(pendiente)


# ======================================================
# PARTE 4 DE 10
# CARGA DE CATÁLOGOS:
# DATOS IMPORTANTES, BASE MEP Y DELEGACIONES
# ======================================================


# ======================================================
# CARGA DE DATOS IMPORTANTES
# Contiene:
# - Provincia
# - Cantón
# - Distrito
# - Dirección Regional
# - Delegación
# - Responde a
# - Actividad Realizada
# - Programa
# ======================================================

@st.cache_data
def cargar_datos_importantes():
    columnas_base = [
        "Provincia",
        "Cantón",
        "Distrito",
        "Dirección Regional",
        "Delegación",
        "Responde a",
        "Actividad Realizada",
        "Programa"
    ]

    if not os.path.exists(ARCHIVO_DATOS_IMPORTANTES):
        return pd.DataFrame(columns=columnas_base)

    try:
        df_original = pd.read_excel(ARCHIVO_DATOS_IMPORTANTES)

        col_provincia = obtener_columna_por_nombre(df_original, ["Provincia"])
        col_canton = obtener_columna_por_nombre(df_original, ["Cantón", "Canton"])
        col_distrito = obtener_columna_por_nombre(df_original, ["Distrito", "Distritos"])

        col_region = obtener_columna_por_nombre(
            df_original,
            ["Dirección Regional", "Direccion Regional"]
        )

        col_delegacion = obtener_columna_por_nombre(
            df_original,
            ["Delegación", "Delegacion"]
        )

        col_responde_a = obtener_columna_por_nombre(
            df_original,
            [
                "Responde a",
                "Responde a:",
                "Responde",
                "Responda a",
                "Responda a:"
            ]
        )

        col_actividad = obtener_columna_por_nombre(
            df_original,
            [
                "Actividad Realizada",
                "Actividad"
            ]
        )

        col_programa = obtener_columna_por_nombre(
            df_original,
            ["Programa"]
        )

        columnas_requeridas = [
            col_provincia,
            col_canton,
            col_distrito,
            col_region,
            col_delegacion,
            col_responde_a,
            col_actividad,
            col_programa
        ]

        if not all(columnas_requeridas):
            st.warning(
                "El archivo Datos Importantes.xlsx no tiene todas las columnas requeridas. "
                "Debe incluir: Provincia, Cantón, Distrito, Dirección Regional, Delegación, "
                "Responde a, Actividad Realizada y Programa."
            )
            return pd.DataFrame(columns=columnas_base)

        df = df_original[
            [
                col_provincia,
                col_canton,
                col_distrito,
                col_region,
                col_delegacion,
                col_responde_a,
                col_actividad,
                col_programa
            ]
        ].copy()

        df.columns = columnas_base

        for col in columnas_base:
            df[col] = df[col].fillna("").astype(str).str.strip()

        df = df.replace("nan", "")
        df = df.drop_duplicates()

        df["Provincia_Normalizada"] = df["Provincia"].apply(normalizar_texto)
        df["Cantón_Normalizado"] = df["Cantón"].apply(normalizar_texto)
        df["Distrito_Normalizado"] = df["Distrito"].apply(normalizar_texto)
        df["Región_Normalizada"] = df["Dirección Regional"].apply(normalizar_texto)
        df["Delegación_Normalizada"] = df["Delegación"].apply(normalizar_texto)
        df["Responde_a_Normalizado"] = df["Responde a"].apply(normalizar_texto)
        df["Actividad_Normalizada"] = df["Actividad Realizada"].apply(normalizar_texto)
        df["Programa_Normalizado"] = df["Programa"].apply(normalizar_texto)

        filas_expandidas = []

        for _, fila in df.iterrows():
            opciones_responde = separar_responde_a(fila["Responde a"])

            if not opciones_responde:
                opciones_responde = [fila["Responde a"]]

            for opcion in opciones_responde:
                nueva_fila = fila.copy()
                nueva_fila["Responde a"] = opcion
                nueva_fila["Responde_a_Normalizado"] = normalizar_texto(opcion)
                filas_expandidas.append(nueva_fila)

        if filas_expandidas:
            df = pd.DataFrame(filas_expandidas)
            df = df.drop_duplicates()

        return df

    except Exception as e:
        st.error(f"Error leyendo {ARCHIVO_DATOS_IMPORTANTES}")
        st.exception(e)
        return pd.DataFrame(columns=columnas_base)


# ======================================================
# OBTENER REGIONES DESDE DATOS IMPORTANTES
# ======================================================

def obtener_regiones_datos():
    df = cargar_datos_importantes()

    if df.empty:
        return ordenar_regiones_numericamente(REGIONES)

    regiones = (
        df["Dirección Regional"]
        .dropna()
        .unique()
        .tolist()
    )

    regiones = [
        x for x in regiones
        if str(x).strip() != ""
    ]

    return ordenar_regiones_numericamente(regiones)


# ======================================================
# OBTENER DELEGACIONES SEGÚN DIRECCIÓN REGIONAL
# ======================================================

def obtener_delegaciones_por_region(region):
    df = cargar_datos_importantes()

    if df.empty or not region:
        return obtener_delegaciones_unicas()

    region_norm = normalizar_texto(region)

    delegaciones = (
        df[df["Región_Normalizada"] == region_norm]["Delegación"]
        .dropna()
        .unique()
        .tolist()
    )

    delegaciones = [
        x for x in delegaciones
        if str(x).strip() != ""
    ]

    return sorted(delegaciones)


# ======================================================
# OBTENER RESPONDE A
# ======================================================

def obtener_responde_a_datos():
    df = cargar_datos_importantes()

    if df.empty or "Responde a" not in df.columns:
        return []

    responde_a = (
        df["Responde a"]
        .dropna()
        .unique()
        .tolist()
    )

    responde_a = [
        x for x in responde_a
        if str(x).strip() != ""
    ]

    return sorted(responde_a)


# ======================================================
# OBTENER PROGRAMAS SEGÚN RESPONDE A
# ======================================================

def obtener_programas_por_responde_a(responde_a):
    df = cargar_datos_importantes()

    if df.empty or not responde_a:
        return obtener_programas_datos()

    responde_norm = normalizar_texto(responde_a)

    programas = (
        df[df["Responde_a_Normalizado"] == responde_norm]["Programa"]
        .dropna()
        .unique()
        .tolist()
    )

    programas = [
        x for x in programas
        if str(x).strip() != ""
    ]

    if not programas:
        return []

    programas_limpios = []

    for programa in programas:
        if normalizar_texto(programa) == "GREAT CAMP":
            if "GREAT" not in programas_limpios:
                programas_limpios.append("GREAT")
        elif programa not in programas_limpios:
            programas_limpios.append(programa)

    orden_oficial = [
        p for p in PROGRAMAS
        if p in programas_limpios
    ]

    extras = sorted(
        [
            p for p in programas_limpios
            if p not in orden_oficial
        ]
    )

    return orden_oficial + extras


# ======================================================
# OBTENER ACTIVIDADES SEGÚN RESPONDE A Y PROGRAMA
# ======================================================

def obtener_actividades_por_responde_a_programa(responde_a, programa):
    df = cargar_datos_importantes()

    if df.empty or not responde_a or not programa:
        return []

    responde_norm = normalizar_texto(responde_a)
    programa_norm = normalizar_texto(programa)

    if programa_norm == "GREAT":
        actividades = (
            df[
                (df["Responde_a_Normalizado"] == responde_norm) &
                (df["Programa_Normalizado"].isin(["GREAT", "GREAT CAMP"]))
            ]["Actividad Realizada"]
            .dropna()
            .unique()
            .tolist()
        )
    else:
        actividades = (
            df[
                (df["Responde_a_Normalizado"] == responde_norm) &
                (df["Programa_Normalizado"] == programa_norm)
            ]["Actividad Realizada"]
            .dropna()
            .unique()
            .tolist()
        )

    actividades = [
        x for x in actividades
        if str(x).strip() != ""
    ]

    return sorted(actividades)


# ======================================================
# OBTENER PROGRAMAS GENERALES
# ======================================================

def obtener_programas_datos():
    df = cargar_datos_importantes()

    if df.empty:
        return PROGRAMAS

    programas = (
        df["Programa"]
        .dropna()
        .unique()
        .tolist()
    )

    programas = [
        x for x in programas
        if str(x).strip() != ""
    ]

    if not programas:
        return PROGRAMAS

    programas_limpios = []

    for programa in programas:
        if normalizar_texto(programa) == "GREAT CAMP":
            if "GREAT" not in programas_limpios:
                programas_limpios.append("GREAT")
        elif programa not in programas_limpios:
            programas_limpios.append(programa)

    orden_oficial = [
        p for p in PROGRAMAS
        if p in programas_limpios
    ]

    extras = sorted(
        [
            p for p in programas_limpios
            if p not in orden_oficial
        ]
    )

    return orden_oficial + extras


# ======================================================
# OBTENER ACTIVIDADES SEGÚN PROGRAMA
# ======================================================

def obtener_actividades_por_programa(programa):
    df = cargar_datos_importantes()

    if df.empty or not programa:
        return []

    programa_norm = normalizar_texto(programa)

    if programa_norm == "GREAT":
        actividades = (
            df[
                df["Programa_Normalizado"].isin(["GREAT", "GREAT CAMP"])
            ]["Actividad Realizada"]
            .dropna()
            .unique()
            .tolist()
        )
    else:
        actividades = (
            df[
                df["Programa_Normalizado"] == programa_norm
            ]["Actividad Realizada"]
            .dropna()
            .unique()
            .tolist()
        )

    actividades = [
        x for x in actividades
        if str(x).strip() != ""
    ]

    return sorted(actividades)


# ======================================================
# OBTENER PROVINCIAS
# ======================================================

def obtener_provincias_datos():
    df = cargar_datos_importantes()

    if df.empty:
        return PROVINCIAS

    provincias = (
        df["Provincia"]
        .dropna()
        .unique()
        .tolist()
    )

    provincias = [
        x for x in provincias
        if str(x).strip() != ""
    ]

    return sorted(provincias)


# ======================================================
# OBTENER CANTONES POR PROVINCIA
# ======================================================

def obtener_cantones_por_provincia(provincia):
    df = cargar_datos_importantes()

    if df.empty or not provincia:
        return []

    provincia_norm = normalizar_texto(provincia)

    cantones = (
        df[df["Provincia_Normalizada"] == provincia_norm]["Cantón"]
        .dropna()
        .unique()
        .tolist()
    )

    cantones = [
        x for x in cantones
        if str(x).strip() != ""
    ]

    return sorted(cantones)


# ======================================================
# OBTENER DISTRITOS POR PROVINCIA Y CANTÓN
# ======================================================

def obtener_distritos_por_provincia_canton(provincia, canton):
    df = cargar_datos_importantes()

    if df.empty or not provincia or not canton:
        return []

    provincia_norm = normalizar_texto(provincia)
    canton_norm = normalizar_texto(canton)

    distritos = (
        df[
            (df["Provincia_Normalizada"] == provincia_norm) &
            (df["Cantón_Normalizado"] == canton_norm)
        ]["Distrito"]
        .dropna()
        .unique()
        .tolist()
    )

    distritos = [
        x for x in distritos
        if str(x).strip() != ""
    ]

    return sorted(distritos)


# ======================================================
# BASE MEP
# Carga centros educativos y código presupuestario.
# ======================================================

@st.cache_data
def cargar_base_mep():
    columnas_base = [
        "PROVINCIA",
        "NOMBRE",
        "CODIGO_PRESUPUESTARIO"
    ]

    if not os.path.exists(ARCHIVO_MEP):
        return pd.DataFrame(columns=columnas_base)

    try:
        df_original = pd.read_excel(ARCHIVO_MEP)

        col_provincia = obtener_columna_por_nombre(
            df_original,
            ["PROVINCIA", "Provincia"]
        )

        col_nombre = obtener_columna_por_nombre(
            df_original,
            [
                "NOMBRE",
                "Nombre",
                "CENTRO EDUCATIVO",
                "Centro Educativo",
                "INSTITUCION",
                "Institución"
            ]
        )

        col_codigo = obtener_columna_por_nombre(
            df_original,
            [
                "CODIGO PRESUPUESTARIO",
                "CÓDIGO PRESUPUESTARIO",
                "Codigo Presupuestario",
                "Código Presupuestario",
                "CODIGO",
                "CÓDIGO"
            ]
        )

        if not col_provincia or not col_nombre:
            return pd.DataFrame(columns=columnas_base)

        if not col_codigo:
            df_original["CODIGO_PRESUPUESTARIO_TMP"] = ""
            col_codigo = "CODIGO_PRESUPUESTARIO_TMP"

        df = df_original[
            [
                col_provincia,
                col_nombre,
                col_codigo
            ]
        ].copy()

        df.columns = columnas_base

        for col in columnas_base:
            df[col] = df[col].fillna("").astype(str).str.strip()

        df = df.dropna(subset=["PROVINCIA", "NOMBRE"])

        df["PROVINCIA_NORMALIZADA"] = df["PROVINCIA"].apply(normalizar_texto)
        df["NOMBRE_NORMALIZADO"] = df["NOMBRE"].apply(normalizar_texto)

        df["CENTRO_MOSTRAR"] = df.apply(
            lambda row: f"{row['NOMBRE']} - Código: {row['CODIGO_PRESUPUESTARIO']}"
            if row["CODIGO_PRESUPUESTARIO"].strip() != ""
            else row["NOMBRE"],
            axis=1
        )

        df = df.drop_duplicates(
            subset=[
                "PROVINCIA_NORMALIZADA",
                "NOMBRE_NORMALIZADO",
                "CODIGO_PRESUPUESTARIO"
            ]
        )

        return df

    except Exception as e:
        st.error(f"Error leyendo {ARCHIVO_MEP}")
        st.exception(e)
        return pd.DataFrame(columns=columnas_base)


# ======================================================
# OBTENER CENTROS EDUCATIVOS POR PROVINCIA
# ======================================================

def obtener_centros_por_provincia(provincia):
    df = cargar_base_mep()

    if df.empty or not provincia:
        return []

    provincia_norm = normalizar_texto(provincia)

    centros = (
        df[df["PROVINCIA_NORMALIZADA"] == provincia_norm]["CENTRO_MOSTRAR"]
        .dropna()
        .astype(str)
        .drop_duplicates()
        .sort_values()
        .tolist()
    )

    return centros


# ======================================================
# OBTENER DATOS DEL CENTRO EDUCATIVO
# ======================================================

def obtener_datos_centro_educativo(centro_mostrar):
    df = cargar_base_mep()

    if df.empty or not centro_mostrar:
        return "", ""

    fila = df[df["CENTRO_MOSTRAR"] == centro_mostrar]

    if fila.empty:
        return centro_mostrar, ""

    nombre = fila.iloc[0].get("NOMBRE", "")
    codigo = fila.iloc[0].get("CODIGO_PRESUPUESTARIO", "")

    return nombre, codigo


# ======================================================
# BASE DELEGACIONES Y DISTRITOS COMO RESPALDO
# ======================================================

@st.cache_data
def cargar_base_delegaciones():
    if not os.path.exists(ARCHIVO_DELEGACIONES):
        return pd.DataFrame(columns=["Delegacion", "Distrito"])

    try:
        df_original = pd.read_excel(ARCHIVO_DELEGACIONES)

        col_delegacion = obtener_columna_por_nombre(
            df_original,
            ["Delegacion", "Delegación"]
        )

        col_distrito = obtener_columna_por_nombre(
            df_original,
            ["Distrito"]
        )

        col_provincia = obtener_columna_por_nombre(
            df_original,
            ["Provincia"]
        )

        col_canton = obtener_columna_por_nombre(
            df_original,
            ["Canton", "Cantón"]
        )

        columnas = {}

        if col_delegacion:
            columnas[col_delegacion] = "Delegacion"

        if col_distrito:
            columnas[col_distrito] = "Distrito"

        if col_provincia:
            columnas[col_provincia] = "Provincia"

        if col_canton:
            columnas[col_canton] = "Cantón"

        if not col_delegacion and not col_distrito:
            return pd.DataFrame(
                columns=[
                    "Delegacion",
                    "Distrito",
                    "Provincia",
                    "Cantón"
                ]
            )

        df = df_original[list(columnas.keys())].copy()
        df = df.rename(columns=columnas)

        for col in ["Delegacion", "Distrito", "Provincia", "Cantón"]:
            if col not in df.columns:
                df[col] = ""

            df[col] = df[col].fillna("").astype(str).str.strip()

        df = df.replace("nan", "")

        df["Delegacion_Normalizada"] = df["Delegacion"].apply(normalizar_texto)
        df["Distrito_Normalizado"] = df["Distrito"].apply(normalizar_texto)
        df["Provincia_Normalizada"] = df["Provincia"].apply(normalizar_texto)
        df["Cantón_Normalizado"] = df["Cantón"].apply(normalizar_texto)

        df = df.drop_duplicates()

        return df

    except Exception as e:
        st.error(f"Error leyendo {ARCHIVO_DELEGACIONES}")
        st.exception(e)
        return pd.DataFrame(columns=["Delegacion", "Distrito"])


# ======================================================
# OBTENER TODAS LAS DELEGACIONES ÚNICAS
# ======================================================

def obtener_delegaciones_unicas():
    df_datos = cargar_datos_importantes()

    if not df_datos.empty:
        delegaciones = (
            df_datos["Delegación"]
            .dropna()
            .unique()
            .tolist()
        )

        delegaciones = [
            x for x in delegaciones
            if str(x).strip() != ""
        ]

        return sorted(delegaciones)

    df = cargar_base_delegaciones()

    if df.empty or "Delegacion" not in df.columns:
        return []

    df_tmp = df[
        ["Delegacion", "Delegacion_Normalizada"]
    ].drop_duplicates(
        subset=["Delegacion_Normalizada"]
    )

    return sorted(
        [
            x for x in df_tmp["Delegacion"].dropna().tolist()
            if str(x).strip() != ""
        ]
    )
# ======================================================
# PARTE 5 DE 10
# MAPA, GEOREFERENCIA, GPS, COORDENADAS MANUALES
# Y MARCADOR VISIBLE SIEMPRE
# ======================================================


# ======================================================
# GEOREFERENCIAR DIRECCIÓN O LUGAR
# Se conserva por compatibilidad, pero ya no se usa
# en el formulario para evitar llenar "Dirección Mapa".
# ======================================================

@st.cache_data(show_spinner=False)
def georreferenciar_direccion(direccion):
    if not direccion or str(direccion).strip() == "":
        return None, None, ""

    try:
        geolocator = Nominatim(
            user_agent="pumi_2026_streamlit_app"
        )

        geocode = RateLimiter(
            geolocator.geocode,
            min_delay_seconds=1,
            max_retries=2,
            error_wait_seconds=2
        )

        consulta = f"{direccion}, Costa Rica"
        location = geocode(consulta)

        if location:
            return location.latitude, location.longitude, location.address

        return None, None, ""

    except Exception:
        return None, None, ""


# ======================================================
# LIMPIAR COORDENADAS
# ======================================================

def limpiar_coordenada(valor):
    try:
        if valor is None or str(valor).strip() == "":
            return None

        return float(str(valor).replace(",", ".").strip())

    except Exception:
        return None


# ======================================================
# COLOR DE MARCADOR SEGÚN PROGRAMA
# ======================================================

def obtener_color_programa(programa):
    return COLORES_PROGRAMA.get(programa, "gray")


# ======================================================
# CREAR MAPA BASE
# ======================================================

def crear_mapa_base(centro, zoom, tipo_mapa):
    mapa = folium.Map(
        location=centro,
        zoom_start=zoom,
        tiles=None
    )

    if tipo_mapa == "OpenStreetMap":
        folium.TileLayer(
            "OpenStreetMap",
            name="OpenStreetMap"
        ).add_to(mapa)

    elif tipo_mapa == "Mapa claro":
        folium.TileLayer(
            "CartoDB positron",
            name="Mapa claro"
        ).add_to(mapa)

    elif tipo_mapa == "Topográfico":
        folium.TileLayer(
            "OpenTopoMap",
            name="Topográfico"
        ).add_to(mapa)

    elif tipo_mapa == "Satélite":
        folium.TileLayer(
            tiles=(
                "https://server.arcgisonline.com/ArcGIS/rest/services/"
                "World_Imagery/MapServer/tile/{z}/{y}/{x}"
            ),
            attr="Esri World Imagery",
            name="Satélite"
        ).add_to(mapa)

    folium.LayerControl().add_to(mapa)

    return mapa


# ======================================================
# OBTENER CENTRO DEL MAPA DE REGISTRO
# ======================================================

def obtener_centro_mapa_registro(provincia):
    lat_actual = limpiar_coordenada(
        st.session_state.get("latitud_registro", "")
    )

    lon_actual = limpiar_coordenada(
        st.session_state.get("longitud_registro", "")
    )

    if lat_actual is not None and lon_actual is not None:
        return [lat_actual, lon_actual], 16

    centro_provincia = CENTROS_PROVINCIA.get(
        provincia,
        [9.7489, -83.7534]
    )

    return centro_provincia, 10


# ======================================================
# CREAR MAPA DE REGISTRO INDIVIDUAL
# ======================================================

def crear_mapa_registro_individual(
    provincia,
    tipo_mapa,
    lugar="",
    delegacion="",
    programa="",
    permitir_click=True
):
    centro_mapa, zoom_mapa = obtener_centro_mapa_registro(provincia)

    mapa = crear_mapa_base(
        centro_mapa,
        zoom_mapa,
        tipo_mapa
    )

    lat_actual = limpiar_coordenada(
        st.session_state.get("latitud_registro", "")
    )

    lon_actual = limpiar_coordenada(
        st.session_state.get("longitud_registro", "")
    )

    if lat_actual is not None and lon_actual is not None:
        direccion_encontrada = st.session_state.get(
            "direccion_encontrada_mapa",
            ""
        )

        popup_html = f"""
        <div style="font-family: Arial; width: 280px;">
            <h4 style="color:#002B7F; margin-bottom:6px;">
                Punto seleccionado para registro
            </h4>
            <b>Delegación:</b> {delegacion}<br>
            <b>Programa:</b> {programa}<br>
            <b>Lugar:</b> {lugar}<br>
            <b>Referencia:</b> {direccion_encontrada}<br>
            <b>Latitud:</b> {lat_actual}<br>
            <b>Longitud:</b> {lon_actual}<br>
        </div>
        """

        folium.Marker(
            location=[lat_actual, lon_actual],
            popup=folium.Popup(popup_html, max_width=330),
            tooltip="Punto registrado",
            icon=folium.Icon(
                color="red",
                icon="map-marker"
            )
        ).add_to(mapa)

        folium.Circle(
            location=[lat_actual, lon_actual],
            radius=180,
            color=COLOR_AZUL,
            fill=True,
            fill_color=COLOR_DORADO,
            fill_opacity=0.18,
            weight=2
        ).add_to(mapa)

    return mapa


# ======================================================
# MOSTRAR MAPA INTERACTIVO DE REGISTRO
# ======================================================

def mostrar_mapa_registro_interactivo(
    provincia,
    tipo_mapa,
    lugar="",
    delegacion="",
    programa="",
    key="mapa_registro_general"
):
    mapa_registro = crear_mapa_registro_individual(
        provincia=provincia,
        tipo_mapa=tipo_mapa,
        lugar=lugar,
        delegacion=delegacion,
        programa=programa
    )

    resultado_mapa = st_folium(
        mapa_registro,
        height=520,
        use_container_width=True,
        key=key
    )

    if resultado_mapa and resultado_mapa.get("last_clicked"):
        lat_click = resultado_mapa["last_clicked"]["lat"]
        lon_click = resultado_mapa["last_clicked"]["lng"]

        st.session_state.latitud_registro = str(lat_click)
        st.session_state.longitud_registro = str(lon_click)
        st.session_state.direccion_encontrada_mapa = "Punto marcado manualmente en el mapa"

        st.success(
            f"Punto seleccionado en el mapa: {lat_click:.6f}, {lon_click:.6f}"
        )

        st.rerun()


# ======================================================
# BLOQUE COMPLETO DE GEOREFERENCIA PARA EL FORMULARIO
# ======================================================

def bloque_georreferencia_formulario(
    provincia,
    canton,
    distrito,
    lugar,
    delegacion,
    programa
):
    st.markdown(
        """
        <div class="bloque-mapa">
            <b>Georreferencia del lugar</b>
        </div>
        """,
        unsafe_allow_html=True
    )

    tipo_mapa_registro = st.selectbox(
        "Tipo de mapa",
        [
            "OpenStreetMap",
            "Mapa claro",
            "Topográfico",
            "Satélite"
        ],
        key="tipo_mapa_registro"
    )

    metodo_ubicacion = st.radio(
        "Método para registrar ubicación",
        [
            "Usar GPS del dispositivo",
            "Ingresar coordenadas manualmente",
            "Marcar punto en el mapa",
            "No registrar ubicación"
        ],
        horizontal=False,
        key="metodo_ubicacion_registro"
    )

    st.session_state.metodo_ubicacion_actual = metodo_ubicacion

    # ==================================================
    # USAR GPS DEL DISPOSITIVO
    # ==================================================

    if metodo_ubicacion == "Usar GPS del dispositivo":
        st.info(
            "El navegador puede solicitar permiso para acceder a la ubicación "
            "del dispositivo. Cuando se obtenga la ubicación, el punto se verá "
            "marcado en el mapa."
        )

        ubicacion_gps = get_geolocation()

        if ubicacion_gps and "coords" in ubicacion_gps:
            lat_gps = ubicacion_gps["coords"].get("latitude", "")
            lon_gps = ubicacion_gps["coords"].get("longitude", "")

            if lat_gps and lon_gps:
                st.session_state.latitud_registro = str(lat_gps)
                st.session_state.longitud_registro = str(lon_gps)
                st.session_state.direccion_encontrada_mapa = "Ubicación obtenida por GPS del dispositivo"

                st.success("Ubicación GPS obtenida correctamente.")
            else:
                st.warning("No se recibieron coordenadas válidas desde el GPS.")
        else:
            st.warning(
                "No se obtuvo ubicación GPS todavía. "
                "Revise que el navegador tenga permisos de ubicación."
            )

    # ==================================================
    # INGRESAR COORDENADAS MANUALMENTE
    # ==================================================

    elif metodo_ubicacion == "Ingresar coordenadas manualmente":
        st.info(
            "Ingrese latitud y longitud. El mapa se actualizará mostrando "
            "el punto registrado."
        )

        col_lat, col_lon = st.columns(2)

        with col_lat:
            lat_manual = st.text_input(
                "Latitud",
                value=st.session_state.latitud_registro,
                key="lat_manual_registro"
            )

        with col_lon:
            lon_manual = st.text_input(
                "Longitud",
                value=st.session_state.longitud_registro,
                key="lon_manual_registro"
            )

        st.session_state.latitud_registro = lat_manual
        st.session_state.longitud_registro = lon_manual
        st.session_state.direccion_encontrada_mapa = "Coordenadas ingresadas manualmente"

    # ==================================================
    # MARCAR PUNTO EN MAPA
    # ==================================================

    elif metodo_ubicacion == "Marcar punto en el mapa":
        st.info(
            "Haga clic sobre el mapa para seleccionar el punto exacto. "
            "El marcador quedará visible después de seleccionar la ubicación."
        )

    # ==================================================
    # NO REGISTRAR UBICACIÓN
    # ==================================================

    elif metodo_ubicacion == "No registrar ubicación":
        st.session_state.latitud_registro = ""
        st.session_state.longitud_registro = ""
        st.session_state.direccion_encontrada_mapa = ""

        st.warning("No se registrará ubicación para este registro.")

    # ==================================================
    # MOSTRAR MAPA SIEMPRE
    # ==================================================

    st.markdown("### Vista del punto en el mapa")

    mostrar_mapa_registro_interactivo(
        provincia=provincia,
        tipo_mapa=tipo_mapa_registro,
        lugar=lugar,
        delegacion=delegacion,
        programa=programa,
        key="mapa_registro_visible_siempre"
    )

    # ==================================================
    # MOSTRAR COORDENADAS REGISTRADAS
    # ==================================================

    col_geo1, col_geo2 = st.columns(2)

    with col_geo1:
        st.text_input(
            "Latitud registrada",
            value=st.session_state.latitud_registro,
            disabled=True,
            key="latitud_registrada_visible"
        )

    with col_geo2:
        st.text_input(
            "Longitud registrada",
            value=st.session_state.longitud_registro,
            disabled=True,
            key="longitud_registrada_visible"
        )

    if st.session_state.get("direccion_encontrada_mapa", ""):
        st.info(
            f"Referencia del punto: {st.session_state.direccion_encontrada_mapa}"
        )

    return ""


# ======================================================
# PREPARAR DATAFRAME PARA MAPA GENERAL
# ======================================================

def preparar_dataframe_mapa(df):
    df_mapa = df.copy()

    if "Latitud" not in df_mapa.columns:
        df_mapa["Latitud"] = ""

    if "Longitud" not in df_mapa.columns:
        df_mapa["Longitud"] = ""

    df_mapa["Latitud_Num"] = df_mapa["Latitud"].apply(limpiar_coordenada)
    df_mapa["Longitud_Num"] = df_mapa["Longitud"].apply(limpiar_coordenada)

    df_mapa = df_mapa.dropna(
        subset=[
            "Latitud_Num",
            "Longitud_Num"
        ]
    )

    return df_mapa


# ======================================================
# CENTRAR MAPA GENERAL POR PROVINCIA
# ======================================================

def obtener_centro_mapa_por_provincia(df):
    if df.empty or "Provincia" not in df.columns:
        return [9.7489, -83.7534], 7

    provincias = (
        df["Provincia"]
        .dropna()
        .astype(str)
        .unique()
        .tolist()
    )

    provincias = [
        p for p in provincias
        if p.strip() != ""
    ]

    if len(provincias) == 1:
        provincia = provincias[0]

        if provincia in CENTROS_PROVINCIA:
            return CENTROS_PROVINCIA[provincia], 10

    return [9.7489, -83.7534], 7


# ======================================================
# CREAR MAPA GENERAL DE REGISTROS
# ======================================================

def crear_mapa_registros(df, zoom_start=8):
    df_mapa = preparar_dataframe_mapa(df)

    if df_mapa.empty:
        centro, zoom = obtener_centro_mapa_por_provincia(df)

        mapa = folium.Map(
            location=centro,
            zoom_start=zoom,
            tiles="OpenStreetMap"
        )

        return mapa

    total_puntos = len(df_mapa)

    if total_puntos == 1:
        centro_lat = float(df_mapa.iloc[0]["Latitud_Num"])
        centro_lon = float(df_mapa.iloc[0]["Longitud_Num"])
        zoom_mapa = 16
    else:
        centro_lat = float(df_mapa["Latitud_Num"].mean())
        centro_lon = float(df_mapa["Longitud_Num"].mean())
        zoom_mapa = zoom_start

    mapa = folium.Map(
        location=[centro_lat, centro_lon],
        zoom_start=zoom_mapa,
        tiles="OpenStreetMap"
    )

    coordenadas = []

    for _, row in df_mapa.iterrows():
        programa = str(row.get("Programa", ""))
        color = obtener_color_programa(programa)

        lat = float(row["Latitud_Num"])
        lon = float(row["Longitud_Num"])

        coordenadas.append([lat, lon])

        popup_html = f"""
        <div style="font-family: Arial; width: 280px;">
            <h4 style="color:#002B7F; margin-bottom:6px;">
                Registro PUMI #{row.get("ID", "")}
            </h4>
            <b>Fecha:</b> {row.get("Fecha Actividad", "")}<br>
            <b>Hora:</b> {row.get("Hora Actividad", "")}<br>
            <b>Dirección Regional:</b> {row.get("Dirección Regional", "")}<br>
            <b>Delegación:</b> {row.get("Delegación", "")}<br>
            <b>Programa:</b> {row.get("Programa", "")}<br>
            <b>Actividad:</b> {row.get("Actividad", "")}<br>
            <b>Provincia:</b> {row.get("Provincia", "")}<br>
            <b>Cantón:</b> {row.get("Cantón", "")}<br>
            <b>Distrito:</b> {row.get("Distrito", "")}<br>
            <b>Lugar:</b> {row.get("Lugar", "")}<br>
            <b>Participantes:</b> {row.get("Cantidad Participantes", "")}<br>
            <b>Hombres:</b> {row.get("Cantidad Hombres", "")}<br>
            <b>Mujeres:</b> {row.get("Cantidad Mujeres", "")}<br>
            <b>Referencia:</b> {row.get("Número de Referencia", "")}<br>
            <b>Expediente:</b> {row.get("Número de Expediente Referencia", "")}<br>
        </div>
        """

        folium.Marker(
            location=[lat, lon],
            popup=folium.Popup(popup_html, max_width=330),
            tooltip=f'{row.get("Programa", "")} - {row.get("Delegación", "")}',
            icon=folium.Icon(
                color=color,
                icon="info-sign"
            )
        ).add_to(mapa)

    if total_puntos == 1:
        folium.Circle(
            location=coordenadas[0],
            radius=1200,
            color=COLOR_AZUL,
            fill=True,
            fill_color=COLOR_DORADO,
            fill_opacity=0.10,
            weight=2,
            popup="Área de referencia del registro filtrado"
        ).add_to(mapa)

    elif total_puntos > 1:
        min_lat = df_mapa["Latitud_Num"].min()
        max_lat = df_mapa["Latitud_Num"].max()
        min_lon = df_mapa["Longitud_Num"].min()
        max_lon = df_mapa["Longitud_Num"].max()

        margen_lat = max((max_lat - min_lat) * 0.25, 0.02)
        margen_lon = max((max_lon - min_lon) * 0.25, 0.02)

        bounds = [
            [min_lat - margen_lat, min_lon - margen_lon],
            [max_lat + margen_lat, max_lon + margen_lon]
        ]

        mapa.fit_bounds(bounds)

        radio = max(
            1500,
            int(max(max_lat - min_lat, max_lon - min_lon) * 111000 / 2)
        )

        folium.Circle(
            location=[centro_lat, centro_lon],
            radius=radio,
            color=COLOR_AZUL,
            fill=True,
            fill_color=COLOR_DORADO,
            fill_opacity=0.08,
            weight=2,
            popup=f"Área aproximada de {total_puntos} registros filtrados"
        ).add_to(mapa)

    return mapa


# ======================================================
# MOSTRAR MAPA GENERAL DE REGISTROS
# ======================================================

def mostrar_mapa_registros(df, height=560, key="mapa_registros"):
    df_mapa = preparar_dataframe_mapa(df)

    if df_mapa.empty:
        st.info(
            "No hay registros con coordenadas para mostrar. "
            "Se mostrará la ubicación aproximada según provincia si está disponible."
        )

    mapa = crear_mapa_registros(df)

    st_folium(
        mapa,
        height=height,
        use_container_width=True,
        key=key
    )
# ======================================================
# PARTE 6 DE 10
# EXPORTACIÓN A EXCEL CON FORMATO INSTITUCIONAL,
# MARCA DE AGUA VISIBLE BLOQUEADA, PROTECCIÓN TOTAL
# Y DESCARGA RÁPIDA
# ======================================================


def convertir_excel(df):
    output = BytesIO()

    df_exportar = df.copy()

    columnas_a_eliminar = [
        "Fecha Registro",
        "Plan Estratégico Relacionado",
        "Dirección Mapa"
    ]

    for col in columnas_a_eliminar:
        if col in df_exportar.columns:
            df_exportar = df_exportar.drop(columns=[col])

    encabezados_excel = [
        col for col in ENCABEZADOS
        if col != "Dirección Mapa"
    ]

    for col in encabezados_excel:
        if col not in df_exportar.columns:
            df_exportar[col] = ""

    df_exportar = df_exportar[encabezados_excel]

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_exportar.to_excel(
            writer,
            index=False,
            sheet_name=NOMBRE_HOJA_EXCEL
        )

    output.seek(0)

    workbook = load_workbook(output)
    worksheet = workbook[NOMBRE_HOJA_EXCEL]

    # ==================================================
    # MARCA DE AGUA VISIBLE EN LA HOJA
    # Archivo requerido:
    # marca_agua.png
    # ==================================================

    if os.path.exists(MARCA_AGUA_EXCEL):
        try:
            marca_agua = Image(MARCA_AGUA_EXCEL)
            marca_agua.width = 900
            marca_agua.height = 520
            worksheet.add_image(marca_agua, "A3")
        except Exception:
            pass

    # ==================================================
    # FORMATO DEL ENCABEZADO
    # ==================================================

    fill_header = PatternFill(
        start_color="002B7F",
        end_color="002B7F",
        fill_type="solid"
    )

    font_header = Font(
        color="FFFFFF",
        bold=True
    )

    border = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF")
    )

    for cell in worksheet[1]:
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        cell.border = border

    # ==================================================
    # FORMATO DEL CUERPO
    # ==================================================

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )
            cell.border = border

    # ==================================================
    # AJUSTE DE COLUMNAS
    # ==================================================

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            try:
                valor = str(cell.value) if cell.value is not None else ""

                if len(valor) > max_length:
                    max_length = len(valor)
            except Exception:
                pass

        adjusted_width = min(
            max(max_length + 2, 14),
            45
        )

        worksheet.column_dimensions[column_letter].width = adjusted_width

    # ==================================================
    # FIJAR ENCABEZADO
    # ==================================================

    worksheet.freeze_panes = "A2"

    # ==================================================
    # MARCA DE AGUA PARA IMPRESIÓN / PDF
    # ==================================================

    worksheet.oddHeader.center.text = "Dirección de Programas Preventivos Policiales"
    worksheet.oddHeader.center.size = 24
    worksheet.oddHeader.center.font = "Arial,Bold"
    worksheet.oddHeader.center.color = "D9D9D9"

    worksheet.evenHeader.center.text = "Dirección de Programas Preventivos Policiales"
    worksheet.evenHeader.center.size = 24
    worksheet.evenHeader.center.font = "Arial,Bold"
    worksheet.evenHeader.center.color = "D9D9D9"

    # ==================================================
    # CONFIGURACIÓN DE IMPRESIÓN
    # ==================================================

    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True

    worksheet.page_margins.left = 0.25
    worksheet.page_margins.right = 0.25
    worksheet.page_margins.top = 0.75
    worksheet.page_margins.bottom = 0.75

    # ==================================================
    # BLOQUEAR TODAS LAS CELDAS
    # ==================================================

    for fila in worksheet.iter_rows():
        for celda in fila:
            celda.protection = Protection(
                locked=True
            )

    # ==================================================
    # PROTEGER HOJA COMPLETA
    # Contraseña: DPPP23
    # ==================================================

    worksheet.protection.sheet = True
    worksheet.protection.password = "DPPP23"

    # Bloquear objetos, imágenes y escenarios
    worksheet.protection.objects = True
    worksheet.protection.scenarios = True

    worksheet.protection.formatCells = False
    worksheet.protection.formatColumns = False
    worksheet.protection.formatRows = False
    worksheet.protection.insertColumns = False
    worksheet.protection.insertRows = False
    worksheet.protection.insertHyperlinks = False
    worksheet.protection.deleteColumns = False
    worksheet.protection.deleteRows = False
    worksheet.protection.sort = False
    worksheet.protection.autoFilter = False
    worksheet.protection.pivotTables = False
    worksheet.protection.selectLockedCells = True
    worksheet.protection.selectUnlockedCells = False

    # ==================================================
    # PROTEGER ESTRUCTURA DEL LIBRO
    # Contraseña: DPPP23
    # ==================================================

    workbook.security.lockStructure = True
    workbook.security.workbookPassword = "DPPP23"

    final_output = BytesIO()
    workbook.save(final_output)
    final_output.seek(0)

    return final_output.getvalue()


def boton_descargar_excel(
    df,
    texto_boton="⬇️ Descargar Excel",
    filtrado=False,
    key=None
):
    if df is None or df.empty:
        st.info("No hay registros para descargar.")
        return

    nombre_archivo = generar_nombre_excel_desde_df(
        df,
        filtrado=filtrado
    )

    st.download_button(
        texto_boton,
        data=convertir_excel(df),
        file_name=nombre_archivo,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=key,
        on_click="ignore"
    )


def boton_descargar_excel_formulario(
    df,
    direccion_regional,
    delegacion,
    texto_boton="⬇️ Descargar Excel actualizado",
    key=None
):
    if df is None or df.empty:
        st.info("No hay registros para descargar.")
        return

    nombre_archivo = generar_nombre_excel(
        direccion_regional=direccion_regional,
        delegacion=delegacion,
        filtrado=False
    )

    st.download_button(
        texto_boton,
        data=convertir_excel(df),
        file_name=nombre_archivo,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=key,
        on_click="ignore"
    )
    
# ======================================================
# LOGIN DESDE EXCEL USUARIOS Y CLAVES PUMI
# La app toma usuarios y claves desde el archivo:
# Usuarios y Claves PUMI.xlsx
#
# Puede ingresar con:
# - Delegación: usuario D27 o D27 Alajuela Norte y su clave de la columna Claves.
# - Región: usuario DR2 o Dirección Regional 2 - Alajuela y su clave de la columna Claves.1.
# - Programa: usuario PSCC, DARE, GMD o nombre del programa y su clave de la columna CLAVES.
# ======================================================


def extraer_numero_delegacion(nombre_delegacion):
    match = re.search(r"\bD\s*([0-9]+)\b", str(nombre_delegacion).upper())
    return match.group(1) if match else ""


def obtener_region_de_delegacion(delegacion):
    df_metas = cargar_metas_regionales()
    if df_metas.empty or not delegacion:
        return ""

    deleg_norm = normalizar_texto(delegacion)
    fila = df_metas[df_metas["Clave Delegación"] == deleg_norm]

    if fila.empty:
        numero = extraer_numero_delegacion(delegacion)
        if numero:
            fila = df_metas[
                df_metas["Delegación"].astype(str).apply(extraer_numero_delegacion) == numero
            ]

    if fila.empty:
        return ""

    return str(fila.iloc[0].get("Dirección Regional", "")).strip()


def construir_usuarios_desde_excel():
    usuarios = {
        "ADMIN": {
            "clave": "ADMIN2026",
            "rol": "ADMIN",
            "region": "",
            "delegacion": "",
            "programa": ""
        }
    }

    df = cargar_catalogo_usuarios_claves()

    if df.empty:
        return usuarios

    # Usuarios por delegación.
    for _, row in df.iterrows():
        delegacion = str(row.get("Delegaciones", "")).strip()
        clave = str(row.get("Claves", "")).strip()

        if delegacion and clave:
            numero = extraer_numero_delegacion(delegacion)
            usuario_codigo = f"D{numero}" if numero else delegacion
            region = obtener_region_de_delegacion(delegacion)

            datos = {
                "clave": clave,
                "rol": "DELEGACION",
                "region": region,
                "delegacion": delegacion,
                "programa": ""
            }

            usuarios[normalizar_texto(usuario_codigo)] = datos
            usuarios[normalizar_texto(delegacion)] = datos

    # Usuarios regionales.
    for _, row in df.iterrows():
        region = str(row.get("Regiones", "")).strip()
        clave = str(row.get("Claves.1", "")).strip()

        if region and clave:
            codigo_region = obtener_codigo_region_desde_texto(clave) or obtener_codigo_region_desde_texto(region) or clave
            datos = {
                "clave": clave,
                "rol": "REGIONAL",
                "region": region,
                "delegacion": "",
                "programa": ""
            }

            usuarios[normalizar_texto(codigo_region)] = datos
            usuarios[normalizar_texto(region)] = datos

    # Usuarios por programa.
    for _, row in df.iterrows():
        programa = str(row.get("Programas", "")).strip()
        clave = str(row.get("CLAVES", "")).strip()

        if programa and clave:
            datos = {
                "clave": clave,
                "rol": "PROGRAMA",
                "region": "",
                "delegacion": "",
                "programa": programa
            }

            usuarios[normalizar_texto(programa)] = datos
            # También permite ingresar con la clave corta como usuario, por ejemplo GMD.
            usuario_corto = clave.replace("23", "").strip()
            if usuario_corto:
                usuarios[normalizar_texto(usuario_corto)] = datos

    return usuarios


def pantalla_login():
    mostrar_encabezado_institucional()
    mostrar_titulo_principal()

    st.markdown(
        """
        <div style="text-align:center; margin-top:25px; margin-bottom:20px;">
            <h1 style="margin-bottom:6px;">Ingreso al sistema</h1>
        </div>
        """,
        unsafe_allow_html=True
    )

    col_login_1, col_login_2, col_login_3 = st.columns([1, 1.15, 1])

    with col_login_2:
        usuario_login = st.text_input("Usuario").strip()
        clave_login = st.text_input("Clave", type="password").strip()

        if st.button("Ingresar"):
            usuarios = construir_usuarios_desde_excel()
            datos_usuario = usuarios.get(normalizar_texto(usuario_login))

            if datos_usuario and clave_login == datos_usuario.get("clave", ""):
                st.session_state.usuario_autenticado = {
                    "usuario": usuario_login.upper(),
                    **datos_usuario
                }
                st.rerun()
            else:
                st.error("Usuario o clave incorrectos.")


if st.session_state.usuario_autenticado is None:
    pantalla_login()
    st.stop()

usuario_actual = st.session_state.usuario_autenticado
rol_actual = usuario_actual.get("rol", "")
es_admin = rol_actual == "ADMIN"
es_regional = rol_actual == "REGIONAL"
es_delegacion = rol_actual == "DELEGACION"
es_programa = rol_actual == "PROGRAMA"
puede_ver_todo = es_admin or es_programa

# ======================================================
# PARTE 7 DE 10
# SIDEBAR, CARGA DE EXCEL EXISTENTE, MENÚ PRINCIPAL
# Y ENCABEZADOS VISUALES DE LA APP
# ======================================================


# ======================================================
# MOSTRAR LOGO EN SIDEBAR
# ======================================================

mostrar_logo()


# ======================================================
# TÍTULO DEL SIDEBAR
# ======================================================

st.sidebar.markdown("## Sistema PUMI 2026")

st.sidebar.markdown(
    f"""
    <div style="
        background:#FFFFFF;
        padding:14px;
        border-radius:14px;
        border-left:6px solid #B88A2A;
        box-shadow:0px 4px 12px rgba(0,0,0,0.10);
        margin-bottom:15px;
        font-size:14px;
    ">
    <b>Usuario:</b> {usuario_actual.get('usuario', '')}<br>
    <b>Perfil:</b> {usuario_actual.get('rol', '')}<br>
    <b>Región:</b> {usuario_actual.get('region', 'Todas') if usuario_actual.get('region', '') else 'Todas'}<br>
    <b>Delegación:</b> {usuario_actual.get('delegacion', 'Todas') if usuario_actual.get('delegacion', '') else 'Todas'}
    </div>
    """,
    unsafe_allow_html=True
)

if st.sidebar.button("Cerrar sesión"):
    st.session_state.usuario_autenticado = None
    st.session_state.registros_pumi = pd.DataFrame(columns=ENCABEZADOS)
    st.rerun()


# ======================================================
# CARGA DE EXCEL EXISTENTE
# Permite subir un Excel generado antes para continuar
# agregando, editando o eliminando registros.
# ======================================================

archivo_excel_cargado = st.sidebar.file_uploader(
    "Subir Excel existente para continuar registros",
    type=["xlsx"]
)

if archivo_excel_cargado is not None:
    if st.session_state.archivo_cargado_nombre != archivo_excel_cargado.name:
        df_cargado = cargar_excel_existente(archivo_excel_cargado)

        if not df_cargado.empty:
            st.session_state.registros_pumi = df_cargado
            st.session_state.archivo_cargado_nombre = archivo_excel_cargado.name
            st.sidebar.success("Excel cargado correctamente.")
        else:
            st.sidebar.warning("El Excel cargado no contiene registros válidos.")


# ======================================================
# MENÚ PRINCIPAL
# ======================================================

if "menu_principal" not in st.session_state:
    st.session_state.menu_principal = "Inicio"

menu = st.sidebar.radio(
    "Menú principal",
    [
        "Inicio",
        "Registrar actividad",
        "Editar y eliminar",
        "Dashboard"
    ],
    key="menu_principal"
)


# ======================================================
# ENCABEZADO INSTITUCIONAL SUPERIOR
# ======================================================

mostrar_encabezado_institucional()


# ======================================================
# TÍTULO PRINCIPAL
# ======================================================

mostrar_titulo_principal()
# ======================================================
# PARTE 8 DE 10
# PANTALLA DE INICIO Y PRIMERA MITAD DEL FORMULARIO:
# DATOS GENERALES, ACTIVIDAD Y PARTICIPANTES
# ======================================================


# ======================================================
# INICIO
# ======================================================

if menu == "Inicio":

    st.markdown(
        f"""
        <div class="card-pumi">
            <div class="subtitulo-pumi">
                Bienvenido al Sistema PUMI 2026
            </div>
            <div class="texto-pumi" style="text-align:center;">
                Seleccione una opción para continuar con el registro, revisión o consulta de avances.
                <br><br>
                <b>Delegación activa:</b> {usuario_actual.get('delegacion', 'Todas') if usuario_actual.get('delegacion', '') else 'Todas'}<br>
                <b>Dirección Regional:</b> {usuario_actual.get('region', 'Todas') if usuario_actual.get('region', '') else 'Todas'}
            </div>
        </div>
        """,
        unsafe_allow_html=True
    )

    st.markdown("### Acceso rápido")

    c1, c2, c3 = st.columns(3)

    with c1:
        if st.button("📝 Registrar actividad", use_container_width=True):
            st.session_state.menu_principal = "Registrar actividad"
            st.rerun()

    with c2:
        if st.button("✏️ Editar y eliminar", use_container_width=True):
            st.session_state.menu_principal = "Editar y eliminar"
            st.rerun()

    with c3:
        if st.button("📊 Dashboard", use_container_width=True):
            st.session_state.menu_principal = "Dashboard"
            st.rerun()

    df_actual = st.session_state.registros_pumi

    if not df_actual.empty:
        st.markdown("---")
        boton_descargar_excel(
            df_actual,
            texto_boton="⬇️ Descargar Excel actual",
            filtrado=False,
            key="descarga_inicio"
        )


# ======================================================
# REGISTRAR ACTIVIDAD
# PRIMERA MITAD DEL FORMULARIO
# ======================================================

elif menu == "Registrar actividad":

    st.markdown("## Registrar nueva actividad preventiva")

    st.markdown(
        """
        <div class="card-azul">
            <div class="texto-pumi">
                Complete la información de la actividad. Cada registro se agregará
                a la tabla temporal y luego podrá descargarse en Excel.
            </div>
        </div>
        """,
        unsafe_allow_html=True
    )


    # ==================================================
    # DATOS GENERALES DEL REGISTRO
    # ==================================================

    st.markdown(
        """
        <div class="bloque-datos">
            <b>Datos generales del registro</b>
        </div>
        """,
        unsafe_allow_html=True
    )

    usar_metas = metas_disponibles()

    col1, col2 = st.columns(2)

    with col1:
        fecha_actividad = st.date_input(
            "Fecha de la actividad",
            value=date.today(),
            format="DD/MM/YYYY"
        )

        hora_actividad = st.time_input(
            "Hora de la actividad",
            value=time(8, 0)
        )

        regiones_lista = obtener_regiones_datos()

        if es_admin or es_programa:
            direccion_regional = st.selectbox(
                "Dirección Regional",
                regiones_lista if regiones_lista else obtener_regiones_metas() or REGIONES
            )
        else:
            direccion_regional = st.selectbox(
                "Dirección Regional",
                [usuario_actual.get("region", "")],
                disabled=True
            )

        st.session_state.ultima_direccion_regional = direccion_regional

        delegaciones_filtradas = obtener_delegaciones_por_region(direccion_regional)

        if es_delegacion:
            delegacion = st.selectbox(
                "Delegación",
                [usuario_actual.get("delegacion", "")],
                disabled=True
            )
        else:
            delegacion = st.selectbox(
                "Delegación",
                delegaciones_filtradas if delegaciones_filtradas else obtener_delegaciones_metas_por_region(direccion_regional) or obtener_delegaciones_unicas()
            )

        st.session_state.ultima_delegacion = delegacion

    with col2:
        responde_a = ""

        if usar_metas:
            programas_lista = obtener_programas_metas_por_delegacion(delegacion)

            programa = st.selectbox(
                "Programa",
                programas_lista if programas_lista else ["Sin datos disponibles"]
            )

            actividades_filtradas = obtener_actividades_metas_por_delegacion_programa(
                delegacion,
                programa
            )

            actividad = st.selectbox(
                "Actividad oficial según meta",
                actividades_filtradas if actividades_filtradas else ["Sin datos disponibles"]
            )

            fila_meta_seleccionada = obtener_fila_meta_por_delegacion(
                delegacion,
                programa,
                actividad
            )

        else:
            programas_lista = obtener_programas_datos()

            programa = st.selectbox(
                "Programa",
                programas_lista if programas_lista else ["Sin datos disponibles"]
            )

            actividades_filtradas = obtener_actividades_por_programa(programa)

            actividad = st.selectbox(
                "Actividad realizada",
                actividades_filtradas if actividades_filtradas else ["Sin datos disponibles"]
            )

            fila_meta_seleccionada = None

    if usar_metas:
        avance_sugerido, meta_oficial, avance_base_oficial, pendiente_oficial = mostrar_resumen_meta_seleccionada(fila_meta_seleccionada)
    else:
        avance_sugerido = 1
        pendiente_oficial = 0

    st.markdown(
        """
        <div class="bloque-avance-registro">
            <b>Registro del avance</b>
        </div>
        """,
        unsafe_allow_html=True
    )

    col_av1, col_av2, col_av3 = st.columns([1, 1, 1])

    max_avance_registro = int(pendiente_oficial) if usar_metas and pendiente_oficial > 0 else None

    with col_av1:
        if max_avance_registro:
            avance_realizado = st.number_input(
                "Avance realizado",
                min_value=0,
                max_value=max_avance_registro,
                value=min(avance_sugerido, max_avance_registro),
                step=1
            )
        else:
            avance_realizado = st.number_input(
                "Avance realizado",
                min_value=0,
                value=avance_sugerido,
                step=1
            )

    with col_av2:
        responsable = st.text_input("Funcionario responsable")

    with col_av3:
        usuario = st.text_input("Usuario que registra")


    # ==================================================
    # PARTICIPANTES
    # ==================================================

    st.markdown(
        """
        <div class="bloque-datos">
            <b>Participación registrada</b>
        </div>
        """,
        unsafe_allow_html=True
    )

    colp1, colp2, colp3 = st.columns(3)

    with colp1:
        cantidad = st.number_input(
            "Cantidad total de participantes",
            min_value=0,
            step=1
        )

    with colp2:
        cantidad_hombres = st.number_input(
            "Cantidad de hombres",
            min_value=0,
            step=1
        )

    with colp3:
        cantidad_mujeres = st.number_input(
            "Cantidad de mujeres",
            min_value=0,
            step=1
        )

    st.markdown("### Rangos de edad")

    colr1, colr2, colr3, colr4 = st.columns(4)

    with colr1:
        edad_10_18 = st.number_input(
            "Edad 10 a 18",
            min_value=0,
            step=1
        )

    with colr2:
        edad_19_30 = st.number_input(
            "Edad 19 a 30",
            min_value=0,
            step=1
        )

    with colr3:
        edad_31_45 = st.number_input(
            "Edad 31 a 45",
            min_value=0,
            step=1
        )

    with colr4:
        edad_46_mas = st.number_input(
            "Edad 46 en adelante",
            min_value=0,
            step=1
        )

    suma_sexo = cantidad_hombres + cantidad_mujeres

    suma_edades = (
        edad_10_18 +
        edad_19_30 +
        edad_31_45 +
        edad_46_mas
    )

    if cantidad > 0:
        if suma_sexo != cantidad:
            st.warning(
                f"La suma de hombres y mujeres ({suma_sexo}) no coincide "
                f"con la cantidad total de participantes ({cantidad})."
            )

        if suma_edades != cantidad:
            st.warning(
                f"La suma de los rangos de edad ({suma_edades}) no coincide "
                f"con la cantidad total de participantes ({cantidad})."
            )
# ======================================================
# PARTE 9 DE 10
# SEGUNDA MITAD DEL FORMULARIO:
# UBICACIÓN TERRITORIAL, LUGAR, GEOREFERENCIA,
# INFORMACIÓN COMPLEMENTARIA, GUARDAR Y DESCARGAR
# ======================================================


    # ==================================================
    # UBICACIÓN TERRITORIAL
    # ==================================================

    st.markdown(
        """
        <div class="bloque-territorio">
            <b>Ubicación territorial</b>
        </div>
        """,
        unsafe_allow_html=True
    )

    provincias_lista = obtener_provincias_datos()

    col3, col4, col5 = st.columns(3)

    with col3:
        provincia = st.selectbox(
            "Provincia",
            provincias_lista if provincias_lista else PROVINCIAS
        )

    with col4:
        cantones_filtrados = obtener_cantones_por_provincia(provincia)

        canton = st.selectbox(
            "Cantón",
            cantones_filtrados
            if cantones_filtrados
            else ["Sin datos disponibles"]
        )

    with col5:
        distritos_filtrados = obtener_distritos_por_provincia_canton(
            provincia,
            canton
        )

        distrito = st.selectbox(
            "Distrito",
            distritos_filtrados
            if distritos_filtrados
            else ["Sin datos disponibles"]
        )


    # ==================================================
    # LUGAR DE REALIZACIÓN
    # ==================================================

    st.markdown(
        """
        <div class="bloque-actividad">
            <b>Lugar de realización de la actividad</b>
        </div>
        """,
        unsafe_allow_html=True
    )

    tipo_lugar = st.radio(
        "Seleccione el tipo de lugar",
        ["Centro educativo", "Otro lugar"],
        horizontal=True
    )

    lugar = ""
    centro_educativo = ""
    codigo_presupuestario = ""

    if tipo_lugar == "Centro educativo":
        centros = obtener_centros_por_provincia(provincia)

        if centros:
            centro_mostrar = st.selectbox(
                "Centro educativo según base MEP 2025",
                centros
            )

            centro_educativo, codigo_presupuestario = obtener_datos_centro_educativo(
                centro_mostrar
            )

            st.info(f"Centro educativo: {centro_educativo}")
            st.info(
                "Código presupuestario: "
                f"{codigo_presupuestario if codigo_presupuestario else 'No disponible'}"
            )

            lugar = centro_educativo

        else:
            st.warning(
                "No se encontraron centros educativos para la provincia seleccionada."
            )

            centro_educativo = st.text_input(
                "Digite el centro educativo manualmente"
            )

            codigo_presupuestario = st.text_input(
                "Código presupuestario"
            )

            lugar = centro_educativo

    else:
        lugar = st.text_input("Lugar donde se realizó la actividad")
        centro_educativo = ""
        codigo_presupuestario = ""


    # ==================================================
    # MAPA Y GEOREFERENCIA
    # La función ya no devuelve ni guarda Dirección Mapa.
    # Esa columna queda siempre vacía y no sale en Excel.
    # ==================================================

    bloque_georreferencia_formulario(
        provincia=provincia,
        canton=canton,
        distrito=distrito,
        lugar=lugar,
        delegacion=delegacion,
        programa=programa
    )


    # ==================================================
    # INFORMACIÓN COMPLEMENTARIA
    # ==================================================

    st.markdown(
        """
        <div class="bloque-datos">
            <b>Información complementaria</b>
        </div>
        """,
        unsafe_allow_html=True
    )

    instituciones_participantes = st.text_area(
        "Instituciones participantes"
    )

    numero_referencia = st.text_input(
        "Número de referencia"
    )

    numero_expediente = st.text_input(
        "Número de expediente referencia"
    )

    observaciones = st.text_area(
        "Observaciones"
    )


    # ==================================================
    # GUARDAR REGISTRO
    # ==================================================

    st.markdown("---")

    if st.button("💾 Agregar registro al Excel temporal"):

        nuevo_id = generar_id_consecutivo()

        registro = {
            "ID": nuevo_id,
            "Fecha Actividad": fecha_actividad.strftime("%d/%m/%Y"),
            "Hora Actividad": hora_actividad.strftime("%H:%M"),
            "Dirección Regional": direccion_regional,
            "Delegación": delegacion,
            "Responde a": "",
            "Programa": programa,
            "Actividad": actividad,
            "Provincia": provincia,
            "Cantón": canton,
            "Distrito": distrito,
            "Tipo Lugar": tipo_lugar,
            "Lugar": lugar,
            "Centro Educativo": centro_educativo,
            "Código Presupuestario": codigo_presupuestario,
            "Dirección Mapa": "",
            "Latitud": st.session_state.latitud_registro,
            "Longitud": st.session_state.longitud_registro,
            "Responsable": responsable,
            "Avance Realizado": avance_realizado,
            "Cantidad Participantes": cantidad,
            "Cantidad Hombres": cantidad_hombres,
            "Cantidad Mujeres": cantidad_mujeres,
            "Edad 10 a 18": edad_10_18,
            "Edad 19 a 30": edad_19_30,
            "Edad 31 a 45": edad_31_45,
            "Edad 46 en adelante": edad_46_mas,
            "Instituciones Participantes": instituciones_participantes,
            "Número de Referencia": numero_referencia,
            "Número de Expediente Referencia": numero_expediente,
            "Observaciones": observaciones,
            "Usuario Registra": usuario
        }

        agregar_registro_session(registro)

        st.success(
            "Registro agregado correctamente. Puede seguir agregando más registros "
            "o descargar el Excel desde esta misma pantalla."
        )

        st.session_state.latitud_registro = ""
        st.session_state.longitud_registro = ""
        st.session_state.direccion_encontrada_mapa = ""

        st.rerun()


    # ==================================================
    # DESCARGA RÁPIDA DESDE EL FORMULARIO
    # ==================================================

    df_actual = st.session_state.registros_pumi

    if not df_actual.empty:
        st.markdown("### Descargar registros actuales")

        boton_descargar_excel_formulario(
            df=df_actual,
            direccion_regional=direccion_regional,
            delegacion=delegacion,
            texto_boton="⬇️ Descargar Excel actualizado",
            key="descarga_formulario"
        )

        st.markdown("### Vista previa de registros cargados")

        st.dataframe(
            df_actual,
            use_container_width=True,
            hide_index=True
        )
# ======================================================
# PARTE 10 DE 10
# REGISTROS CARGADOS, EDICIÓN, ELIMINACIÓN,
# DASHBOARD, MAPA GENERAL Y DESCARGA GLOBAL
# ======================================================


elif menu == "Editar y eliminar":

    st.markdown("## Editar / eliminar registros PUMI")

    df_actual = st.session_state.registros_pumi.copy()

    if es_delegacion and not df_actual.empty:
        df_actual = df_actual[
            df_actual["Delegación"].apply(normalizar_texto) == normalizar_texto(usuario_actual.get("delegacion", ""))
        ].copy()

    if df_actual.empty:
        st.info("Aún no hay registros para editar o eliminar.")
    else:
        st.markdown(
            """
            <div class="card-dorado">
                <div class="texto-pumi">
                    En esta sección puede revisar los registros agregados,
                    eliminar registros específicos, editar información y descargar
                    el Excel actualizado.
                </div>
            </div>
            """,
            unsafe_allow_html=True
        )

        st.dataframe(
            df_actual,
            use_container_width=True,
            hide_index=True
        )

        st.markdown("### Descargar Excel actualizado")

        boton_descargar_excel(
            df_actual,
            texto_boton="⬇️ Descargar Excel",
            filtrado=False,
            key="descarga_registros"
        )

        st.markdown("---")
        st.markdown("### Eliminar registro")

        ids_disponibles = df_actual["ID"].astype(str).tolist()

        id_eliminar = st.selectbox(
            "Seleccione el ID del registro que desea revisar/eliminar",
            ids_disponibles,
            key="id_eliminar_registro"
        )

        fila_eliminar = df_actual[
            df_actual["ID"].astype(str) == str(id_eliminar)
        ]

        if not fila_eliminar.empty:
            registro_eliminar = fila_eliminar.iloc[0].to_dict()

            st.markdown("#### Revisión del registro seleccionado")

            columnas_revision = [
                "ID", "Fecha Actividad", "Hora Actividad", "Dirección Regional",
                "Delegación", "Programa", "Actividad", "Provincia", "Cantón",
                "Distrito", "Lugar", "Avance Realizado", "Cantidad Participantes",
                "Responsable", "Usuario Registra", "Observaciones"
            ]
            columnas_revision = [c for c in columnas_revision if c in fila_eliminar.columns]

            st.dataframe(
                fila_eliminar[columnas_revision],
                use_container_width=True,
                hide_index=True
            )

            confirmar_eliminacion = st.checkbox(
                "Confirmo que revisé el registro y deseo eliminarlo definitivamente de esta sesión",
                key="confirmar_eliminacion_registro"
            )

            if st.button("🗑️ Eliminar registro seleccionado", disabled=not confirmar_eliminacion):
                eliminar_registro_session(id_eliminar)
                st.success("Registro eliminado correctamente.")
                st.rerun()

        st.markdown("---")
        st.markdown("### Editar registro existente")

        id_editar = st.selectbox(
            "Seleccione el ID del registro que desea editar",
            ids_disponibles,
            key="id_editar_registro"
        )

        fila_editar = df_actual[
            df_actual["ID"].astype(str) == str(id_editar)
        ]

        if not fila_editar.empty:
            fila = fila_editar.iloc[0].to_dict()

            with st.expander("Abrir formulario de edición", expanded=False):

                def indice_opcion(lista, valor):
                    valor_norm = normalizar_texto(valor)
                    for i, opcion in enumerate(lista):
                        if normalizar_texto(opcion) == valor_norm:
                            return i
                    return 0

                def preparar_opciones(lista, valor_actual):
                    opciones = [str(x) for x in lista if str(x).strip() != ""]
                    if valor_actual and str(valor_actual).strip() != "" and normalizar_texto(valor_actual) not in [normalizar_texto(x) for x in opciones]:
                        opciones.insert(0, str(valor_actual))
                    if not opciones:
                        opciones = ["Sin datos disponibles"]
                    return opciones

                col1, col2 = st.columns(2)

                with col1:
                    fecha_actual_edit = pd.to_datetime(
                        fila.get("Fecha Actividad", date.today()),
                        errors="coerce",
                        dayfirst=True
                    )
                    if pd.isna(fecha_actual_edit):
                        fecha_actual_edit = pd.Timestamp(date.today())

                    fecha_actividad_edit_dt = st.date_input(
                        "Fecha Actividad",
                        value=fecha_actual_edit.date(),
                        format="DD/MM/YYYY"
                    )
                    fecha_actividad_edit = fecha_actividad_edit_dt.strftime("%d/%m/%Y")

                    hora_actual_texto = str(fila.get("Hora Actividad", "08:00"))
                    try:
                        hora_actual_edit = datetime.strptime(hora_actual_texto[:5], "%H:%M").time()
                    except Exception:
                        hora_actual_edit = time(8, 0)

                    hora_actividad_edit_dt = st.time_input(
                        "Hora Actividad",
                        value=hora_actual_edit
                    )
                    hora_actividad_edit = hora_actividad_edit_dt.strftime("%H:%M")

                    regiones_edit_lista = preparar_opciones(
                        obtener_regiones_datos(),
                        fila.get("Dirección Regional", "")
                    )

                    if es_admin or es_programa:
                        direccion_regional_edit = st.selectbox(
                            "Dirección Regional",
                            regiones_edit_lista,
                            index=indice_opcion(regiones_edit_lista, fila.get("Dirección Regional", "")),
                            key="edit_direccion_regional"
                        )
                    else:
                        direccion_regional_edit = st.selectbox(
                            "Dirección Regional",
                            [usuario_actual.get("region", fila.get("Dirección Regional", ""))],
                            disabled=True,
                            key="edit_direccion_regional"
                        )

                    delegaciones_edit_lista = preparar_opciones(
                        obtener_delegaciones_por_region(direccion_regional_edit),
                        fila.get("Delegación", "")
                    )

                    if es_delegacion:
                        delegacion_edit = st.selectbox(
                            "Delegación",
                            [usuario_actual.get("delegacion", fila.get("Delegación", ""))],
                            disabled=True,
                            key="edit_delegacion"
                        )
                    else:
                        delegacion_edit = st.selectbox(
                            "Delegación",
                            delegaciones_edit_lista,
                            index=indice_opcion(delegaciones_edit_lista, fila.get("Delegación", "")),
                            key="edit_delegacion"
                        )

                    programas_edit_lista = obtener_programas_metas_por_delegacion(delegacion_edit)
                    if not programas_edit_lista:
                        programas_edit_lista = obtener_programas_datos()

                    programas_edit_lista = preparar_opciones(
                        programas_edit_lista,
                        fila.get("Programa", "")
                    )

                    programa_edit = st.selectbox(
                        "Programa",
                        programas_edit_lista,
                        index=indice_opcion(programas_edit_lista, fila.get("Programa", "")),
                        key="edit_programa"
                    )

                    actividades_edit_lista = obtener_actividades_metas_por_delegacion_programa(
                        delegacion_edit,
                        programa_edit
                    )
                    if not actividades_edit_lista:
                        actividades_edit_lista = obtener_actividades_por_programa(programa_edit)

                    actividades_edit_lista = preparar_opciones(
                        actividades_edit_lista,
                        fila.get("Actividad", "")
                    )

                    actividad_edit = st.selectbox(
                        "Actividad",
                        actividades_edit_lista,
                        index=indice_opcion(actividades_edit_lista, fila.get("Actividad", "")),
                        key="edit_actividad"
                    )

                    provincias_edit_lista = preparar_opciones(
                        obtener_provincias_datos(),
                        fila.get("Provincia", "")
                    )
                    provincia_edit = st.selectbox(
                        "Provincia",
                        provincias_edit_lista,
                        index=indice_opcion(provincias_edit_lista, fila.get("Provincia", "")),
                        key="edit_provincia"
                    )

                    cantones_edit_lista = preparar_opciones(
                        obtener_cantones_por_provincia(provincia_edit),
                        fila.get("Cantón", "")
                    )
                    canton_edit = st.selectbox(
                        "Cantón",
                        cantones_edit_lista,
                        index=indice_opcion(cantones_edit_lista, fila.get("Cantón", "")),
                        key="edit_canton"
                    )

                    distritos_edit_lista = preparar_opciones(
                        obtener_distritos_por_provincia_canton(provincia_edit, canton_edit),
                        fila.get("Distrito", "")
                    )
                    distrito_edit = st.selectbox(
                        "Distrito",
                        distritos_edit_lista,
                        index=indice_opcion(distritos_edit_lista, fila.get("Distrito", "")),
                        key="edit_distrito"
                    )

                    tipo_lugar_opciones = preparar_opciones(
                        ["Centro educativo", "Otro lugar"],
                        fila.get("Tipo Lugar", "")
                    )
                    tipo_lugar_edit = st.selectbox(
                        "Tipo Lugar",
                        tipo_lugar_opciones,
                        index=indice_opcion(tipo_lugar_opciones, fila.get("Tipo Lugar", "")),
                        key="edit_tipo_lugar"
                    )

                    lugar_edit = st.text_input(
                        "Lugar",
                        value=str(fila.get("Lugar", ""))
                    )

                    centro_educativo_edit = st.text_input(
                        "Centro Educativo",
                        value=str(fila.get("Centro Educativo", ""))
                    )

                    codigo_presupuestario_edit = st.text_input(
                        "Código Presupuestario",
                        value=str(fila.get("Código Presupuestario", ""))
                    )

                    # ==============================================
                    # Corrección de ubicación en mapa para edición
                    # ==============================================
                    st.markdown("#### Corregir ubicación en mapa")

                    lat_key = f"edit_latitud_{id_editar}"
                    lon_key = f"edit_longitud_{id_editar}"

                    if lat_key not in st.session_state:
                        st.session_state[lat_key] = str(fila.get("Latitud", ""))
                    if lon_key not in st.session_state:
                        st.session_state[lon_key] = str(fila.get("Longitud", ""))

                    tipo_mapa_edit = st.selectbox(
                        "Tipo de mapa para corregir marca",
                        ["OpenStreetMap", "Mapa claro", "Topográfico", "Satélite"],
                        key=f"edit_tipo_mapa_{id_editar}"
                    )

                    lat_actual_edit = limpiar_coordenada(st.session_state.get(lat_key, ""))
                    lon_actual_edit = limpiar_coordenada(st.session_state.get(lon_key, ""))

                    if lat_actual_edit is not None and lon_actual_edit is not None:
                        centro_edit = [lat_actual_edit, lon_actual_edit]
                        zoom_edit = 16
                    else:
                        centro_edit = CENTROS_PROVINCIA.get(provincia_edit, [9.7489, -83.7534])
                        zoom_edit = 10

                    mapa_edit = crear_mapa_base(
                        centro=centro_edit,
                        zoom=zoom_edit,
                        tipo_mapa=tipo_mapa_edit
                    )

                    if lat_actual_edit is not None and lon_actual_edit is not None:
                        folium.Marker(
                            location=[lat_actual_edit, lon_actual_edit],
                            popup="Ubicación actual del registro",
                            tooltip="Marca actual",
                            icon=folium.Icon(color="red", icon="map-marker")
                        ).add_to(mapa_edit)

                        folium.Circle(
                            location=[lat_actual_edit, lon_actual_edit],
                            radius=180,
                            color=COLOR_AZUL,
                            fill=True,
                            fill_color=COLOR_DORADO,
                            fill_opacity=0.18,
                            weight=2
                        ).add_to(mapa_edit)

                    resultado_mapa_edit = st_folium(
                        mapa_edit,
                        height=420,
                        use_container_width=True,
                        key=f"mapa_edicion_{id_editar}"
                    )

                    if resultado_mapa_edit and resultado_mapa_edit.get("last_clicked"):
                        st.session_state[lat_key] = str(resultado_mapa_edit["last_clicked"]["lat"])
                        st.session_state[lon_key] = str(resultado_mapa_edit["last_clicked"]["lng"])
                        st.success("Marca actualizada en el mapa. Revise las coordenadas antes de guardar.")
                        st.rerun()

                    latitud_edit = st.text_input(
                        "Latitud",
                        value=st.session_state.get(lat_key, ""),
                        key=f"input_latitud_edit_{id_editar}"
                    )

                    longitud_edit = st.text_input(
                        "Longitud",
                        value=st.session_state.get(lon_key, ""),
                        key=f"input_longitud_edit_{id_editar}"
                    )

                    st.session_state[lat_key] = latitud_edit
                    st.session_state[lon_key] = longitud_edit

                with col2:
                    responsable_edit = st.text_input(
                        "Responsable",
                        value=str(fila.get("Responsable", ""))
                    )

                    avance_realizado_edit = st.number_input(
                        "Avance Realizado",
                        min_value=0,
                        step=1,
                        value=int(
                            pd.to_numeric(
                                fila.get("Avance Realizado", 1),
                                errors="coerce"
                            ) or 0
                        )
                    )

                    cantidad_edit = st.number_input(
                        "Cantidad Participantes",
                        min_value=0,
                        step=1,
                        value=int(
                            pd.to_numeric(
                                fila.get("Cantidad Participantes", 0),
                                errors="coerce"
                            ) or 0
                        )
                    )

                    hombres_edit = st.number_input(
                        "Cantidad Hombres",
                        min_value=0,
                        step=1,
                        value=int(
                            pd.to_numeric(
                                fila.get("Cantidad Hombres", 0),
                                errors="coerce"
                            ) or 0
                        )
                    )

                    mujeres_edit = st.number_input(
                        "Cantidad Mujeres",
                        min_value=0,
                        step=1,
                        value=int(
                            pd.to_numeric(
                                fila.get("Cantidad Mujeres", 0),
                                errors="coerce"
                            ) or 0
                        )
                    )

                    edad_10_18_edit = st.number_input(
                        "Edad 10 a 18",
                        min_value=0,
                        step=1,
                        value=int(
                            pd.to_numeric(
                                fila.get("Edad 10 a 18", 0),
                                errors="coerce"
                            ) or 0
                        )
                    )

                    edad_19_30_edit = st.number_input(
                        "Edad 19 a 30",
                        min_value=0,
                        step=1,
                        value=int(
                            pd.to_numeric(
                                fila.get("Edad 19 a 30", 0),
                                errors="coerce"
                            ) or 0
                        )
                    )

                    edad_31_45_edit = st.number_input(
                        "Edad 31 a 45",
                        min_value=0,
                        step=1,
                        value=int(
                            pd.to_numeric(
                                fila.get("Edad 31 a 45", 0),
                                errors="coerce"
                            ) or 0
                        )
                    )

                    edad_46_mas_edit = st.number_input(
                        "Edad 46 en adelante",
                        min_value=0,
                        step=1,
                        value=int(
                            pd.to_numeric(
                                fila.get("Edad 46 en adelante", 0),
                                errors="coerce"
                            ) or 0
                        )
                    )

                    instituciones_edit = st.text_area(
                        "Instituciones Participantes",
                        value=str(fila.get("Instituciones Participantes", ""))
                    )

                    referencia_edit = st.text_input(
                        "Número de Referencia",
                        value=str(fila.get("Número de Referencia", ""))
                    )

                    expediente_edit = st.text_input(
                        "Número de Expediente Referencia",
                        value=str(fila.get("Número de Expediente Referencia", ""))
                    )

                    observaciones_edit = st.text_area(
                        "Observaciones",
                        value=str(fila.get("Observaciones", ""))
                    )

                    usuario_edit = st.text_input(
                        "Usuario Registra",
                        value=str(fila.get("Usuario Registra", ""))
                    )

                suma_sexo_edit = hombres_edit + mujeres_edit

                suma_edades_edit = (
                    edad_10_18_edit +
                    edad_19_30_edit +
                    edad_31_45_edit +
                    edad_46_mas_edit
                )

                if cantidad_edit > 0:
                    if suma_sexo_edit != cantidad_edit:
                        st.warning(
                            f"La suma de hombres y mujeres ({suma_sexo_edit}) "
                            f"no coincide con el total ({cantidad_edit})."
                        )

                    if suma_edades_edit != cantidad_edit:
                        st.warning(
                            f"La suma de rangos de edad ({suma_edades_edit}) "
                            f"no coincide con el total ({cantidad_edit})."
                        )

                if st.button("💾 Guardar cambios del registro"):

                    if cantidad_edit > 0 and suma_sexo_edit != cantidad_edit:
                        st.error(
                            "No se puede guardar porque la suma de hombres y mujeres "
                            "no coincide con el total de participantes."
                        )
                        st.stop()

                    if cantidad_edit > 0 and suma_edades_edit != cantidad_edit:
                        st.error(
                            "No se puede guardar porque la suma de rangos de edad "
                            "no coincide con el total de participantes."
                        )
                        st.stop()

                    nuevos_datos = {
                        "ID": id_editar,
                        "Fecha Actividad": fecha_actividad_edit,
                        "Hora Actividad": hora_actividad_edit,
                        "Dirección Regional": direccion_regional_edit,
                        "Delegación": delegacion_edit,
                        "Programa": programa_edit,
                        "Actividad": actividad_edit,
                        "Provincia": provincia_edit,
                        "Cantón": canton_edit,
                        "Distrito": distrito_edit,
                        "Tipo Lugar": tipo_lugar_edit,
                        "Lugar": lugar_edit,
                        "Centro Educativo": centro_educativo_edit,
                        "Código Presupuestario": codigo_presupuestario_edit,
                        "Dirección Mapa": "",
                        "Latitud": latitud_edit,
                        "Longitud": longitud_edit,
                        "Responsable": responsable_edit,
                        "Avance Realizado": avance_realizado_edit,
                        "Cantidad Participantes": cantidad_edit,
                        "Cantidad Hombres": hombres_edit,
                        "Cantidad Mujeres": mujeres_edit,
                        "Edad 10 a 18": edad_10_18_edit,
                        "Edad 19 a 30": edad_19_30_edit,
                        "Edad 31 a 45": edad_31_45_edit,
                        "Edad 46 en adelante": edad_46_mas_edit,
                        "Instituciones Participantes": instituciones_edit,
                        "Número de Referencia": referencia_edit,
                        "Número de Expediente Referencia": expediente_edit,
                        "Observaciones": observaciones_edit,
                        "Usuario Registra": usuario_edit
                    }

                    actualizado = actualizar_registro_session(
                        id_editar,
                        nuevos_datos
                    )

                    if actualizado:
                        st.success("Registro actualizado correctamente.")
                        st.rerun()
                    else:
                        st.error("No se pudo actualizar el registro.")


elif menu == "Dashboard":

    st.markdown("## Dashboard de avances")

    df_actual = st.session_state.registros_pumi.copy()

    # La delegación solo ve sus registros y sus metas.
    delegacion_login = usuario_actual.get("delegacion", "") if es_delegacion else ""

    if es_delegacion and not df_actual.empty:
        df_actual = df_actual[
            df_actual["Delegación"].apply(normalizar_texto) == normalizar_texto(delegacion_login)
        ].copy()

    # Este dataframe SIEMPRE se genera, aunque no existan registros todavía,
    # para que el usuario pueda ver la meta oficial pendiente.
    df_avance = generar_avance_contra_metas(df_actual)

    if es_delegacion and not df_avance.empty:
        df_avance = df_avance[
            df_avance["Delegación"].apply(normalizar_texto) == normalizar_texto(delegacion_login)
        ].copy()

    if es_regional and not df_avance.empty:
        df_avance = df_avance[
            df_avance["Dirección Regional"].apply(normalizar_texto) == normalizar_texto(usuario_actual.get("region", ""))
        ].copy()

    if es_programa and not df_avance.empty:
        df_avance = df_avance[
            df_avance["Programa"].apply(normalizar_texto) == normalizar_texto(usuario_actual.get("programa", ""))
        ].copy()

    if df_avance.empty:
        st.info("No se encontraron metas oficiales para esta delegación.")
    else:
        st.markdown("### Filtros del avance")

        colf1, colf2, colf3 = st.columns(3)

        with colf1:
            programas_dash = sorted(
                df_avance["Programa"]
                .dropna()
                .astype(str)
                .unique()
                .tolist()
            )
            filtro_programa_dash = st.multiselect(
                "Programa",
                programas_dash,
                key="dash_filtro_programa_metas"
            )

        with colf2:
            estados_dash = sorted(
                df_avance["Estado cumplimiento"]
                .dropna()
                .astype(str)
                .unique()
                .tolist()
            )
            filtro_estado_dash = st.multiselect(
                "Estado",
                estados_dash,
                key="dash_filtro_estado_metas"
            )

        with colf3:
            if not es_delegacion:
                delegaciones_dash = sorted(
                    df_avance["Delegación"]
                    .dropna()
                    .astype(str)
                    .unique()
                    .tolist()
                )
                filtro_delegacion_dash = st.multiselect(
                    "Delegación",
                    delegaciones_dash,
                    key="dash_filtro_delegacion_metas"
                )
            else:
                st.text_input(
                    "Delegación",
                    value=delegacion_login,
                    disabled=True,
                    key="dash_delegacion_login_visible"
                )
                filtro_delegacion_dash = []

        df_avance_filtrado = df_avance.copy()

        if filtro_programa_dash:
            df_avance_filtrado = df_avance_filtrado[
                df_avance_filtrado["Programa"].isin(filtro_programa_dash)
            ]

        if filtro_estado_dash:
            df_avance_filtrado = df_avance_filtrado[
                df_avance_filtrado["Estado cumplimiento"].isin(filtro_estado_dash)
            ]

        if not es_delegacion and filtro_delegacion_dash:
            df_avance_filtrado = df_avance_filtrado[
                df_avance_filtrado["Delegación"].isin(filtro_delegacion_dash)
            ]

        st.markdown("### Resumen de cumplimiento")

        meta_total_dash = df_avance_filtrado["Meta 2026"].sum()
        avance_base_dash = df_avance_filtrado["Avance base"].sum()
        realizado_pumi_dash = df_avance_filtrado["Realizado PUMI"].sum()
        realizado_total_dash = df_avance_filtrado["Realizado total"].sum()
        pendiente_total_dash = df_avance_filtrado["Pendiente"].sum()
        porcentaje_total_dash = realizado_total_dash / meta_total_dash if meta_total_dash else 0

        cdm1, cdm2, cdm3, cdm4, cdm5 = st.columns(5)
        cdm1.metric("Meta oficial", f"{meta_total_dash:,.0f}")
        cdm2.metric("Avance base", f"{avance_base_dash:,.0f}")
        cdm3.metric("Registrado PUMI", f"{realizado_pumi_dash:,.0f}")
        cdm4.metric("Pendiente", f"{pendiente_total_dash:,.0f}")
        cdm5.metric("% avance", f"{porcentaje_total_dash:.1%}")

        st.markdown("### Gráfico circular por estado")

        # El gráfico circular se basa en el estado de cada actividad oficial:
        # Completa: % Cumplimiento >= 100%
        # En avance: % Cumplimiento mayor a 0% y menor a 100%
        # Pendiente: % Cumplimiento igual a 0%
        orden_estados = ["Completa", "En avance", "Pendiente"]

        df_circular = (
            df_avance_filtrado
            .groupby("Estado cumplimiento", as_index=False)
            .agg({
                "Actividad": "count",
                "Meta 2026": "sum",
                "Realizado total": "sum",
                "Pendiente": "sum"
            })
            .rename(columns={
                "Estado cumplimiento": "Estado",
                "Actividad": "Actividades"
            })
        )

        df_circular["Orden"] = df_circular["Estado"].apply(
            lambda x: orden_estados.index(x) if x in orden_estados else 99
        )
        df_circular = df_circular.sort_values("Orden").drop(columns=["Orden"])

        if not df_circular.empty:
            fig_circular = px.pie(
                df_circular,
                names="Estado",
                values="Actividades",
                title="Actividades por estado de cumplimiento",
                hole=0.45,
                hover_data=["Meta 2026", "Realizado total", "Pendiente"]
            )
            fig_circular.update_traces(
                textinfo="label+percent+value"
            )
            fig_circular.update_layout(
                title_x=0.5,
                paper_bgcolor="white",
                plot_bgcolor="white"
            )
            st.plotly_chart(fig_circular, use_container_width=True)
        else:
            st.info("No hay datos para graficar por estado.")

        st.markdown("### Avance por actividad oficial")

        df_avance_vista = df_avance_filtrado.copy()
        df_avance_vista["% Cumplimiento"] = df_avance_vista["% Cumplimiento"].map(lambda x: f"{x:.1%}")

        columnas_vista_avance = [
            "Delegación", "Programa", "Actividad", "Meta 2026", "Avance base",
            "Realizado PUMI", "Realizado total", "Pendiente", "% Cumplimiento",
            "Estado cumplimiento"
        ]

        columnas_vista_avance = [c for c in columnas_vista_avance if c in df_avance_vista.columns]

        st.dataframe(
            df_avance_vista[columnas_vista_avance],
            use_container_width=True,
            hide_index=True
        )

        st.markdown("### Descarga del avance")

        boton_descargar_excel(
            df_avance_filtrado[columnas_vista_avance],
            texto_boton="⬇️ Descargar avance filtrado",
            filtrado=True,
            key="descarga_dashboard_avance_filtrado"
        )


# ======================================================
# DESCARGA GLOBAL EN SIDEBAR
# ======================================================

st.sidebar.markdown("---")
st.sidebar.markdown("### Descargar información")

df_sidebar = st.session_state.registros_pumi

if df_sidebar.empty:
    st.sidebar.info("No hay registros para descargar.")
else:
    nombre_sidebar = generar_nombre_excel_desde_df(
        df_sidebar,
        filtrado=False
    )

    st.sidebar.download_button(
        "⬇️ Descargar Excel actual",
        data=convertir_excel(df_sidebar),
        file_name=nombre_sidebar,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="descarga_sidebar_global",
        on_click="ignore"
    )

    if st.sidebar.button("🧹 Limpiar registros de la sesión"):
        st.session_state.registros_pumi = pd.DataFrame(columns=ENCABEZADOS)
        st.session_state.archivo_cargado_nombre = ""
        st.session_state.latitud_registro = ""
        st.session_state.longitud_registro = ""
        st.session_state.direccion_encontrada_mapa = ""
        st.sidebar.success("Registros limpiados.")
        st.rerun()
