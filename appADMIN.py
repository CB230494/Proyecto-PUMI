# ======================================================
# appADMIN.py
# PARTE 1 DE 5
# CONFIGURACIÓN GENERAL, LIBRERÍAS, CONSTANTES,
# ENCABEZADOS, ESTILOS, LOGOS Y FUNCIONES BASE
# ======================================================

import streamlit as st
import pandas as pd
import plotly.express as px
import folium
import os
import base64
import unicodedata
import re
import textwrap

from io import BytesIO
from datetime import datetime, date
from streamlit_folium import st_folium

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    PageBreak,
    Image,
    KeepTogether
)


# ======================================================
# CONFIGURACIÓN GENERAL STREAMLIT
# ======================================================

st.set_page_config(
    page_title="PUMI 2026 - Coordinadores Nacionales",
    page_icon="🛡️",
    layout="wide",
    initial_sidebar_state="expanded"
)


# ======================================================
# ARCHIVOS Y LOGOS
# Deben estar al mismo nivel que appADMIN.py
# ======================================================

LOGO_MINISTERIO = "Logo2.jpeg"
LOGO_PUMI = "logo_pumi.jpeg"
LOGO_FUERZA_PUBLICA = "Logo1.jpeg"

ARCHIVO_DATOS_IMPORTANTES = "Datos Importantes.xlsx"
ARCHIVO_MEP = "BASE DE DATOS MEP 2025.xlsx"

NOMBRE_HOJA_EXCEL = "REGISTRO_PUMI_2026"


# ======================================================
# ENCABEZADOS BASE DE LA APP PUMI
# ======================================================

ENCABEZADOS_PUMI = [
    "ID",
    "Fecha Actividad",
    "Hora Actividad",
    "Dirección Regional",
    "Delegación",
    "Responde a",
    "Programa",
    "Actividad",
    "Provincia",
    "Cantón",
    "Distrito",
    "Tipo Lugar",
    "Lugar",
    "Centro Educativo",
    "Código Presupuestario",
    "Dirección Mapa",
    "Latitud",
    "Longitud",
    "Responsable",
    "Cantidad Participantes",
    "Cantidad Hombres",
    "Cantidad Mujeres",
    "Edad 10 a 18",
    "Edad 19 a 30",
    "Edad 31 a 45",
    "Edad 46 en adelante",
    "Instituciones Participantes",
    "Número de Referencia",
    "Número de Expediente Referencia",
    "Observaciones",
    "Usuario Registra"
]


# ======================================================
# ENCABEZADOS ADMINISTRATIVOS
# Estos campos se agregan para validación.
# ======================================================

ENCABEZADOS_VERIFICACION_REGIONAL = [
    "Estado Verificación Regional",
    "Observaciones Verificación Regional",
    "Coordinador Regional",
    "Fecha Verificación Regional"
]

ENCABEZADOS_VALIDACION = [
    "Estado Validación",
    "Observaciones Validación",
    "Funcionario Validador",
    "Fecha Validación"
]


ENCABEZADOS_EXTRA_CONTROL = ["Archivo Origen"]

ENCABEZADOS_COMPLETOS = ENCABEZADOS_EXTRA_CONTROL + ENCABEZADOS_PUMI + ENCABEZADOS_VERIFICACION_REGIONAL + ENCABEZADOS_VALIDACION


# ======================================================
# COLORES INSTITUCIONALES
# ======================================================

COLOR_AZUL = "#002B7F"
COLOR_AZUL_MEDIO = "#1E4FA3"
COLOR_AZUL_CLARO = "#DCE8FF"

COLOR_DORADO = "#B88A2A"
COLOR_DORADO_OSCURO = "#8A6418"
COLOR_DORADO_CLARO = "#F4E6C1"

COLOR_BLANCO = "#FFFFFF"
COLOR_GRIS = "#F4F6F8"
COLOR_GRIS_OSCURO = "#2F3542"

COLOR_VERDE = "#1F8A4C"
COLOR_ROJO = "#B42318"
COLOR_AMARILLO = "#F2C94C"


COLORES_VALIDACION = {
    "Pendiente": "#F2C94C",
    "Aprobada": "#1F8A4C",
    "Rechazada": "#B42318",
    "Con observaciones": "#B88A2A"
}


COLORES_PROGRAMA = {
    "DARE": "purple",
    "GREAT": "orange",
    "GREAT CAMP": "orange",
    "MPAS": "darkblue",
    "PSCC": "blue",
    "VIF": "orange",
    "Política Pública": "green"
}


# ======================================================
# ACCESOS POR PROGRAMA PARA COORDINADORES NACIONALES
# Se leen desde Usuarios y Claves PUMI.xlsx
# Hoja esperada: columnas Programas y CLAVES
# ======================================================

ARCHIVO_USUARIOS_CLAVES = "Usuarios y Claves PUMI.xlsx"

# Respaldo si el Excel no está disponible.
PROGRAMAS_ACCESO_ADMIN_RESPALDO = {
    "PSCC": "PSCC23",
    "GREAT/MPAS/DARE": "GMD23",
    "DARE": "DARE23",
    "PLAN DE POLITICA LOCAL E IMPLEMENTACIÓN DE LA POLITICA LOCAL DE SEGURIDAD": "PLAN23"
}

PROGRAMAS_BASE = [
    "DARE",
    "GREAT",
    "MPAS",
    "PSCC",
    "VIF",
    "Política Pública"
]

REGIONES_BASE = [
    "R1 San José Central",
    "R2 San José Norte",
    "R3 San José Sur",
    "R4 Alajuela",
    "R5 Cartago",
    "R6 Heredia",
    "R7 Chorotega",
    "R8 Puntarenas",
    "R9 Limón",
    "R10 Brunca",
    "R11 Chorotega Norte",
    "R12"
]

PROVINCIAS_BASE = [
    "San José",
    "Alajuela",
    "Cartago",
    "Heredia",
    "Guanacaste",
    "Puntarenas",
    "Limón"
]


CENTROS_PROVINCIA = {
    "San José": [9.9281, -84.0907],
    "Alajuela": [10.0162, -84.2116],
    "Cartago": [9.8644, -83.9194],
    "Heredia": [10.0024, -84.1165],
    "Guanacaste": [10.6267, -85.4437],
    "Puntarenas": [9.9763, -84.8384],
    "Limón": [9.9917, -83.0360]
}


# ======================================================
# SESSION STATE
# ======================================================

def inicializar_session_state_admin():
    if "df_admin" not in st.session_state:
        st.session_state.df_admin = pd.DataFrame(columns=ENCABEZADOS_COMPLETOS)

    if "archivo_admin_nombre" not in st.session_state:
        st.session_state.archivo_admin_nombre = ""

    if "filtros_admin_aplicados" not in st.session_state:
        st.session_state.filtros_admin_aplicados = {}

    if "mapa_imagen_referencia" not in st.session_state:
        st.session_state.mapa_imagen_referencia = None

    if "grafico_imagen_referencia" not in st.session_state:
        st.session_state.grafico_imagen_referencia = None

    if "admin_programa_autenticado" not in st.session_state:
        st.session_state.admin_programa_autenticado = ""

    if "admin_acceso_programa" not in st.session_state:
        st.session_state.admin_acceso_programa = False

    if "admin_nombre_acceso" not in st.session_state:
        st.session_state.admin_nombre_acceso = ""


inicializar_session_state_admin()


# ======================================================
# FUNCIONES BASE
# ======================================================

def imagen_a_base64(ruta):
    if not os.path.exists(ruta):
        return ""

    try:
        with open(ruta, "rb") as archivo:
            return base64.b64encode(archivo.read()).decode()
    except Exception:
        return ""


def normalizar_texto(valor):
    if pd.isna(valor):
        return ""

    texto = str(valor).strip().upper()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(
        caracter for caracter in texto
        if not unicodedata.combining(caracter)
    )
    texto = " ".join(texto.split())

    return texto


def limpiar_nombre_archivo(valor):
    texto = str(valor).strip()

    if texto == "" or texto.lower() == "nan":
        return "SIN_DATO"

    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(
        caracter for caracter in texto
        if not unicodedata.combining(caracter)
    )

    texto = texto.upper()
    texto = re.sub(r"[^\w\s-]", "", texto)
    texto = re.sub(r"\s+", "_", texto)
    texto = texto.strip("_")

    if texto == "":
        return "SIN_DATO"

    return texto


def limpiar_coordenada(valor):
    try:
        if valor is None or str(valor).strip() == "":
            return None

        return float(str(valor).replace(",", ".").strip())

    except Exception:
        return None


def obtener_color_programa(programa):
    return COLORES_PROGRAMA.get(str(programa), "gray")


def generar_nombre_reporte(df, tipo="PDF"):
    if df is None or df.empty:
        return f"INFORME_VALIDACION_PUMI_2026.{tipo.lower()}"

    region = "VARIAS_REGIONES"
    delegacion = "VARIAS_DELEGACIONES"

    if "Dirección Regional" in df.columns:
        regiones = (
            df["Dirección Regional"]
            .replace("", pd.NA)
            .dropna()
            .astype(str)
            .unique()
            .tolist()
        )

        if len(regiones) == 1:
            region = regiones[0]

    if "Delegación" in df.columns:
        delegaciones = (
            df["Delegación"]
            .replace("", pd.NA)
            .dropna()
            .astype(str)
            .unique()
            .tolist()
        )

        if len(delegaciones) == 1:
            delegacion = delegaciones[0]

    region_limpia = limpiar_nombre_archivo(region)
    delegacion_limpia = limpiar_nombre_archivo(delegacion)

    return f"INFORME_VALIDACION_PUMI_2026_{region_limpia}_{delegacion_limpia}.{tipo.lower()}"



def normalizar_programa_pumi(valor):
    texto = normalizar_texto(valor)
    texto = texto.replace(" Y ", "/")
    texto = texto.replace(" - ", "/")
    texto = re.sub(r"\s*/\s*", "/", texto)
    return texto


@st.cache_data
def cargar_accesos_programas_admin():
    columnas = ["Programa", "Clave"]

    if not os.path.exists(ARCHIVO_USUARIOS_CLAVES):
        return pd.DataFrame([
            {"Programa": k, "Clave": v}
            for k, v in PROGRAMAS_ACCESO_ADMIN_RESPALDO.items()
        ])

    try:
        df = pd.read_excel(ARCHIVO_USUARIOS_CLAVES)
        df.columns = [str(c).strip() for c in df.columns]

        col_programa = None
        col_clave = None

        for col in df.columns:
            if normalizar_texto(col) in ["PROGRAMAS", "PROGRAMA"]:
                col_programa = col
            if normalizar_texto(col) in ["CLAVES", "CLAVE"]:
                col_clave = col

        if not col_programa or not col_clave:
            return pd.DataFrame([
                {"Programa": k, "Clave": v}
                for k, v in PROGRAMAS_ACCESO_ADMIN_RESPALDO.items()
            ])

        accesos = df[[col_programa, col_clave]].copy()
        accesos.columns = columnas
        accesos["Programa"] = accesos["Programa"].fillna("").astype(str).str.strip()
        accesos["Clave"] = accesos["Clave"].fillna("").astype(str).str.strip()
        accesos = accesos[(accesos["Programa"] != "") & (accesos["Clave"] != "")]
        accesos = accesos.drop_duplicates()

        if accesos.empty:
            return pd.DataFrame([
                {"Programa": k, "Clave": v}
                for k, v in PROGRAMAS_ACCESO_ADMIN_RESPALDO.items()
            ])

        return accesos

    except Exception:
        return pd.DataFrame([
            {"Programa": k, "Clave": v}
            for k, v in PROGRAMAS_ACCESO_ADMIN_RESPALDO.items()
        ])


def programa_permitido_por_acceso(programa_registro, programa_acceso):
    registro = normalizar_programa_pumi(programa_registro)
    acceso = normalizar_programa_pumi(programa_acceso)

    if not registro or not acceso:
        return False

    # Regla especial solicitada:
    # - DARE puede ver DARE exclusivo y la capa compartida GREAT/MPAS/DARE.
    # - GREAT/MPAS/DARE puede ver solo la capa compartida y actividades GREAT/MPAS si vienen separadas,
    #   pero NO puede ver el DARE exclusivo.
    if acceso == "DARE":
        return registro in ["DARE", "GREAT/MPAS/DARE"]

    if acceso == "GREAT/MPAS/DARE":
        return registro in ["GREAT/MPAS/DARE", "GREAT", "GREAT CAMP", "MPAS"]

    if acceso.startswith("PLAN DE POLITICA LOCAL"):
        return registro.startswith("PLAN DE POLITICA LOCAL")

    return registro == acceso


def filtrar_dataframe_por_programa_admin(df, programa_autorizado):
    if df is None or df.empty:
        return df

    if not programa_autorizado or "Programa" not in df.columns:
        return pd.DataFrame(columns=df.columns)

    mascara = df["Programa"].apply(
        lambda x: programa_permitido_por_acceso(x, programa_autorizado)
    )

    return df[mascara].copy()


def mostrar_acceso_por_programa_admin():
    st.sidebar.markdown("---")
    st.sidebar.markdown("### Acceso coordinador nacional")

    if st.session_state.get("admin_acceso_programa"):
        programa_actual = st.session_state.get("admin_programa_autenticado", "")
        st.sidebar.success(f"Acceso activo: {programa_actual}")

        if st.sidebar.button("🔒 Cerrar acceso"):
            st.session_state.admin_acceso_programa = False
            st.session_state.admin_programa_autenticado = ""
            st.session_state.admin_nombre_acceso = ""
            st.rerun()

        return True

    accesos = cargar_accesos_programas_admin()
    programas = accesos["Programa"].dropna().astype(str).tolist()

    programa = st.sidebar.selectbox(
        "Programa",
        programas,
        key="programa_acceso_admin"
    )

    clave = st.sidebar.text_input(
        "Clave",
        type="password",
        key="clave_acceso_admin"
    )

    if st.sidebar.button("Ingresar"):
        fila = accesos[
            accesos["Programa"].apply(normalizar_texto) == normalizar_texto(programa)
        ]

        clave_esperada = ""
        if not fila.empty:
            clave_esperada = str(fila.iloc[0].get("Clave", "")).strip()

        if clave.strip() == clave_esperada and clave_esperada:
            st.session_state.admin_acceso_programa = True
            st.session_state.admin_programa_autenticado = programa
            st.session_state.admin_nombre_acceso = programa
            st.sidebar.success("Acceso autorizado.")
            st.rerun()
        else:
            st.sidebar.error("Usuario o clave incorrectos.")

    return False




# ======================================================
# LOGIN CENTRAL PARA COORDINADORES NACIONALES
# ======================================================

def mostrar_login_coordinador_nacional():
    st.markdown(
        """
        <div style="max-width:720px;margin:20px auto 0 auto;text-align:center;">
            <h1 style="color:#002B7F;font-weight:900;margin-bottom:8px;">
                Ingreso al sistema
            </h1>
            <p style="font-size:18px;color:#2F3542;margin-bottom:28px;">
                Coordinadores Nacionales PUMI 2026
            </p>
        </div>
        """,
        unsafe_allow_html=True
    )

    accesos = cargar_accesos_programas_admin()

    col_izq, col_centro, col_der = st.columns([1, 1.35, 1])

    with col_centro:
        usuario = st.text_input(
            "Usuario",
            key="login_usuario_coordinador_nacional",
            placeholder="Digite el programa asignado"
        )

        clave = st.text_input(
            "Clave",
            type="password",
            key="login_clave_coordinador_nacional"
        )

        if st.button("Ingresar", key="btn_login_coordinador_nacional"):
            usuario_norm = normalizar_texto(usuario)
            clave_txt = str(clave).strip()

            fila = accesos[
                accesos["Programa"].apply(normalizar_texto) == usuario_norm
            ]

            if not fila.empty:
                clave_esperada = str(fila.iloc[0].get("Clave", "")).strip()
                programa_real = str(fila.iloc[0].get("Programa", "")).strip()
            else:
                clave_esperada = ""
                programa_real = ""

            if programa_real and clave_txt == clave_esperada:
                st.session_state.admin_acceso_programa = True
                st.session_state.admin_programa_autenticado = programa_real
                st.session_state.admin_nombre_acceso = programa_real
                st.success("Acceso autorizado.")
                st.rerun()
            else:
                st.error("Usuario o clave incorrectos.")

    st.markdown(
        """
        <div style="max-width:900px;margin:34px auto 0 auto;background:#FFFFFF;border-radius:20px;
                    padding:22px;border-left:8px solid #B88A2A;box-shadow:0 6px 16px rgba(0,0,0,.12);">
            <b>Acceso por programa:</b> cada coordinación nacional visualiza, valida y genera reportes
            únicamente con la información que le corresponde según el usuario autorizado.
        </div>
        """,
        unsafe_allow_html=True
    )


# ======================================================
# ESTILOS CSS
# Misma línea visual de la app PUMI principal.
# ======================================================

st.markdown(
    f"""
    <style>

    .stApp {{
        background:
            radial-gradient(circle at top left, #FFFFFF 0, transparent 28%),
            radial-gradient(circle at top right, #DCE8FF 0, transparent 30%),
            linear-gradient(180deg, #FFFFFF 0%, #F4F7FB 48%, #FFFFFF 100%);
    }}

    section[data-testid="stSidebar"] {{
        background: linear-gradient(180deg, #FFFFFF 0%, #EAF1FF 45%, #FFFFFF 100%);
        border-right: 5px solid {COLOR_DORADO};
        box-shadow: 4px 0px 12px rgba(0,0,0,0.10);
    }}

    .sidebar-logo-pumi {{
        background: #FFFFFF;
        border-radius: 16px;
        padding: 12px;
        margin: 8px 0px 18px 0px;
        box-shadow: 0px 4px 14px rgba(0,0,0,0.10);
        text-align: center;
    }}

    .sidebar-logo-pumi img {{
        width: 100%;
        max-height: 145px;
        object-fit: contain;
    }}

    .encabezado-institucional {{
        background: #FFFFFF;
        border-radius: 22px;
        padding: 22px 34px;
        margin-bottom: 24px;
        box-shadow: 0px 8px 22px rgba(0,0,0,0.11);
        border-bottom: 7px solid {COLOR_DORADO};
        display: grid;
        grid-template-columns: 1fr 1fr 1fr;
        align-items: center;
        gap: 22px;
        min-height: 155px;
    }}

    .encabezado-logo-izq,
    .encabezado-logo-centro,
    .encabezado-logo-der {{
        height: 125px;
        display: flex;
        align-items: center;
    }}

    .encabezado-logo-izq {{
        justify-content: flex-start;
    }}

    .encabezado-logo-centro {{
        justify-content: center;
    }}

    .encabezado-logo-der {{
        justify-content: flex-end;
    }}

    .logo-ministerio-app {{
        width: 350px;
        height: 100px;
        object-fit: contain;
    }}

    .logo-pumi-app {{
        width: 300px;
        height: 100px;
        object-fit: contain;
    }}

    .logo-fp-app {{
        width: 400px;
        height: 150px;
        object-fit: contain;
    }}

    .titulo-principal {{
        background: linear-gradient(135deg, {COLOR_AZUL}, #243B63, {COLOR_DORADO});
        padding: 34px;
        border-radius: 22px;
        text-align: center;
        color: white;
        margin-bottom: 30px;
        box-shadow: 0px 8px 22px rgba(0,0,0,0.28);
        border: 3px solid {COLOR_BLANCO};
    }}

    .titulo-principal h1 {{
        font-size: 46px;
        margin-bottom: 8px;
        font-weight: 900;
        letter-spacing: 1.5px;
        color: white;
        text-shadow: 1px 2px 5px rgba(0,0,0,0.35);
    }}

    .titulo-principal h3 {{
        font-size: 25px;
        margin-top: 8px;
        font-weight: 600;
        color: white;
    }}

    .card-admin {{
        padding: 24px;
        border-radius: 20px;
        box-shadow: 0px 6px 16px rgba(0,0,0,0.13);
        margin-bottom: 20px;
        background: linear-gradient(135deg, #FFFFFF 0%, #F8FAFF 100%);
        border-left: 9px solid {COLOR_AZUL};
    }}

    .card-validacion {{
        padding: 22px;
        border-radius: 18px;
        box-shadow: 0px 5px 14px rgba(0,0,0,0.12);
        margin-bottom: 18px;
        background: linear-gradient(135deg, #FFFFFF 0%, #FFF8E8 100%);
        border-left: 8px solid {COLOR_DORADO};
    }}

    .subtitulo-admin {{
        color: {COLOR_AZUL};
        font-weight: 900;
        font-size: 29px;
        margin-bottom: 12px;
        text-align: center;
    }}

    .texto-admin {{
        color: {COLOR_GRIS_OSCURO};
        font-size: 17px;
        line-height: 1.7;
        text-align: justify;
    }}

    div[data-testid="stMetric"] {{
        background: linear-gradient(145deg, #FFFFFF 0%, #F3F6FF 100%);
        padding: 20px;
        border-radius: 20px;
        box-shadow: 0px 5px 15px rgba(0,0,0,0.13);
        border-bottom: 6px solid {COLOR_DORADO};
    }}

    div[data-baseweb="select"] > div {{
        background-color: #FFFFFF !important;
        border-radius: 12px !important;
        border: 1.8px solid #AFC3EE !important;
    }}

    input {{
        background-color: #FFFFFF !important;
        border-radius: 12px !important;
        border: 1.8px solid #AFC3EE !important;
    }}

    textarea {{
        background-color: #FFFFFF !important;
        border-radius: 12px !important;
        border: 1.8px solid {COLOR_DORADO_CLARO} !important;
    }}

    .stButton > button {{
        background: linear-gradient(90deg, {COLOR_AZUL}, {COLOR_DORADO});
        color: white;
        border-radius: 12px;
        border: none;
        font-weight: 800;
        padding: 0.7rem 1.2rem;
        box-shadow: 0px 4px 10px rgba(0,0,0,0.18);
    }}

    .stDownloadButton > button {{
        background: linear-gradient(90deg, {COLOR_DORADO}, {COLOR_AZUL});
        color: white;
        border-radius: 12px;
        border: none;
        font-weight: 800;
        padding: 0.7rem 1.2rem;
        box-shadow: 0px 4px 10px rgba(0,0,0,0.18);
    }}

    div[data-testid="stDataFrame"] {{
        border-radius: 18px;
        overflow: hidden;
        box-shadow: 0px 5px 16px rgba(0,0,0,0.13);
        background-color: white;
    }}

    iframe {{
        border-radius: 18px !important;
        box-shadow: 0px 5px 16px rgba(0,0,0,0.13);
        width: 100% !important;
    }}

    hr {{
        border: none;
        height: 3px;
        background: linear-gradient(90deg, {COLOR_AZUL}, {COLOR_BLANCO}, {COLOR_DORADO});
        margin-top: 25px;
        margin-bottom: 25px;
    }}

    h1, h2, h3 {{
        color: {COLOR_AZUL};
        font-weight: 900;
    }}

    </style>
    """,
    unsafe_allow_html=True
)


# ======================================================
# LOGOS Y ENCABEZADOS
# ======================================================

def mostrar_logo_sidebar():
    logo_pumi_b64 = imagen_a_base64(LOGO_PUMI)

    if logo_pumi_b64:
        st.sidebar.markdown(
            f"""
            <div class="sidebar-logo-pumi">
                <img src="data:image/jpeg;base64,{logo_pumi_b64}">
            </div>
            """,
            unsafe_allow_html=True
        )
    else:
        st.sidebar.warning("Logo PUMI no encontrado.")


def mostrar_encabezado_institucional():
    logo_min_b64 = imagen_a_base64(LOGO_MINISTERIO)
    logo_pumi_b64 = imagen_a_base64(LOGO_PUMI)
    logo_fp_b64 = imagen_a_base64(LOGO_FUERZA_PUBLICA)

    img_min = (
        f'<img class="logo-ministerio-app" src="data:image/jpeg;base64,{logo_min_b64}">'
        if logo_min_b64 else ""
    )

    img_pumi = (
        f'<img class="logo-pumi-app" src="data:image/jpeg;base64,{logo_pumi_b64}">'
        if logo_pumi_b64 else ""
    )

    img_fp = (
        f'<img class="logo-fp-app" src="data:image/jpeg;base64,{logo_fp_b64}">'
        if logo_fp_b64 else ""
    )

    st.markdown(
        f"""
        <div class="encabezado-institucional">
            <div class="encabezado-logo-izq">{img_min}</div>
            <div class="encabezado-logo-centro">{img_pumi}</div>
            <div class="encabezado-logo-der">{img_fp}</div>
        </div>
        """,
        unsafe_allow_html=True
    )


def mostrar_titulo_admin():
    st.markdown(
        """
        <div class="titulo-principal">
            <h1>🛡️ P.U.M.I. 2026 - Coordinadores Nacionales</h1>
            <h3>Validación nacional, análisis, dashboard e informe PDF por programa</h3>
        </div>
        """,
        unsafe_allow_html=True
    )
# ======================================================
# appADMIN.py
# PARTE 2 DE 5 CORREGIDA
# CARGA DE EXCEL PUMI, PREPARACIÓN DE DATOS,
# FILTROS GENERALES, MAPA CON TIPO DE MAPA
# Y DASHBOARD CON COLORES INSTITUCIONALES
# ======================================================


# ======================================================
# PREPARAR DATAFRAME ADMINISTRATIVO
# ======================================================

def preparar_dataframe_admin(df):
    df = df.copy()

    for col in ENCABEZADOS_EXTRA_CONTROL:
        if col not in df.columns:
            df[col] = ""

    for col in ENCABEZADOS_PUMI:
        if col not in df.columns:
            df[col] = ""

    for col in ENCABEZADOS_VERIFICACION_REGIONAL:
        if col not in df.columns:
            df[col] = ""

    for col in ENCABEZADOS_VALIDACION:
        if col not in df.columns:
            df[col] = ""

    df = df[ENCABEZADOS_COMPLETOS]

    for col in ENCABEZADOS_COMPLETOS:
        df[col] = df[col].fillna("").astype(str)

    columnas_numericas = [
        "ID",
        "Cantidad Participantes",
        "Cantidad Hombres",
        "Cantidad Mujeres",
        "Edad 10 a 18",
        "Edad 19 a 30",
        "Edad 31 a 45",
        "Edad 46 en adelante"
    ]

    for col in columnas_numericas:
        if col in df.columns:
            df[col] = pd.to_numeric(
                df[col],
                errors="coerce"
            ).fillna(0).astype(int)

    df["Estado Verificación Regional"] = df["Estado Verificación Regional"].replace(
        "",
        "Pendiente de verificación"
    )

    df["Estado Validación"] = df["Estado Validación"].replace(
        "",
        "Pendiente"
    )

    return df


# ======================================================
# CARGAR EXCEL GENERADO POR APP PUMI
# ======================================================

def cargar_excel_pumi_admin(archivo_excel):
    try:
        df = pd.read_excel(
            archivo_excel,
            sheet_name=NOMBRE_HOJA_EXCEL
        )
        if "Archivo Origen" not in df.columns:
            df["Archivo Origen"] = getattr(archivo_excel, "name", "")
        else:
            df["Archivo Origen"] = df["Archivo Origen"].replace("", getattr(archivo_excel, "name", ""))
    except Exception:
        try:
            df = pd.read_excel(archivo_excel)
            if "Archivo Origen" not in df.columns:
                df["Archivo Origen"] = getattr(archivo_excel, "name", "")
            else:
                df["Archivo Origen"] = df["Archivo Origen"].replace("", getattr(archivo_excel, "name", ""))
        except Exception as e:
            st.error("No se pudo leer el Excel cargado.")
            st.exception(e)
            return pd.DataFrame(columns=ENCABEZADOS_COMPLETOS)

    return preparar_dataframe_admin(df)


# ======================================================
# EXPORTAR EXCEL ADMINISTRATIVO
# CON COLORES EN COLUMNAS DE VALIDACIÓN
# ======================================================

def convertir_excel_admin(df):
    output = BytesIO()

    df_exportar = preparar_dataframe_admin(df)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_exportar.to_excel(
            writer,
            index=False,
            sheet_name=NOMBRE_HOJA_EXCEL
        )

    output.seek(0)

    workbook = load_workbook(output)
    worksheet = workbook[NOMBRE_HOJA_EXCEL]

    fill_header = PatternFill(
        start_color="002B7F",
        end_color="002B7F",
        fill_type="solid"
    )

    fill_header_admin = PatternFill(
        start_color="B88A2A",
        end_color="B88A2A",
        fill_type="solid"
    )

    font_header = Font(
        color="FFFFFF",
        bold=True
    )

    border = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF")
    )

    columnas_regional = {
        "Estado Verificación Regional",
        "Observaciones Verificación Regional",
        "Coordinador Regional",
        "Fecha Verificación Regional"
    }

    columnas_validacion = {
        "Estado Validación",
        "Observaciones Validación",
        "Funcionario Validador",
        "Fecha Validación"
    }

    for cell in worksheet[1]:
        if cell.value in columnas_validacion:
            cell.fill = fill_header_admin
        elif cell.value in columnas_regional:
            cell.fill = PatternFill(start_color="1E4FA3", end_color="1E4FA3", fill_type="solid")
        else:
            cell.fill = fill_header

        cell.font = font_header
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        cell.border = border

    colores_estado = {
        "Pendiente": "F2C94C",
        "Aprobada": "1F8A4C",
        "Rechazada": "B42318",
        "Con observaciones": "B88A2A"
    }

    col_estado = None
    col_estado_regional = None

    for idx, cell in enumerate(worksheet[1], start=1):
        if cell.value == "Estado Validación":
            col_estado = idx
        if cell.value == "Estado Verificación Regional":
            col_estado_regional = idx

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )
            cell.border = border

        if col_estado_regional:
            celda_estado_regional = row[col_estado_regional - 1]
            estado_regional = str(celda_estado_regional.value).strip()
            colores_regional = {
                "Pendiente de verificación": "F2C94C",
                "Verificada para envío": "1F8A4C",
                "Devuelta a delegación": "B42318",
                "Con observaciones regionales": "B88A2A"
            }
            if estado_regional in colores_regional:
                celda_estado_regional.fill = PatternFill(
                    start_color=colores_regional[estado_regional],
                    end_color=colores_regional[estado_regional],
                    fill_type="solid"
                )
                celda_estado_regional.font = Font(color="FFFFFF", bold=True)
                celda_estado_regional.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        if col_estado:
            celda_estado = row[col_estado - 1]
            estado = str(celda_estado.value).strip()

            if estado in colores_estado:
                celda_estado.fill = PatternFill(
                    start_color=colores_estado[estado],
                    end_color=colores_estado[estado],
                    fill_type="solid"
                )
                celda_estado.font = Font(
                    color="FFFFFF",
                    bold=True
                )
                celda_estado.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            valor = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(valor))

        worksheet.column_dimensions[column_letter].width = min(
            max(max_length + 2, 14),
            45
        )

    worksheet.freeze_panes = "A2"

    final_output = BytesIO()
    workbook.save(final_output)
    final_output.seek(0)

    return final_output.getvalue()


# ======================================================
# LIMPIEZA PARA MÉTRICAS Y FILTROS
# ======================================================

def limpiar_dataframe_metricas_admin(df):
    df = df.copy()

    columnas_numericas = [
        "Cantidad Participantes",
        "Cantidad Hombres",
        "Cantidad Mujeres",
        "Edad 10 a 18",
        "Edad 19 a 30",
        "Edad 31 a 45",
        "Edad 46 en adelante"
    ]

    for col in columnas_numericas:
        if col in df.columns:
            df[col] = pd.to_numeric(
                df[col],
                errors="coerce"
            ).fillna(0)

    if "Fecha Actividad" in df.columns:
        df["Fecha Actividad"] = pd.to_datetime(
            df["Fecha Actividad"],
            errors="coerce",
            dayfirst=True
        )

    return df


# ======================================================
# FILTROS GENERALES DEL PANEL ADMIN
# ======================================================

def aplicar_filtros_admin(df):
    if df is None or df.empty:
        return df

    df_filtrado = df.copy()

    st.markdown("### Filtros de consulta")

    col1, col2, col3 = st.columns(3)

    with col1:
        filtro_region = st.multiselect(
            "Dirección Regional",
            sorted(
                df_filtrado["Dirección Regional"]
                .replace("", pd.NA)
                .dropna()
                .astype(str)
                .unique()
                .tolist()
            )
        )

        filtro_delegacion = st.multiselect(
            "Delegación",
            sorted(
                df_filtrado["Delegación"]
                .replace("", pd.NA)
                .dropna()
                .astype(str)
                .unique()
                .tolist()
            )
        )

    with col2:
        filtro_programa = st.multiselect(
            "Programa",
            sorted(
                df_filtrado["Programa"]
                .replace("", pd.NA)
                .dropna()
                .astype(str)
                .unique()
                .tolist()
            )
        )

        filtro_actividad = st.multiselect(
            "Actividad",
            sorted(
                df_filtrado["Actividad"]
                .replace("", pd.NA)
                .dropna()
                .astype(str)
                .unique()
                .tolist()
            )
        )

    with col3:
        filtro_estado_regional = st.multiselect(
            "Estado de verificación regional",
            [
                "Pendiente de verificación",
                "Verificada para envío",
                "Devuelta a delegación",
                "Con observaciones regionales"
            ],
            default=["Verificada para envío"]
        )

        filtro_estado = st.multiselect(
            "Estado de validación nacional",
            [
                "Pendiente",
                "Aprobada",
                "Rechazada",
                "Con observaciones"
            ]
        )

        filtro_provincia = st.multiselect(
            "Provincia",
            sorted(
                df_filtrado["Provincia"]
                .replace("", pd.NA)
                .dropna()
                .astype(str)
                .unique()
                .tolist()
            )
        )

    colf1, colf2 = st.columns(2)

    with colf1:
        fecha_inicio = st.date_input(
            "Fecha inicio",
            value=None,
            format="DD/MM/YYYY"
        )

    with colf2:
        fecha_fin = st.date_input(
            "Fecha final",
            value=None,
            format="DD/MM/YYYY"
        )

    if filtro_region:
        df_filtrado = df_filtrado[
            df_filtrado["Dirección Regional"].isin(filtro_region)
        ]

    if filtro_delegacion:
        df_filtrado = df_filtrado[
            df_filtrado["Delegación"].isin(filtro_delegacion)
        ]

    if filtro_programa:
        df_filtrado = df_filtrado[
            df_filtrado["Programa"].isin(filtro_programa)
        ]

    if filtro_actividad:
        df_filtrado = df_filtrado[
            df_filtrado["Actividad"].isin(filtro_actividad)
        ]

    if filtro_estado_regional:
        df_filtrado = df_filtrado[
            df_filtrado["Estado Verificación Regional"].isin(filtro_estado_regional)
        ]

    if filtro_estado:
        df_filtrado = df_filtrado[
            df_filtrado["Estado Validación"].isin(filtro_estado)
        ]

    if filtro_provincia:
        df_filtrado = df_filtrado[
            df_filtrado["Provincia"].isin(filtro_provincia)
        ]

    if fecha_inicio or fecha_fin:
        df_fecha = df_filtrado.copy()

        df_fecha["Fecha_Actividad_DT"] = pd.to_datetime(
            df_fecha["Fecha Actividad"],
            errors="coerce",
            dayfirst=True
        )

        if fecha_inicio:
            df_fecha = df_fecha[
                df_fecha["Fecha_Actividad_DT"].dt.date >= fecha_inicio
            ]

        if fecha_fin:
            df_fecha = df_fecha[
                df_fecha["Fecha_Actividad_DT"].dt.date <= fecha_fin
            ]

        df_filtrado = df_fecha.drop(columns=["Fecha_Actividad_DT"])

    st.session_state.filtros_admin_aplicados = {
        "Dirección Regional": filtro_region,
        "Delegación": filtro_delegacion,
        "Programa": filtro_programa,
        "Actividad": filtro_actividad,
        "Estado Verificación Regional": filtro_estado_regional,
        "Estado Validación Nacional": filtro_estado,
        "Provincia": filtro_provincia,
        "Fecha Inicio": str(fecha_inicio) if fecha_inicio else "",
        "Fecha Final": str(fecha_fin) if fecha_fin else ""
    }

    return df_filtrado


# ======================================================
# CREAR MAPA BASE ADMIN CON TIPO DE MAPA
# ======================================================

def crear_mapa_base_admin(centro, zoom, tipo_mapa):
    mapa = folium.Map(
        location=centro,
        zoom_start=zoom,
        tiles=None
    )

    if tipo_mapa == "OpenStreetMap":
        folium.TileLayer(
            "OpenStreetMap",
            name="OpenStreetMap"
        ).add_to(mapa)

    elif tipo_mapa == "Mapa claro":
        folium.TileLayer(
            "CartoDB positron",
            name="Mapa claro"
        ).add_to(mapa)

    elif tipo_mapa == "Mapa oscuro":
        folium.TileLayer(
            "CartoDB dark_matter",
            name="Mapa oscuro"
        ).add_to(mapa)

    elif tipo_mapa == "Topográfico":
        folium.TileLayer(
            "OpenTopoMap",
            name="Topográfico"
        ).add_to(mapa)

    elif tipo_mapa == "Satélite":
        folium.TileLayer(
            tiles=(
                "https://server.arcgisonline.com/ArcGIS/rest/services/"
                "World_Imagery/MapServer/tile/{z}/{y}/{x}"
            ),
            attr="Esri World Imagery",
            name="Satélite"
        ).add_to(mapa)

    folium.LayerControl().add_to(mapa)

    return mapa


# ======================================================
# DATAFRAME PARA MAPA
# ======================================================

def preparar_dataframe_mapa_admin(df):
    df_mapa = df.copy()

    if "Latitud" not in df_mapa.columns:
        df_mapa["Latitud"] = ""

    if "Longitud" not in df_mapa.columns:
        df_mapa["Longitud"] = ""

    df_mapa["Latitud_Num"] = df_mapa["Latitud"].apply(limpiar_coordenada)
    df_mapa["Longitud_Num"] = df_mapa["Longitud"].apply(limpiar_coordenada)

    df_mapa = df_mapa.dropna(
        subset=[
            "Latitud_Num",
            "Longitud_Num"
        ]
    )

    return df_mapa


# ======================================================
# MAPA GENERAL DE REGISTROS ADMIN
# ======================================================

def crear_mapa_admin(df, tipo_mapa="OpenStreetMap"):
    df_mapa = preparar_dataframe_mapa_admin(df)

    if df_mapa.empty:
        return crear_mapa_base_admin(
            centro=[9.7489, -83.7534],
            zoom=7,
            tipo_mapa=tipo_mapa
        )

    if len(df_mapa) == 1:
        centro = [
            float(df_mapa.iloc[0]["Latitud_Num"]),
            float(df_mapa.iloc[0]["Longitud_Num"])
        ]
        zoom = 16
    else:
        centro = [
            float(df_mapa["Latitud_Num"].mean()),
            float(df_mapa["Longitud_Num"].mean())
        ]
        zoom = 8

    mapa = crear_mapa_base_admin(
        centro=centro,
        zoom=zoom,
        tipo_mapa=tipo_mapa
    )

    for _, row in df_mapa.iterrows():
        estado = str(row.get("Estado Validación", "Pendiente"))
        programa = str(row.get("Programa", ""))

        color = obtener_color_programa(programa)

        popup_html = f"""
        <div style="font-family:Arial; width:300px;">
            <h4 style="color:#002B7F;">Registro PUMI #{row.get("ID", "")}</h4>
            <b>Estado nacional:</b> {estado}<br>
            <b>Estado regional:</b> {row.get("Estado Verificación Regional", "")}<br>
            <b>Dirección Regional:</b> {row.get("Dirección Regional", "")}<br>
            <b>Delegación:</b> {row.get("Delegación", "")}<br>
            <b>Programa:</b> {row.get("Programa", "")}<br>
            <b>Actividad:</b> {row.get("Actividad", "")}<br>
            <b>Fecha:</b> {row.get("Fecha Actividad", "")}<br>
            <b>Lugar:</b> {row.get("Lugar", "")}<br>
            <b>Participantes:</b> {row.get("Cantidad Participantes", "")}<br>
            <b>Observación validación:</b> {row.get("Observaciones Validación", "")}<br>
        </div>
        """

        folium.Marker(
            location=[
                float(row["Latitud_Num"]),
                float(row["Longitud_Num"])
            ],
            popup=folium.Popup(popup_html, max_width=350),
            tooltip=f'{row.get("Delegación", "")} - {estado}',
            icon=folium.Icon(
                color=color,
                icon="info-sign"
            )
        ).add_to(mapa)

    if len(df_mapa) > 1:
        bounds = [
            [
                df_mapa["Latitud_Num"].min(),
                df_mapa["Longitud_Num"].min()
            ],
            [
                df_mapa["Latitud_Num"].max(),
                df_mapa["Longitud_Num"].max()
            ]
        ]

        mapa.fit_bounds(bounds)

    return mapa


def mostrar_mapa_admin(df, key="mapa_admin"):
    st.markdown("### Mapa de actividades filtradas")

    tipo_mapa_admin = st.selectbox(
        "Tipo de mapa",
        [
            "OpenStreetMap",
            "Mapa claro",
            "Mapa oscuro",
            "Topográfico",
            "Satélite"
        ],
        key=f"tipo_mapa_{key}"
    )

    st.session_state["tipo_mapa_admin_actual"] = tipo_mapa_admin

    df_mapa = preparar_dataframe_mapa_admin(df)

    if df_mapa.empty:
        st.info("No hay registros con coordenadas para mostrar en el mapa.")

    mapa = crear_mapa_admin(
        df,
        tipo_mapa=tipo_mapa_admin
    )

    st_folium(
        mapa,
        height=560,
        use_container_width=True,
        key=key
    )


# ======================================================
# MÉTRICAS ADMIN
# ======================================================

def mostrar_metricas_admin(df):
    df_metricas = limpiar_dataframe_metricas_admin(df)

    total_registros = len(df_metricas)

    total_participantes = (
        int(df_metricas["Cantidad Participantes"].sum())
        if not df_metricas.empty
        else 0
    )

    total_aprobadas = (
        len(df_metricas[df_metricas["Estado Validación"] == "Aprobada"])
        if not df_metricas.empty
        else 0
    )

    total_pendientes = (
        len(df_metricas[df_metricas["Estado Validación"] == "Pendiente"])
        if not df_metricas.empty
        else 0
    )

    total_rechazadas = (
        len(df_metricas[df_metricas["Estado Validación"] == "Rechazada"])
        if not df_metricas.empty
        else 0
    )

    total_observadas = (
        len(df_metricas[df_metricas["Estado Validación"] == "Con observaciones"])
        if not df_metricas.empty
        else 0
    )

    col1, col2, col3, col4, col5 = st.columns(5)

    with col1:
        st.metric("Registros", total_registros)

    with col2:
        st.metric("Participantes", total_participantes)

    with col3:
        st.metric("Aprobadas", total_aprobadas)

    with col4:
        st.metric("Pendientes", total_pendientes)

    with col5:
        st.metric("Con observaciones", total_observadas)

    col6, col7 = st.columns(2)

    with col6:
        st.metric("Rechazadas", total_rechazadas)

    with col7:
        porcentaje_aprobacion = 0

        if total_registros > 0:
            porcentaje_aprobacion = (total_aprobadas / total_registros) * 100

        st.metric("Porcentaje aprobación", f"{porcentaje_aprobacion:.1f}%")


# ======================================================
# APLICAR DISEÑO INSTITUCIONAL A GRÁFICOS
# Evita gráficos negros y mantiene azul/dorado.
# ======================================================

def aplicar_estilo_grafico_institucional(fig):
    fig.update_layout(
        title_x=0.5,
        plot_bgcolor="white",
        paper_bgcolor="white",
        font=dict(
            color=COLOR_GRIS_OSCURO,
            size=14
        ),
        title_font=dict(
            color=COLOR_AZUL,
            size=20
        ),
        xaxis=dict(
            title_font=dict(color=COLOR_AZUL),
            tickfont=dict(color=COLOR_GRIS_OSCURO),
            gridcolor="#E6EAF2",
            zerolinecolor="#D7DCE5"
        ),
        yaxis=dict(
            title_font=dict(color=COLOR_AZUL),
            tickfont=dict(color=COLOR_GRIS_OSCURO),
            gridcolor="#E6EAF2",
            zerolinecolor="#D7DCE5"
        ),
        legend=dict(
            font=dict(color=COLOR_GRIS_OSCURO),
            bgcolor="rgba(255,255,255,0.85)"
        )
    )

    return fig


# ======================================================
# GRÁFICOS ADMIN
# ======================================================

def mostrar_graficos_admin(df):
    df_metricas = limpiar_dataframe_metricas_admin(df)

    if df_metricas.empty:
        st.info("No hay datos para graficar.")
        return

    st.markdown("### Gráficos de análisis")

    # ==================================================
    # GRÁFICO POR ESTADO DE VALIDACIÓN
    # ==================================================

    if "Estado Validación" in df_metricas.columns:
        conteo_estado = (
            df_metricas
            .groupby("Estado Validación")
            .size()
            .reset_index(name="Cantidad")
            .sort_values("Cantidad", ascending=False)
        )

        if not conteo_estado.empty:
            fig_estado = px.pie(
                conteo_estado,
                names="Estado Validación",
                values="Cantidad",
                title="Distribución por estado de validación",
                hole=0.35,
                color="Estado Validación",
                color_discrete_map={
                    "Pendiente": COLOR_AMARILLO,
                    "Aprobada": COLOR_VERDE,
                    "Rechazada": COLOR_ROJO,
                    "Con observaciones": COLOR_DORADO
                }
            )

            fig_estado.update_traces(
                textinfo="percent+label+value",
                marker=dict(
                    line=dict(
                        color="white",
                        width=2
                    )
                )
            )

            fig_estado = aplicar_estilo_grafico_institucional(fig_estado)

            st.plotly_chart(
                fig_estado,
                use_container_width=True
            )

    # ==================================================
    # GRÁFICO POR PROGRAMA
    # ==================================================

    if "Programa" in df_metricas.columns:
        conteo_programa = (
            df_metricas
            .groupby("Programa")
            .size()
            .reset_index(name="Cantidad")
            .sort_values("Cantidad", ascending=False)
        )

        if not conteo_programa.empty:
            fig_programa = px.bar(
                conteo_programa,
                x="Programa",
                y="Cantidad",
                title="Cantidad de registros por programa",
                text="Cantidad",
                color_discrete_sequence=[COLOR_DORADO]
            )

            fig_programa.update_traces(
                marker_line_color=COLOR_AZUL,
                marker_line_width=1.5,
                textposition="outside",
                textfont=dict(
                    color=COLOR_AZUL,
                    size=14
                )
            )

            fig_programa.update_yaxes(
                rangemode="tozero",
                dtick=1
            )

            fig_programa = aplicar_estilo_grafico_institucional(fig_programa)

            st.plotly_chart(
                fig_programa,
                use_container_width=True
            )

    # ==================================================
    # GRÁFICO POR DIRECCIÓN REGIONAL
    # ==================================================

    if "Dirección Regional" in df_metricas.columns:
        conteo_region = (
            df_metricas
            .groupby("Dirección Regional")
            .size()
            .reset_index(name="Cantidad")
            .sort_values("Cantidad", ascending=False)
        )

        if not conteo_region.empty:
            fig_region = px.bar(
                conteo_region,
                x="Dirección Regional",
                y="Cantidad",
                title="Cantidad de registros por Dirección Regional",
                text="Cantidad",
                color_discrete_sequence=[COLOR_AZUL]
            )

            fig_region.update_traces(
                marker_line_color=COLOR_DORADO,
                marker_line_width=1.5,
                textposition="outside",
                textfont=dict(
                    color=COLOR_AZUL,
                    size=14
                )
            )

            fig_region.update_yaxes(
                rangemode="tozero",
                dtick=1
            )

            fig_region.update_xaxes(
                tickangle=-30
            )

            fig_region = aplicar_estilo_grafico_institucional(fig_region)

            st.plotly_chart(
                fig_region,
                use_container_width=True
            )

    # ==================================================
    # GRÁFICO POR DELEGACIÓN
    # ==================================================

    if "Delegación" in df_metricas.columns:
        conteo_delegacion = (
            df_metricas
            .groupby("Delegación")
            .size()
            .reset_index(name="Cantidad")
            .sort_values("Cantidad", ascending=False)
            .head(15)
        )

        if not conteo_delegacion.empty:
            fig_delegacion = px.bar(
                conteo_delegacion,
                x="Delegación",
                y="Cantidad",
                title="Top delegaciones por cantidad de registros",
                text="Cantidad",
                color_discrete_sequence=[COLOR_DORADO]
            )

            fig_delegacion.update_traces(
                marker_line_color=COLOR_AZUL,
                marker_line_width=1.5,
                textposition="outside",
                textfont=dict(
                    color=COLOR_AZUL,
                    size=14
                )
            )

            fig_delegacion.update_yaxes(
                rangemode="tozero",
                dtick=1
            )

            fig_delegacion.update_xaxes(
                tickangle=-30
            )

            fig_delegacion = aplicar_estilo_grafico_institucional(fig_delegacion)

            st.plotly_chart(
                fig_delegacion,
                use_container_width=True
            )
# ======================================================
# appADMIN.py
# PARTE 3 DE 5 CORREGIDA
# MÓDULO DE VALIDACIÓN INDIVIDUAL DE ACTIVIDADES,
# ACTUALIZACIÓN DE ESTADOS Y DESCARGA DE EXCEL VALIDADO
# FECHA DE VALIDACIÓN SIN HORA
# ======================================================


# ======================================================
# ACTUALIZAR VALIDACIÓN POR ID
# ======================================================

def actualizar_validacion_registro(
    indice_registro,
    estado_validacion,
    observaciones_validacion,
    funcionario_validador
):
    df = st.session_state.df_admin.copy()

    try:
        indice_registro = int(indice_registro)
    except Exception:
        return False

    if indice_registro not in df.index:
        return False

    df.loc[indice_registro, "Estado Validación"] = estado_validacion
    df.loc[indice_registro, "Observaciones Validación"] = observaciones_validacion
    df.loc[indice_registro, "Funcionario Validador"] = funcionario_validador
    df.loc[indice_registro, "Fecha Validación"] = datetime.now().strftime("%d/%m/%Y")

    st.session_state.df_admin = preparar_dataframe_admin(df.reset_index(drop=True))

    return True


def etiqueta_registro_admin(df, indice):
    try:
        fila = df.loc[indice]
    except Exception:
        return str(indice)

    delegacion = str(fila.get("Delegación", "")).strip() or "Delegación sin dato"
    id_registro = str(fila.get("ID", "")).strip() or "Sin ID"
    archivo = str(fila.get("Archivo Origen", "")).strip()
    actividad = str(fila.get("Actividad", "")).strip()

    etiqueta = f"{delegacion} | ID {id_registro}"
    if archivo:
        etiqueta += f" | Archivo: {archivo}"
    if actividad:
        etiqueta += f" | {actividad}"
    return etiqueta


# ======================================================
# TARJETAS RESUMEN DE VALIDACIÓN
# ======================================================

def mostrar_resumen_validacion(df):
    if df.empty:
        return

    total = len(df)
    aprobadas = len(df[df["Estado Validación"] == "Aprobada"])
    rechazadas = len(df[df["Estado Validación"] == "Rechazada"])
    pendientes = len(df[df["Estado Validación"] == "Pendiente"])
    observadas = len(df[df["Estado Validación"] == "Con observaciones"])

    col1, col2, col3, col4, col5 = st.columns(5)

    with col1:
        st.metric("Total filtrado", total)

    with col2:
        st.metric("Aprobadas", aprobadas)

    with col3:
        st.metric("Pendientes", pendientes)

    with col4:
        st.metric("Rechazadas", rechazadas)

    with col5:
        st.metric("Con observaciones", observadas)


# ======================================================
# MOSTRAR BADGE DE ESTADO
# ======================================================

def mostrar_badge_estado(estado):
    color = COLORES_VALIDACION.get(estado, COLOR_AMARILLO)

    st.markdown(
        f"""
        <div style="
            background:{color};
            color:white;
            padding:9px 16px;
            border-radius:14px;
            text-align:center;
            font-weight:900;
            margin-bottom:12px;
            box-shadow:0px 4px 10px rgba(0,0,0,0.16);
        ">
            Estado actual: {estado}
        </div>
        """,
        unsafe_allow_html=True
    )


# ======================================================
# MOSTRAR TARJETA DEL REGISTRO
# ======================================================

def mostrar_tarjeta_registro_validacion(fila):
    st.markdown(
        """
        <div class="card-validacion">
            <div class="texto-admin">
                Detalle del registro seleccionado para revisión administrativa.
            </div>
        </div>
        """,
        unsafe_allow_html=True
    )

    col1, col2 = st.columns(2)

    with col1:
        st.markdown("#### Datos generales")
        st.write(f"**ID:** {fila.get('ID', '')}")
        st.write(f"**Fecha:** {fila.get('Fecha Actividad', '')}")
        st.write(f"**Hora:** {fila.get('Hora Actividad', '')}")
        st.write(f"**Dirección Regional:** {fila.get('Dirección Regional', '')}")
        st.write(f"**Delegación:** {fila.get('Delegación', '')}")
        st.write(f"**Responsable:** {fila.get('Responsable', '')}")
        st.write(f"**Usuario Registra:** {fila.get('Usuario Registra', '')}")

    with col2:
        st.markdown("#### Actividad")
        st.write(f"**Responde a:** {fila.get('Responde a', '')}")
        st.write(f"**Programa:** {fila.get('Programa', '')}")
        st.write(f"**Actividad:** {fila.get('Actividad', '')}")
        st.write(f"**Lugar:** {fila.get('Lugar', '')}")
        st.write(f"**Provincia:** {fila.get('Provincia', '')}")
        st.write(f"**Cantón:** {fila.get('Cantón', '')}")
        st.write(f"**Distrito:** {fila.get('Distrito', '')}")

    st.markdown("#### Participantes")

    colp1, colp2, colp3, colp4 = st.columns(4)

    with colp1:
        st.metric("Total", fila.get("Cantidad Participantes", 0))

    with colp2:
        st.metric("Hombres", fila.get("Cantidad Hombres", 0))

    with colp3:
        st.metric("Mujeres", fila.get("Cantidad Mujeres", 0))

    with colp4:
        st.metric("Referencia", fila.get("Número de Referencia", ""))

    colr1, colr2, colr3, colr4 = st.columns(4)

    with colr1:
        st.metric("Edad 10 a 18", fila.get("Edad 10 a 18", 0))

    with colr2:
        st.metric("Edad 19 a 30", fila.get("Edad 19 a 30", 0))

    with colr3:
        st.metric("Edad 31 a 45", fila.get("Edad 31 a 45", 0))

    with colr4:
        st.metric("Edad 46+", fila.get("Edad 46 en adelante", 0))

    st.markdown("#### Verificación regional previa")
    colvr1, colvr2 = st.columns(2)
    with colvr1:
        st.write(f"**Estado regional:** {fila.get('Estado Verificación Regional', '')}")
        st.write(f"**Coordinador regional:** {fila.get('Coordinador Regional', '')}")
    with colvr2:
        st.write(f"**Fecha verificación regional:** {fila.get('Fecha Verificación Regional', '')}")
        st.write(f"**Observaciones regionales:** {fila.get('Observaciones Verificación Regional', '')}")

    st.markdown("#### Observaciones del registro original")

    observacion_original = str(fila.get("Observaciones", "")).strip()

    if observacion_original:
        st.info(observacion_original)
    else:
        st.info("Sin observaciones registradas.")

    st.markdown("#### Datos de referencia")

    colref1, colref2 = st.columns(2)

    with colref1:
        st.write(f"**Número de referencia:** {fila.get('Número de Referencia', '')}")
        st.write(f"**Expediente referencia:** {fila.get('Número de Expediente Referencia', '')}")

    with colref2:
        st.write(f"**Instituciones participantes:** {fila.get('Instituciones Participantes', '')}")
        st.write(f"**Dirección mapa:** {fila.get('Dirección Mapa', '')}")


# ======================================================
# MÓDULO PRINCIPAL DE VALIDACIÓN INDIVIDUAL
# ======================================================

def modulo_validacion_actividades(df_filtrado):
    st.markdown("## Validación individual de actividades")

    if df_filtrado.empty:
        st.info("No hay registros disponibles con los filtros aplicados.")
        return

    mostrar_resumen_validacion(df_filtrado)

    st.markdown("### Selección del registro")

    df_filtrado = df_filtrado[df_filtrado["Estado Verificación Regional"].astype(str) == "Verificada para envío"].copy()

    if df_filtrado.empty:
        st.warning("No hay actividades verificadas por la Dirección Regional para validación nacional con los filtros aplicados.")
        return

    indices_filtrados = df_filtrado.index.tolist()

    indice_validar = st.selectbox(
        "Seleccione el registro a validar",
        indices_filtrados,
        format_func=lambda idx: etiqueta_registro_admin(df_filtrado, idx),
        key="indice_validar_individual_admin"
    )

    if indice_validar not in df_filtrado.index:
        st.warning("No se encontró el registro seleccionado.")
        return

    fila = df_filtrado.loc[indice_validar].to_dict()

    estado_actual = fila.get("Estado Validación", "Pendiente")

    mostrar_badge_estado(estado_actual)

    with st.expander("Ver detalle completo del registro", expanded=True):
        mostrar_tarjeta_registro_validacion(fila)

        lat = str(fila.get("Latitud", "")).strip()
        lon = str(fila.get("Longitud", "")).strip()

        if lat and lon:
            st.markdown("#### Ubicación del registro")
            df_unico = pd.DataFrame([fila])
            mostrar_mapa_admin(
                df_unico,
                key=f"mapa_validacion_{indice_validar}"
            )
        else:
            st.info("Este registro no tiene coordenadas disponibles.")

    st.markdown("---")
    st.markdown("### Actualizar validación del registro")

    estados = [
        "Pendiente",
        "Aprobada",
        "Rechazada",
        "Con observaciones"
    ]

    indice_estado = estados.index(estado_actual) if estado_actual in estados else 0

    colv1, colv2 = st.columns(2)

    with colv1:
        nuevo_estado = st.selectbox(
            "Estado de validación",
            estados,
            index=indice_estado,
            key="nuevo_estado_individual"
        )

    with colv2:
        funcionario_validador = st.text_input(
            "Funcionario validador",
            value=str(fila.get("Funcionario Validador", "")),
            key="funcionario_validador_individual"
        )

    observaciones_validacion = st.text_area(
        "Observaciones de validación",
        value=str(fila.get("Observaciones Validación", "")),
        key="observaciones_validacion_individual"
    )

    if nuevo_estado in ["Rechazada", "Con observaciones"] and observaciones_validacion.strip() == "":
        st.warning("Para registros rechazados o con observaciones, se recomienda indicar claramente el motivo.")

    if st.button("💾 Guardar validación del registro"):
        if funcionario_validador.strip() == "":
            st.warning("Debe indicar el funcionario validador.")
        elif nuevo_estado in ["Rechazada", "Con observaciones"] and observaciones_validacion.strip() == "":
            st.warning("Debe agregar una observación cuando el estado sea Rechazada o Con observaciones.")
        else:
            actualizado = actualizar_validacion_registro(
                indice_registro=indice_validar,
                estado_validacion=nuevo_estado,
                observaciones_validacion=observaciones_validacion,
                funcionario_validador=funcionario_validador
            )

            if actualizado:
                st.success("Validación guardada correctamente.")
                st.rerun()
            else:
                st.error("No se pudo guardar la validación.")


# ======================================================
# CATÁLOGOS BASE PARA EDICIÓN ADMIN
# Estas funciones usan las mismas bases de la app donde
# las delegaciones registran: Datos Importantes.xlsx y
# BASE DE DATOS MEP 2025.xlsx.
# ======================================================

def obtener_columna_por_nombre(df, posibles_nombres):
    columnas = list(df.columns)
    for posible in posibles_nombres:
        posible_norm = normalizar_texto(posible)
        for col in columnas:
            if normalizar_texto(col) == posible_norm:
                return col
    return None


def ordenar_regiones_numericamente(lista_regiones):
    def extraer_numero_region(region):
        texto = str(region)
        numero = ""
        for caracter in texto:
            if caracter.isdigit():
                numero += caracter
            elif numero:
                break
        return int(numero) if numero else 999
    return sorted(lista_regiones, key=extraer_numero_region)


def separar_responde_a(valor):
    if pd.isna(valor):
        return []
    texto = str(valor).strip()
    if texto == "" or texto.lower() == "nan":
        return []
    return [parte.strip() for parte in texto.split("/") if parte.strip()]


@st.cache_data
def cargar_datos_importantes():
    columnas_base = [
        "Provincia", "Cantón", "Distrito", "Dirección Regional",
        "Delegación", "Responde a", "Actividad Realizada", "Programa"
    ]

    if not os.path.exists(ARCHIVO_DATOS_IMPORTANTES):
        return pd.DataFrame(columns=columnas_base)

    try:
        df_original = pd.read_excel(ARCHIVO_DATOS_IMPORTANTES)

        col_provincia = obtener_columna_por_nombre(df_original, ["Provincia"])
        col_canton = obtener_columna_por_nombre(df_original, ["Cantón", "Canton"])
        col_distrito = obtener_columna_por_nombre(df_original, ["Distrito", "Distritos"])
        col_region = obtener_columna_por_nombre(df_original, ["Dirección Regional", "Direccion Regional"])
        col_delegacion = obtener_columna_por_nombre(df_original, ["Delegación", "Delegacion"])
        col_responde_a = obtener_columna_por_nombre(df_original, ["Responde a", "Responde a:", "Responde", "Responda a", "Responda a:"])
        col_actividad = obtener_columna_por_nombre(df_original, ["Actividad Realizada", "Actividad"])
        col_programa = obtener_columna_por_nombre(df_original, ["Programa"])

        columnas_requeridas = [
            col_provincia, col_canton, col_distrito, col_region,
            col_delegacion, col_responde_a, col_actividad, col_programa
        ]

        if not all(columnas_requeridas):
            st.warning(
                "El archivo Datos Importantes.xlsx no tiene todas las columnas requeridas. "
                "Debe incluir: Provincia, Cantón, Distrito, Dirección Regional, Delegación, "
                "Responde a, Actividad Realizada y Programa."
            )
            return pd.DataFrame(columns=columnas_base)

        df = df_original[[
            col_provincia, col_canton, col_distrito, col_region,
            col_delegacion, col_responde_a, col_actividad, col_programa
        ]].copy()
        df.columns = columnas_base

        for col in columnas_base:
            df[col] = df[col].fillna("").astype(str).str.strip()

        df = df.replace("nan", "")
        df = df.drop_duplicates()

        df["Provincia_Normalizada"] = df["Provincia"].apply(normalizar_texto)
        df["Cantón_Normalizado"] = df["Cantón"].apply(normalizar_texto)
        df["Distrito_Normalizado"] = df["Distrito"].apply(normalizar_texto)
        df["Región_Normalizada"] = df["Dirección Regional"].apply(normalizar_texto)
        df["Delegación_Normalizada"] = df["Delegación"].apply(normalizar_texto)
        df["Responde_a_Normalizado"] = df["Responde a"].apply(normalizar_texto)
        df["Actividad_Normalizada"] = df["Actividad Realizada"].apply(normalizar_texto)
        df["Programa_Normalizado"] = df["Programa"].apply(normalizar_texto)

        filas_expandidas = []
        for _, fila in df.iterrows():
            opciones_responde = separar_responde_a(fila["Responde a"])
            if not opciones_responde:
                opciones_responde = [fila["Responde a"]]
            for opcion in opciones_responde:
                nueva_fila = fila.copy()
                nueva_fila["Responde a"] = opcion
                nueva_fila["Responde_a_Normalizado"] = normalizar_texto(opcion)
                filas_expandidas.append(nueva_fila)

        if filas_expandidas:
            df = pd.DataFrame(filas_expandidas).drop_duplicates()

        return df

    except Exception as e:
        st.error(f"Error leyendo {ARCHIVO_DATOS_IMPORTANTES}")
        st.exception(e)
        return pd.DataFrame(columns=columnas_base)


def agregar_valor_actual(opciones, valor_actual):
    opciones = [str(x).strip() for x in opciones if str(x).strip() and str(x).strip().lower() != "nan"]
    opciones = list(dict.fromkeys(opciones))
    valor_actual = str(valor_actual).strip()
    if valor_actual and valor_actual.lower() != "nan" and valor_actual not in opciones:
        opciones.insert(0, valor_actual)
    if not opciones:
        opciones = [valor_actual] if valor_actual else [""]
    return opciones


def obtener_regiones_datos(valor_actual=""):
    df = cargar_datos_importantes()
    if df.empty:
        return agregar_valor_actual(ordenar_regiones_numericamente(REGIONES_BASE), valor_actual)
    regiones = df["Dirección Regional"].dropna().unique().tolist()
    regiones = [x for x in regiones if str(x).strip()]
    return agregar_valor_actual(ordenar_regiones_numericamente(regiones), valor_actual)


def obtener_delegaciones_por_region(region, valor_actual=""):
    df = cargar_datos_importantes()
    if df.empty or not region:
        return agregar_valor_actual([], valor_actual)
    region_norm = normalizar_texto(region)
    delegaciones = df[df["Región_Normalizada"] == region_norm]["Delegación"].dropna().unique().tolist()
    return agregar_valor_actual(sorted(delegaciones), valor_actual)


def obtener_responde_a_datos(valor_actual=""):
    df = cargar_datos_importantes()
    if df.empty or "Responde a" not in df.columns:
        return agregar_valor_actual([], valor_actual)
    responde_a = df["Responde a"].dropna().unique().tolist()
    return agregar_valor_actual(sorted(responde_a), valor_actual)


def obtener_programas_por_responde_a(responde_a, valor_actual=""):
    df = cargar_datos_importantes()
    if df.empty or not responde_a:
        return agregar_valor_actual(PROGRAMAS_BASE, valor_actual)
    responde_norm = normalizar_texto(responde_a)
    programas = df[df["Responde_a_Normalizado"] == responde_norm]["Programa"].dropna().unique().tolist()

    programas_limpios = []
    for programa in programas:
        if normalizar_texto(programa) == "GREAT CAMP":
            if "GREAT" not in programas_limpios:
                programas_limpios.append("GREAT")
        elif programa not in programas_limpios:
            programas_limpios.append(programa)

    orden_oficial = [p for p in PROGRAMAS_BASE if p in programas_limpios]
    extras = sorted([p for p in programas_limpios if p not in orden_oficial])
    return agregar_valor_actual(orden_oficial + extras, valor_actual)


def obtener_actividades_por_responde_a_programa(responde_a, programa, valor_actual=""):
    df = cargar_datos_importantes()
    if df.empty or not responde_a or not programa:
        return agregar_valor_actual([], valor_actual)

    responde_norm = normalizar_texto(responde_a)
    programa_norm = normalizar_texto(programa)

    if programa_norm == "GREAT":
        actividades = df[
            (df["Responde_a_Normalizado"] == responde_norm) &
            (df["Programa_Normalizado"].isin(["GREAT", "GREAT CAMP"]))
        ]["Actividad Realizada"].dropna().unique().tolist()
    else:
        actividades = df[
            (df["Responde_a_Normalizado"] == responde_norm) &
            (df["Programa_Normalizado"] == programa_norm)
        ]["Actividad Realizada"].dropna().unique().tolist()

    return agregar_valor_actual(sorted(actividades), valor_actual)


def obtener_provincias_datos(valor_actual=""):
    df = cargar_datos_importantes()
    if df.empty:
        return agregar_valor_actual(PROVINCIAS_BASE, valor_actual)
    provincias = df["Provincia"].dropna().unique().tolist()
    return agregar_valor_actual(sorted(provincias), valor_actual)


def obtener_cantones_por_provincia(provincia, valor_actual=""):
    df = cargar_datos_importantes()
    if df.empty or not provincia:
        return agregar_valor_actual([], valor_actual)
    provincia_norm = normalizar_texto(provincia)
    cantones = df[df["Provincia_Normalizada"] == provincia_norm]["Cantón"].dropna().unique().tolist()
    return agregar_valor_actual(sorted(cantones), valor_actual)


def obtener_distritos_por_provincia_canton(provincia, canton, valor_actual=""):
    df = cargar_datos_importantes()
    if df.empty or not provincia or not canton:
        return agregar_valor_actual([], valor_actual)
    provincia_norm = normalizar_texto(provincia)
    canton_norm = normalizar_texto(canton)
    distritos = df[
        (df["Provincia_Normalizada"] == provincia_norm) &
        (df["Cantón_Normalizado"] == canton_norm)
    ]["Distrito"].dropna().unique().tolist()
    return agregar_valor_actual(sorted(distritos), valor_actual)


@st.cache_data
def cargar_base_mep():
    columnas_base = ["PROVINCIA", "NOMBRE", "CODIGO_PRESUPUESTARIO"]
    if not os.path.exists(ARCHIVO_MEP):
        return pd.DataFrame(columns=columnas_base)
    try:
        df_original = pd.read_excel(ARCHIVO_MEP)
        col_provincia = obtener_columna_por_nombre(df_original, ["PROVINCIA", "Provincia"])
        col_nombre = obtener_columna_por_nombre(df_original, ["NOMBRE", "Nombre", "CENTRO EDUCATIVO", "Centro Educativo", "INSTITUCION", "Institución"])
        col_codigo = obtener_columna_por_nombre(df_original, ["CODIGO PRESUPUESTARIO", "CÓDIGO PRESUPUESTARIO", "Codigo Presupuestario", "Código Presupuestario", "CODIGO", "CÓDIGO"])

        if not col_provincia or not col_nombre:
            return pd.DataFrame(columns=columnas_base)
        if not col_codigo:
            df_original["CODIGO_PRESUPUESTARIO_TMP"] = ""
            col_codigo = "CODIGO_PRESUPUESTARIO_TMP"

        df = df_original[[col_provincia, col_nombre, col_codigo]].copy()
        df.columns = columnas_base
        for col in columnas_base:
            df[col] = df[col].fillna("").astype(str).str.strip()

        df["PROVINCIA_NORMALIZADA"] = df["PROVINCIA"].apply(normalizar_texto)
        df["NOMBRE_NORMALIZADO"] = df["NOMBRE"].apply(normalizar_texto)
        df["CENTRO_MOSTRAR"] = df.apply(
            lambda row: f"{row['NOMBRE']} - Código: {row['CODIGO_PRESUPUESTARIO']}"
            if row["CODIGO_PRESUPUESTARIO"].strip() != "" else row["NOMBRE"],
            axis=1
        )
        return df.drop_duplicates(subset=["PROVINCIA_NORMALIZADA", "NOMBRE_NORMALIZADO", "CODIGO_PRESUPUESTARIO"])
    except Exception as e:
        st.error(f"Error leyendo {ARCHIVO_MEP}")
        st.exception(e)
        return pd.DataFrame(columns=columnas_base)


def obtener_centros_por_provincia(provincia, valor_actual=""):
    df = cargar_base_mep()
    if df.empty or not provincia:
        return agregar_valor_actual([], valor_actual)
    provincia_norm = normalizar_texto(provincia)
    centros = df[df["PROVINCIA_NORMALIZADA"] == provincia_norm]["CENTRO_MOSTRAR"].dropna().astype(str).drop_duplicates().sort_values().tolist()
    return agregar_valor_actual(centros, valor_actual)


def obtener_datos_centro_educativo(centro_mostrar):
    df = cargar_base_mep()
    if df.empty or not centro_mostrar:
        return centro_mostrar, ""
    fila = df[df["CENTRO_MOSTRAR"] == centro_mostrar]
    if fila.empty:
        return centro_mostrar, ""
    nombre = fila.iloc[0].get("NOMBRE", "")
    codigo = fila.iloc[0].get("CODIGO_PRESUPUESTARIO", "")
    return nombre, codigo


def selectbox_con_valor_actual(label, opciones, valor_actual, key):
    opciones = agregar_valor_actual(opciones, valor_actual)
    return st.selectbox(
        label,
        opciones,
        index=indice_opcion_seguro(opciones, valor_actual),
        key=key
    )

# ======================================================
# EDITAR Y ELIMINAR REGISTROS ADMINISTRATIVOS
# ======================================================

def eliminar_registro_admin(indice_registro):
    df = st.session_state.df_admin.copy()

    try:
        indice_registro = int(indice_registro)
    except Exception:
        return False

    if indice_registro not in df.index:
        return False

    df = df.drop(index=indice_registro).reset_index(drop=True)
    st.session_state.df_admin = preparar_dataframe_admin(df)
    return True


def actualizar_registro_admin_completo(indice_registro, nuevos_datos):
    df = st.session_state.df_admin.copy()

    try:
        indice_registro = int(indice_registro)
    except Exception:
        return False

    if indice_registro not in df.index:
        return False

    # Evita errores de pandas cuando una columna quedó como numérica y se
    # intenta actualizar con texto desde la edición.
    df = df.astype("object")

    for col, valor in nuevos_datos.items():
        if col in df.columns:
            df.loc[indice_registro, col] = valor

    st.session_state.df_admin = preparar_dataframe_admin(df.reset_index(drop=True))
    return True


def convertir_entero_seguro(valor):
    try:
        return int(pd.to_numeric(valor, errors="coerce"))
    except Exception:
        return 0



def obtener_opciones_columna_admin(columna, df_base=None, filtros=None, valor_actual=""):
    """Devuelve opciones limpias para selectbox usando los datos cargados."""
    if df_base is None:
        df_base = st.session_state.df_admin.copy()
    else:
        df_base = df_base.copy()

    if filtros:
        for col_filtro, valor_filtro in filtros.items():
            if col_filtro in df_base.columns and str(valor_filtro).strip() != "":
                df_base = df_base[
                    df_base[col_filtro].astype(str).apply(normalizar_texto)
                    == normalizar_texto(valor_filtro)
                ]

    opciones = []
    if columna in df_base.columns:
        opciones = (
            df_base[columna]
            .replace("", pd.NA)
            .dropna()
            .astype(str)
            .map(lambda x: x.strip())
            .drop_duplicates()
            .tolist()
        )

    opciones = [x for x in opciones if x and x.lower() != "nan"]
    opciones = sorted(opciones)

    valor_actual = str(valor_actual).strip()
    if valor_actual and valor_actual.lower() != "nan" and valor_actual not in opciones:
        opciones.insert(0, valor_actual)

    if not opciones:
        opciones = [valor_actual] if valor_actual else [""]

    return opciones


def indice_opcion_seguro(opciones, valor_actual):
    valor_actual = str(valor_actual).strip()
    try:
        return opciones.index(valor_actual)
    except Exception:
        return 0


def selectbox_edicion_admin(label, columna, fila, key, df_base=None, filtros=None):
    valor_actual = str(fila.get(columna, "")).strip()
    opciones = obtener_opciones_columna_admin(
        columna=columna,
        df_base=df_base,
        filtros=filtros,
        valor_actual=valor_actual
    )
    return st.selectbox(
        label,
        opciones,
        index=indice_opcion_seguro(opciones, valor_actual),
        key=key
    )

def modulo_editar_eliminar_registros(df_filtrado):
    st.markdown("## Editar o eliminar registros")

    if df_filtrado.empty:
        st.info("No hay registros disponibles con los filtros aplicados.")
        return

    st.markdown(
        """
        <div class="card-validacion">
            <div class="texto-admin">
                Revise, edite o elimine registros antes de descargar el Excel validado.
            </div>
        </div>
        """,
        unsafe_allow_html=True
    )

    indices_filtrados = df_filtrado.index.tolist()

    indice_editar = st.selectbox(
        "Seleccione el registro a editar o eliminar",
        indices_filtrados,
        format_func=lambda idx: etiqueta_registro_admin(df_filtrado, idx),
        key="indice_editar_eliminar_admin"
    )

    if indice_editar not in df_filtrado.index:
        st.warning("No se encontró el registro seleccionado.")
        return

    fila = df_filtrado.loc[indice_editar].to_dict()

    mostrar_badge_estado(fila.get("Estado Validación", "Pendiente"))

    with st.expander("Ver detalle actual del registro", expanded=False):
        mostrar_tarjeta_registro_validacion(fila)

    st.markdown("### Editar registro")
    # IMPORTANTE:
    # No se usa st.form en esta sección porque dentro de un formulario Streamlit
    # no actualiza los selectbox dependientes hasta presionar guardar.
    # Al dejar los widgets fuera del form, cada cambio refresca la pantalla y
    # los catálogos se actualizan inmediatamente, igual que en la app de delegaciones.

    col1, col2 = st.columns(2)

    with col1:
        fecha_actividad = st.text_input(
            "Fecha Actividad",
            value=str(fila.get("Fecha Actividad", "")),
            key=f"edit_fecha_{indice_editar}"
        )

        hora_actividad = st.text_input(
            "Hora Actividad",
            value=str(fila.get("Hora Actividad", "")),
            key=f"edit_hora_{indice_editar}"
        )

        direccion_admin = selectbox_con_valor_actual(
            "Dirección Regional",
            obtener_regiones_datos(fila.get("Dirección Regional", "")),
            fila.get("Dirección Regional", ""),
            key=f"edit_region_{indice_editar}"
        )

        valor_delegacion_actual = (
            fila.get("Delegación", "")
            if normalizar_texto(direccion_admin) == normalizar_texto(fila.get("Dirección Regional", ""))
            else ""
        )

        delegacion = selectbox_con_valor_actual(
            "Delegación",
            obtener_delegaciones_por_region(direccion_admin, valor_delegacion_actual),
            valor_delegacion_actual,
            key=f"edit_delegacion_{indice_editar}"
        )

        responde_a = selectbox_con_valor_actual(
            "Responde a",
            obtener_responde_a_datos(fila.get("Responde a", "")),
            fila.get("Responde a", ""),
            key=f"edit_responde_{indice_editar}"
        )

        valor_programa_actual = (
            fila.get("Programa", "")
            if normalizar_texto(responde_a) == normalizar_texto(fila.get("Responde a", ""))
            else ""
        )

        programa = selectbox_con_valor_actual(
            "Programa",
            obtener_programas_por_responde_a(responde_a, valor_programa_actual),
            valor_programa_actual,
            key=f"edit_programa_{indice_editar}"
        )

        valor_actividad_actual = (
            fila.get("Actividad", "")
            if (
                normalizar_texto(responde_a) == normalizar_texto(fila.get("Responde a", "")) and
                normalizar_texto(programa) == normalizar_texto(fila.get("Programa", ""))
            )
            else ""
        )

        actividad = selectbox_con_valor_actual(
            "Actividad",
            obtener_actividades_por_responde_a_programa(responde_a, programa, valor_actividad_actual),
            valor_actividad_actual,
            key=f"edit_actividad_{indice_editar}"
        )

        provincia = selectbox_con_valor_actual(
            "Provincia",
            obtener_provincias_datos(fila.get("Provincia", "")),
            fila.get("Provincia", ""),
            key=f"edit_provincia_{indice_editar}"
        )

        valor_canton_actual = (
            fila.get("Cantón", "")
            if normalizar_texto(provincia) == normalizar_texto(fila.get("Provincia", ""))
            else ""
        )

        canton = selectbox_con_valor_actual(
            "Cantón",
            obtener_cantones_por_provincia(provincia, valor_canton_actual),
            valor_canton_actual,
            key=f"edit_canton_{indice_editar}"
        )

        valor_distrito_actual = (
            fila.get("Distrito", "")
            if (
                normalizar_texto(provincia) == normalizar_texto(fila.get("Provincia", "")) and
                normalizar_texto(canton) == normalizar_texto(fila.get("Cantón", ""))
            )
            else ""
        )

        distrito = selectbox_con_valor_actual(
            "Distrito",
            obtener_distritos_por_provincia_canton(provincia, canton, valor_distrito_actual),
            valor_distrito_actual,
            key=f"edit_distrito_{indice_editar}"
        )

        tipo_lugar_opciones = ["Centro educativo", "Otro lugar"]
        tipo_lugar_actual = str(fila.get("Tipo Lugar", "")).strip()
        if tipo_lugar_actual and tipo_lugar_actual not in tipo_lugar_opciones:
            tipo_lugar_opciones.insert(0, tipo_lugar_actual)

        tipo_lugar = st.selectbox(
            "Tipo Lugar",
            tipo_lugar_opciones,
            index=indice_opcion_seguro(tipo_lugar_opciones, tipo_lugar_actual),
            key=f"edit_tipo_lugar_{indice_editar}"
        )

        if tipo_lugar == "Centro educativo":
            valor_centro_actual = (
                fila.get("Centro Educativo", "")
                if normalizar_texto(provincia) == normalizar_texto(fila.get("Provincia", ""))
                else ""
            )

            centro_opciones = obtener_centros_por_provincia(provincia, valor_centro_actual)

            centro_mostrar = selectbox_con_valor_actual(
                "Centro educativo según base MEP 2025",
                centro_opciones,
                valor_centro_actual,
                key=f"edit_centro_mep_{indice_editar}"
            )

            centro_educativo, codigo_presupuestario_auto = obtener_datos_centro_educativo(centro_mostrar)
            codigo_actual = str(fila.get("Código Presupuestario", "")).strip()

            codigo_presupuestario = (
                codigo_presupuestario_auto
                if codigo_presupuestario_auto
                else codigo_actual
            )

            lugar = centro_educativo

            st.info(f"Centro educativo seleccionado: {centro_educativo}")
            st.info(f"Código presupuestario: {codigo_presupuestario if codigo_presupuestario else 'No disponible'}")
        else:
            lugar = st.text_input(
                "Lugar",
                value=str(fila.get("Lugar", "")),
                key=f"edit_lugar_{indice_editar}"
            )
            centro_educativo = ""
            codigo_presupuestario = ""

        latitud = st.text_input(
            "Latitud",
            value=str(fila.get("Latitud", "")),
            key=f"edit_latitud_{indice_editar}"
        )

        longitud = st.text_input(
            "Longitud",
            value=str(fila.get("Longitud", "")),
            key=f"edit_longitud_{indice_editar}"
        )

    with col2:
        responsable = st.text_input(
            "Responsable",
            value=str(fila.get("Responsable", "")),
            key=f"edit_responsable_{indice_editar}"
        )

        cantidad_participantes = st.number_input(
            "Cantidad Participantes",
            min_value=0,
            step=1,
            value=convertir_entero_seguro(fila.get("Cantidad Participantes", 0)),
            key=f"edit_cantidad_{indice_editar}"
        )

        cantidad_hombres = st.number_input(
            "Cantidad Hombres",
            min_value=0,
            step=1,
            value=convertir_entero_seguro(fila.get("Cantidad Hombres", 0)),
            key=f"edit_hombres_{indice_editar}"
        )

        cantidad_mujeres = st.number_input(
            "Cantidad Mujeres",
            min_value=0,
            step=1,
            value=convertir_entero_seguro(fila.get("Cantidad Mujeres", 0)),
            key=f"edit_mujeres_{indice_editar}"
        )

        edad_10_18 = st.number_input(
            "Edad 10 a 18",
            min_value=0,
            step=1,
            value=convertir_entero_seguro(fila.get("Edad 10 a 18", 0)),
            key=f"edit_edad_10_18_{indice_editar}"
        )

        edad_19_30 = st.number_input(
            "Edad 19 a 30",
            min_value=0,
            step=1,
            value=convertir_entero_seguro(fila.get("Edad 19 a 30", 0)),
            key=f"edit_edad_19_30_{indice_editar}"
        )

        edad_31_45 = st.number_input(
            "Edad 31 a 45",
            min_value=0,
            step=1,
            value=convertir_entero_seguro(fila.get("Edad 31 a 45", 0)),
            key=f"edit_edad_31_45_{indice_editar}"
        )

        edad_46_mas = st.number_input(
            "Edad 46 en adelante",
            min_value=0,
            step=1,
            value=convertir_entero_seguro(fila.get("Edad 46 en adelante", 0)),
            key=f"edit_edad_46_{indice_editar}"
        )

        instituciones = st.text_area(
            "Instituciones Participantes",
            value=str(fila.get("Instituciones Participantes", "")),
            height=80,
            key=f"edit_instituciones_{indice_editar}"
        )

        numero_referencia = st.text_input(
            "Número de Referencia",
            value=str(fila.get("Número de Referencia", "")),
            key=f"edit_referencia_{indice_editar}"
        )

        numero_expediente = st.text_input(
            "Número de Expediente Referencia",
            value=str(fila.get("Número de Expediente Referencia", "")),
            key=f"edit_expediente_{indice_editar}"
        )

        observaciones = st.text_area(
            "Observaciones del registro",
            value=str(fila.get("Observaciones", "")),
            height=90,
            key=f"edit_observaciones_{indice_editar}"
        )

        usuario_registra = st.text_input(
            "Usuario Registra",
            value=str(fila.get("Usuario Registra", "")),
            key=f"edit_usuario_{indice_editar}"
        )

        archivo_origen = st.text_input(
            "Archivo Origen",
            value=str(fila.get("Archivo Origen", "")),
            disabled=True,
            key=f"edit_archivo_origen_{indice_editar}"
        )

    suma_sexo = cantidad_hombres + cantidad_mujeres
    suma_edades = edad_10_18 + edad_19_30 + edad_31_45 + edad_46_mas

    if cantidad_participantes > 0:
        if suma_sexo != cantidad_participantes:
            st.warning(
                f"La suma de hombres y mujeres ({suma_sexo}) no coincide con el total de participantes ({cantidad_participantes})."
            )
        if suma_edades != cantidad_participantes:
            st.warning(
                f"La suma de rangos de edad ({suma_edades}) no coincide con el total de participantes ({cantidad_participantes})."
            )

    st.markdown("#### Validación nacional")

    estados = [
        "Pendiente",
        "Aprobada",
        "Rechazada",
        "Con observaciones"
    ]

    estado_actual = str(fila.get("Estado Validación", "Pendiente"))
    indice_estado = estados.index(estado_actual) if estado_actual in estados else 0

    colv1, colv2 = st.columns(2)

    with colv1:
        estado_validacion = st.selectbox(
            "Estado Validación",
            estados,
            index=indice_estado,
            key=f"edit_estado_validacion_{indice_editar}"
        )

        funcionario_validador = st.text_input(
            "Funcionario Validador",
            value=str(fila.get("Funcionario Validador", "")),
            key=f"edit_funcionario_validador_{indice_editar}"
        )

    with colv2:
        fecha_validacion = st.text_input(
            "Fecha Validación",
            value=str(fila.get("Fecha Validación", "")),
            key=f"edit_fecha_validacion_{indice_editar}"
        )

        observaciones_validacion = st.text_area(
            "Observaciones Validación",
            value=str(fila.get("Observaciones Validación", "")),
            height=95,
            key=f"edit_obs_validacion_{indice_editar}"
        )

    if st.button("💾 Guardar cambios del registro", key=f"btn_guardar_edicion_{indice_editar}"):
        if cantidad_participantes > 0 and suma_sexo != cantidad_participantes:
            st.error("No se puede guardar: hombres + mujeres no coincide con el total de participantes.")
            return

        if cantidad_participantes > 0 and suma_edades != cantidad_participantes:
            st.error("No se puede guardar: los rangos de edad no coinciden con el total de participantes.")
            return

        nuevos_datos = {
            "Fecha Actividad": fecha_actividad,
            "Hora Actividad": hora_actividad,
            "Dirección Regional": direccion_admin,
            "Delegación": delegacion,
            "Responde a": responde_a,
            "Programa": programa,
            "Actividad": actividad,
            "Provincia": provincia,
            "Cantón": canton,
            "Distrito": distrito,
            "Tipo Lugar": tipo_lugar,
            "Lugar": lugar,
            "Centro Educativo": centro_educativo,
            "Código Presupuestario": codigo_presupuestario,
            "Dirección Mapa": "",
            "Latitud": latitud,
            "Longitud": longitud,
            "Responsable": responsable,
            "Cantidad Participantes": cantidad_participantes,
            "Cantidad Hombres": cantidad_hombres,
            "Cantidad Mujeres": cantidad_mujeres,
            "Edad 10 a 18": edad_10_18,
            "Edad 19 a 30": edad_19_30,
            "Edad 31 a 45": edad_31_45,
            "Edad 46 en adelante": edad_46_mas,
            "Instituciones Participantes": instituciones,
            "Número de Referencia": numero_referencia,
            "Número de Expediente Referencia": numero_expediente,
            "Observaciones": observaciones,
            "Usuario Registra": usuario_registra,
            "Estado Validación": estado_validacion,
            "Observaciones Validación": observaciones_validacion,
            "Funcionario Validador": funcionario_validador,
            "Fecha Validación": fecha_validacion,
        }

        actualizado = actualizar_registro_admin_completo(indice_editar, nuevos_datos)

        if actualizado:
            st.success("Registro actualizado correctamente.")
            st.rerun()
        else:
            st.error("No se pudo actualizar el registro.")

    st.markdown("---")
    st.markdown("### Eliminar registro")

    st.warning(
        "Esta acción elimina el registro de la sesión actual. "
        "El cambio se reflejará en el Excel que descargue después."
    )

    confirmar_eliminacion = st.checkbox(
        "Confirmo que deseo eliminar este registro",
        key=f"confirmar_eliminacion_admin_{indice_editar}"
    )

    if st.button("🗑️ Eliminar registro seleccionado", key=f"btn_eliminar_admin_{indice_editar}"):
        if not confirmar_eliminacion:
            st.warning("Debe marcar la confirmación antes de eliminar.")
        else:
            eliminado = eliminar_registro_admin(indice_editar)
            if eliminado:
                st.success("Registro eliminado correctamente.")
                st.rerun()
            else:
                st.error("No se pudo eliminar el registro.")


# ======================================================
# DESCARGA DE EXCEL ADMIN VALIDADO
# ======================================================

def boton_descargar_excel_admin(df, key="descarga_excel_admin"):
    if df is None or df.empty:
        st.info("No hay registros para descargar.")
        return

    nombre_archivo = generar_nombre_reporte(df, tipo="XLSX")

    st.download_button(
        "⬇️ Descargar Excel nacional validado",
        data=convertir_excel_admin(df),
        file_name=nombre_archivo,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=key
    )


# ======================================================
# TABLA RESUMIDA PARA REVISIÓN
# ======================================================

def mostrar_tabla_resumen_validacion(df):
    if df.empty:
        st.info("No hay registros para mostrar.")
        return

    columnas_resumen = [
        "Archivo Origen",
        "ID",
        "Fecha Actividad",
        "Dirección Regional",
        "Delegación",
        "Programa",
        "Actividad",
        "Cantidad Participantes",
        "Estado Verificación Regional",
        "Observaciones Verificación Regional",
        "Coordinador Regional",
        "Fecha Verificación Regional",
        "Estado Validación",
        "Observaciones Validación",
        "Funcionario Validador",
        "Fecha Validación"
    ]

    columnas_existentes = [
        col for col in columnas_resumen
        if col in df.columns
    ]

    st.dataframe(
        df[columnas_existentes],
        use_container_width=True,
        hide_index=True
    )


# ======================================================
# TABLA DETALLADA ADMINISTRATIVA
# ======================================================

def mostrar_tabla_detallada_admin(df):
    if df.empty:
        st.info("No hay registros para mostrar.")
        return

    st.dataframe(
        df,
        use_container_width=True,
        hide_index=True
    )

# ======================================================
# DASHBOARD NACIONAL POR PROGRAMA AUTORIZADO
# Muestra la información que compete al usuario autenticado
# por región, delegación, programa y estado.
# ======================================================

def mostrar_dashboard_coordinador_nacional(df):
    if df is None or df.empty:
        st.info("No hay datos para mostrar con el programa y filtros seleccionados.")
        return

    df_metricas = limpiar_dataframe_metricas_admin(df)
    programa_actual = st.session_state.get("admin_programa_autenticado", "")

    st.markdown("### Resumen nacional del programa")
    st.caption(f"Información visible para: {programa_actual}")

    total_registros = len(df_metricas)
    total_participantes = int(df_metricas["Cantidad Participantes"].sum()) if "Cantidad Participantes" in df_metricas.columns else 0
    total_regiones = df_metricas["Dirección Regional"].replace("", pd.NA).dropna().nunique() if "Dirección Regional" in df_metricas.columns else 0
    total_delegaciones = df_metricas["Delegación"].replace("", pd.NA).dropna().nunique() if "Delegación" in df_metricas.columns else 0
    total_aprobadas = len(df_metricas[df_metricas["Estado Validación"] == "Aprobada"]) if "Estado Validación" in df_metricas.columns else 0

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Registros visibles", total_registros)
    c2.metric("Participantes", total_participantes)
    c3.metric("Regiones", total_regiones)
    c4.metric("Delegaciones", total_delegaciones)
    c5.metric("Aprobadas", total_aprobadas)

    st.markdown("---")

    if "Estado Validación" in df_metricas.columns:
        estado = (
            df_metricas.groupby("Estado Validación")
            .size()
            .reset_index(name="Cantidad")
            .sort_values("Cantidad", ascending=False)
        )
        if not estado.empty:
            fig_estado = px.pie(
                estado,
                names="Estado Validación",
                values="Cantidad",
                title="Estado de validación nacional",
                hole=0.38,
                color="Estado Validación",
                color_discrete_map={
                    "Pendiente": COLOR_AMARILLO,
                    "Aprobada": COLOR_VERDE,
                    "Rechazada": COLOR_ROJO,
                    "Con observaciones": COLOR_DORADO,
                },
            )
            fig_estado.update_traces(textinfo="percent+label+value")
            fig_estado = aplicar_estilo_grafico_institucional(fig_estado)
            st.plotly_chart(fig_estado, use_container_width=True)

    st.markdown("### Avance por Dirección Regional")

    if "Dirección Regional" in df_metricas.columns:
        resumen_region = (
            df_metricas.groupby("Dirección Regional", dropna=False)
            .agg(
                Registros=("ID", "count"),
                Participantes=("Cantidad Participantes", "sum"),
                Aprobadas=("Estado Validación", lambda x: (x == "Aprobada").sum()),
                Pendientes=("Estado Validación", lambda x: (x == "Pendiente").sum()),
                Rechazadas=("Estado Validación", lambda x: (x == "Rechazada").sum()),
                Observaciones=("Estado Validación", lambda x: (x == "Con observaciones").sum()),
            )
            .reset_index()
            .sort_values("Registros", ascending=False)
        )

        st.dataframe(resumen_region, use_container_width=True, hide_index=True)

        fig_region = px.bar(
            resumen_region,
            x="Dirección Regional",
            y="Registros",
            text="Registros",
            title="Registros visibles por Dirección Regional",
            color_discrete_sequence=[COLOR_AZUL],
        )
        fig_region.update_traces(textposition="outside", marker_line_color=COLOR_DORADO, marker_line_width=1.4)
        fig_region.update_xaxes(tickangle=-25)
        fig_region.update_yaxes(rangemode="tozero", dtick=1)
        fig_region = aplicar_estilo_grafico_institucional(fig_region)
        st.plotly_chart(fig_region, use_container_width=True)

    st.markdown("### Avance por Delegación")

    if "Delegación" in df_metricas.columns:
        resumen_delegacion = (
            df_metricas.groupby(["Dirección Regional", "Delegación"], dropna=False)
            .agg(
                Registros=("ID", "count"),
                Participantes=("Cantidad Participantes", "sum"),
                Aprobadas=("Estado Validación", lambda x: (x == "Aprobada").sum()),
                Pendientes=("Estado Validación", lambda x: (x == "Pendiente").sum()),
                Rechazadas=("Estado Validación", lambda x: (x == "Rechazada").sum()),
                Observaciones=("Estado Validación", lambda x: (x == "Con observaciones").sum()),
            )
            .reset_index()
            .sort_values(["Dirección Regional", "Delegación"])
        )

        st.dataframe(resumen_delegacion, use_container_width=True, hide_index=True)

        top_delegaciones = resumen_delegacion.sort_values("Registros", ascending=False).head(20)
        if not top_delegaciones.empty:
            fig_delegacion = px.bar(
                top_delegaciones,
                x="Delegación",
                y="Registros",
                color="Dirección Regional",
                text="Registros",
                title="Delegaciones con más registros visibles",
            )
            fig_delegacion.update_traces(textposition="outside")
            fig_delegacion.update_xaxes(tickangle=-30)
            fig_delegacion.update_yaxes(rangemode="tozero", dtick=1)
            fig_delegacion = aplicar_estilo_grafico_institucional(fig_delegacion)
            st.plotly_chart(fig_delegacion, use_container_width=True)

    st.markdown("### Detalle incluido en el dashboard")
    mostrar_tabla_resumen_validacion(df_metricas)



# ======================================================
# LECTURA AUTOMÁTICA DE METAS REGIONALES PARA DASHBOARD
# Lee archivos locales tipo DR2 AVANCE JUNIO.xlsx, DR3...
# y muestra metas/avance base según el programa autorizado.
# ======================================================

MESES_META = [
    "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
    "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"
]

MAPA_REGION_ARCHIVO = {
    "DR1 C": "Dirección Regional 1 - San José Central",
    "DR1 N": "Dirección Regional 1 - San José Norte",
    "DR1 S": "Dirección Regional 1 - San José Sur",
    "DR2": "Dirección Regional 2 - Alajuela",
    "DR3": "Dirección Regional 3 - Cartago",
    "DR4": "Dirección Regional 4 - Heredia",
    "DR5": "Dirección Regional 5",
    "DR6": "Dirección Regional 6",
    "DR7": "Dirección Regional 7",
    "DR8": "Dirección Regional 8",
    "DR9": "Dirección Regional 9",
    "DR10": "Dirección Regional 10",
    "DR11": "Dirección Regional 11",
    "DR12": "Dirección Regional 12 - Caribe",
}


def limpiar_numero_meta(valor):
    if pd.isna(valor) or valor == "":
        return 0.0
    if isinstance(valor, str):
        valor = valor.replace("%", "").replace(",", ".").strip()
    try:
        return float(valor)
    except Exception:
        return 0.0


def normalizar_columnas_meta(cols):
    nuevas = []
    for c in cols:
        c = str(c).strip().upper()
        c = re.sub(r"\s+", " ", c)
        nuevas.append(c)
    return nuevas


def obtener_region_desde_archivo_meta(nombre_archivo):
    nombre = str(nombre_archivo).upper().replace("_", " ")
    nombre = re.sub(r"\s+", " ", nombre).strip()
    for clave in sorted(MAPA_REGION_ARCHIVO.keys(), key=len, reverse=True):
        if nombre.startswith(clave):
            return MAPA_REGION_ARCHIVO[clave]
    m = re.match(r"(DR\d+)", nombre)
    if m:
        return MAPA_REGION_ARCHIVO.get(m.group(1), m.group(1))
    return "Región no identificada"


def clasificar_estado_meta(meta, avance):
    meta = limpiar_numero_meta(meta)
    avance = limpiar_numero_meta(avance)
    if meta > 0 and avance >= meta:
        return "Completa"
    if avance > 0:
        return "En avance"
    return "Pendiente"


@st.cache_data(show_spinner=False)
def cargar_metas_regionales_nacionales():
    """
    Carga automáticamente todos los Excel de metas regionales ubicados al mismo
    nivel de la app. No requiere subirlos manualmente desde la interfaz.
    """
    import glob

    patrones = [
        "DR* AVANCE JUNIO.xlsx",
        "DR*AVANCE JUNIO.xlsx",
        "DR* AVANCE JUNIO.xlsm",
        "DR*AVANCE JUNIO.xlsm",
    ]

    archivos = []
    for patron in patrones:
        archivos.extend(glob.glob(patron))

    # Respaldo para pruebas locales dentro del sandbox.
    if not archivos:
        for patron in patrones:
            archivos.extend(glob.glob(os.path.join("/mnt/data", patron)))

    archivos = sorted(list(dict.fromkeys(archivos)))
    registros = []

    for ruta in archivos:
        nombre_archivo = os.path.basename(ruta)
        region = obtener_region_desde_archivo_meta(nombre_archivo)

        try:
            xl = pd.ExcelFile(ruta)
        except Exception:
            continue

        for hoja in xl.sheet_names:
            nombre_hoja = str(hoja).strip()
            if not nombre_hoja or nombre_hoja.upper().startswith("TOTAL"):
                continue

            try:
                raw = pd.read_excel(ruta, sheet_name=hoja, header=None)
            except Exception:
                continue

            if raw.empty:
                continue

            fila_header = None
            for i in range(min(12, len(raw))):
                fila = " ".join(raw.iloc[i].astype(str).fillna("").str.upper().tolist())
                if "PROGRAMA" in fila and "META" in fila:
                    fila_header = i
                    break

            if fila_header is None:
                continue

            try:
                df = pd.read_excel(ruta, sheet_name=hoja, header=fila_header)
            except Exception:
                continue

            if df.empty or len(df.columns) < 4:
                continue

            df.columns = normalizar_columnas_meta(df.columns)
            df = df.dropna(how="all")

            col_programa = df.columns[0]
            col_actividad = df.columns[1]
            col_meta = next((c for c in df.columns if "META" in c), None)
            col_avance = next((c for c in df.columns if "AVANCE" in c), None)

            if not col_meta:
                continue

            if not col_avance:
                df["AVANCE"] = 0
                col_avance = "AVANCE"

            df[col_programa] = df[col_programa].ffill()

            for _, row in df.iterrows():
                programa = str(row.get(col_programa, "")).strip()
                actividad = str(row.get(col_actividad, "")).strip()

                if not actividad or actividad.upper() in ["NAN", "TOTAL", "TOTALES"]:
                    continue
                if not programa or programa.upper() == "NAN":
                    continue

                meta = limpiar_numero_meta(row.get(col_meta, 0))
                avance = limpiar_numero_meta(row.get(col_avance, 0))
                pendiente = max(meta - avance, 0)
                porcentaje = avance / meta if meta else 0

                item = {
                    "Archivo meta": nombre_archivo,
                    "Dirección Regional": region,
                    "Delegación": nombre_hoja,
                    "Programa": programa,
                    "Actividad": actividad,
                    "Meta oficial": meta,
                    "Avance base": avance,
                    "Pendiente base": pendiente,
                    "% avance base": porcentaje,
                    "Estado": clasificar_estado_meta(meta, avance),
                }

                for mes in MESES_META:
                    col_mes = next((c for c in df.columns if c == mes), None)
                    item[mes.title()] = limpiar_numero_meta(row.get(col_mes, 0)) if col_mes else 0

                registros.append(item)

    if not registros:
        return pd.DataFrame(columns=[
            "Archivo meta", "Dirección Regional", "Delegación", "Programa", "Actividad",
            "Meta oficial", "Avance base", "Pendiente base", "% avance base", "Estado"
        ])

    return pd.DataFrame(registros)


def filtrar_metas_por_programa_autorizado(df_metas, programa_autorizado):
    if df_metas is None or df_metas.empty:
        return df_metas
    if not programa_autorizado or "Programa" not in df_metas.columns:
        return pd.DataFrame(columns=df_metas.columns)
    mascara = df_metas["Programa"].apply(lambda x: programa_permitido_por_acceso(x, programa_autorizado))
    return df_metas[mascara].copy()


def aplicar_filtros_dashboard_metas(df_metas):
    if df_metas is None or df_metas.empty:
        return df_metas

    st.markdown("### Filtros del dashboard")
    col1, col2, col3 = st.columns(3)

    with col1:
        regiones = sorted(df_metas["Dirección Regional"].replace("", pd.NA).dropna().astype(str).unique().tolist())
        filtro_region = st.multiselect("Dirección Regional", regiones, key="filtro_meta_region")

    with col2:
        base_delegaciones = df_metas.copy()
        if filtro_region:
            base_delegaciones = base_delegaciones[base_delegaciones["Dirección Regional"].isin(filtro_region)]
        delegaciones = sorted(base_delegaciones["Delegación"].replace("", pd.NA).dropna().astype(str).unique().tolist())
        filtro_delegacion = st.multiselect("Delegación", delegaciones, key="filtro_meta_delegacion")

    with col3:
        programas = sorted(df_metas["Programa"].replace("", pd.NA).dropna().astype(str).unique().tolist())
        filtro_programa = st.multiselect("Programa", programas, key="filtro_meta_programa")
        filtro_estado = st.multiselect("Estado", ["Completa", "En avance", "Pendiente"], key="filtro_meta_estado")

    df_filtrado = df_metas.copy()
    if filtro_region:
        df_filtrado = df_filtrado[df_filtrado["Dirección Regional"].isin(filtro_region)]
    if filtro_delegacion:
        df_filtrado = df_filtrado[df_filtrado["Delegación"].isin(filtro_delegacion)]
    if filtro_programa:
        df_filtrado = df_filtrado[df_filtrado["Programa"].isin(filtro_programa)]
    if filtro_estado:
        df_filtrado = df_filtrado[df_filtrado["Estado"].isin(filtro_estado)]

    st.session_state.filtros_admin_aplicados = {
        "Dirección Regional": filtro_region,
        "Delegación": filtro_delegacion,
        "Programa": filtro_programa,
        "Estado": filtro_estado,
        "Fuente": "Metas regionales locales"
    }

    return df_filtrado


def mostrar_dashboard_metas_nacionales(df_metas, df_registros=None):
    programa_actual = st.session_state.get("admin_programa_autenticado", "")

    if df_metas is None or df_metas.empty:
        st.warning("No se encontraron archivos de metas regionales al mismo nivel de la app.")
        st.info("Verifique que estén cargados archivos como DR2 AVANCE JUNIO.xlsx, DR3 AVANCE JUNIO.xlsx, etc.")
        return

    df_filtrado = aplicar_filtros_dashboard_metas(df_metas)

    if df_filtrado.empty:
        st.info("No hay metas para mostrar con los filtros seleccionados.")
        return

    meta_total = float(df_filtrado["Meta oficial"].sum())
    avance_base = float(df_filtrado["Avance base"].sum())
    pendiente = max(meta_total - avance_base, 0)
    porcentaje = avance_base / meta_total if meta_total else 0
    regiones = df_filtrado["Dirección Regional"].nunique()
    delegaciones = df_filtrado["Delegación"].nunique()

    st.markdown("### Resumen nacional según metas oficiales")
    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric("Programa", programa_actual if programa_actual else "Sin acceso")
    c2.metric("Meta oficial", f"{meta_total:,.0f}")
    c3.metric("Avance base", f"{avance_base:,.0f}")
    c4.metric("Pendiente", f"{pendiente:,.0f}")
    c5.metric("% avance", f"{porcentaje:.1%}")
    c6.metric("Delegaciones", f"{delegaciones:,}")

    st.caption(f"Regiones con metas visibles: {regiones}")

    st.markdown("---")
    col_g1, col_g2 = st.columns(2)

    with col_g1:
        estado = (
            df_filtrado.groupby("Estado", as_index=False)
            .agg(Actividades=("Actividad", "count"), Meta=("Meta oficial", "sum"), Avance=("Avance base", "sum"))
        )
        if not estado.empty:
            fig_estado = px.pie(
                estado,
                names="Estado",
                values="Actividades",
                title="Actividades por estado de avance",
                hole=0.38,
                color="Estado",
                color_discrete_map={"Completa": COLOR_VERDE, "En avance": COLOR_DORADO, "Pendiente": COLOR_AMARILLO},
            )
            fig_estado.update_traces(textinfo="percent+label+value")
            fig_estado = aplicar_estilo_grafico_institucional(fig_estado)
            st.plotly_chart(fig_estado, use_container_width=True)

    with col_g2:
        avance_pend = pd.DataFrame({
            "Categoría": ["Avance base", "Pendiente"],
            "Cantidad": [avance_base, pendiente]
        })
        fig_cumplimiento = px.pie(
            avance_pend,
            names="Categoría",
            values="Cantidad",
            title="Avance base vs pendiente",
            hole=0.38,
            color="Categoría",
            color_discrete_map={"Avance base": COLOR_AZUL, "Pendiente": COLOR_AMARILLO},
        )
        fig_cumplimiento.update_traces(textinfo="percent+label+value")
        fig_cumplimiento = aplicar_estilo_grafico_institucional(fig_cumplimiento)
        st.plotly_chart(fig_cumplimiento, use_container_width=True)

    st.markdown("### Avance por Dirección Regional")
    resumen_region = (
        df_filtrado.groupby("Dirección Regional", as_index=False)
        .agg(
            Actividades=("Actividad", "count"),
            Delegaciones=("Delegación", "nunique"),
            Meta=("Meta oficial", "sum"),
            Avance=("Avance base", "sum"),
            Pendiente=("Pendiente base", "sum"),
        )
    )
    resumen_region["% avance"] = resumen_region.apply(lambda x: x["Avance"] / x["Meta"] if x["Meta"] else 0, axis=1)
    resumen_region = resumen_region.sort_values("Meta", ascending=False)

    vista_region = resumen_region.copy()
    vista_region["% avance"] = vista_region["% avance"].map(lambda x: f"{x:.1%}")
    for col in ["Meta", "Avance", "Pendiente"]:
        vista_region[col] = vista_region[col].map(lambda x: f"{x:,.0f}")
    st.dataframe(vista_region, use_container_width=True, hide_index=True)

    if not resumen_region.empty:
        fig_region = px.bar(
            resumen_region,
            x="Dirección Regional",
            y=["Meta", "Avance", "Pendiente"],
            barmode="group",
            title="Meta, avance base y pendiente por Dirección Regional",
        )
        fig_region.update_xaxes(tickangle=-25)
        fig_region = aplicar_estilo_grafico_institucional(fig_region)
        st.plotly_chart(fig_region, use_container_width=True)

    st.markdown("### Avance por Delegación")
    resumen_delegacion = (
        df_filtrado.groupby(["Dirección Regional", "Delegación"], as_index=False)
        .agg(
            Actividades=("Actividad", "count"),
            Meta=("Meta oficial", "sum"),
            Avance=("Avance base", "sum"),
            Pendiente=("Pendiente base", "sum"),
        )
    )
    resumen_delegacion["% avance"] = resumen_delegacion.apply(lambda x: x["Avance"] / x["Meta"] if x["Meta"] else 0, axis=1)
    resumen_delegacion = resumen_delegacion.sort_values(["Dirección Regional", "Delegación"])

    vista_delegacion = resumen_delegacion.copy()
    vista_delegacion["% avance"] = vista_delegacion["% avance"].map(lambda x: f"{x:.1%}")
    for col in ["Meta", "Avance", "Pendiente"]:
        vista_delegacion[col] = vista_delegacion[col].map(lambda x: f"{x:,.0f}")
    st.dataframe(vista_delegacion, use_container_width=True, hide_index=True)

    st.markdown("### Detalle de metas visibles")
    detalle = df_filtrado.copy()
    detalle["% avance base"] = detalle["% avance base"].map(lambda x: f"{x:.1%}")
    for col in ["Meta oficial", "Avance base", "Pendiente base"]:
        detalle[col] = detalle[col].map(lambda x: f"{x:,.0f}")
    columnas = [
        "Archivo meta", "Dirección Regional", "Delegación", "Programa", "Actividad",
        "Meta oficial", "Avance base", "Pendiente base", "% avance base", "Estado"
    ]
    st.dataframe(detalle[columnas], use_container_width=True, hide_index=True)

    excel_bytes = BytesIO()
    with pd.ExcelWriter(excel_bytes, engine="openpyxl") as writer:
        df_filtrado.to_excel(writer, index=False, sheet_name="DASHBOARD_METAS")
        resumen_region.to_excel(writer, index=False, sheet_name="RESUMEN_REGION")
        resumen_delegacion.to_excel(writer, index=False, sheet_name="RESUMEN_DELEGACION")
    excel_bytes.seek(0)

    st.download_button(
        "⬇️ Descargar dashboard de metas",
        data=excel_bytes.getvalue(),
        file_name=f"DASHBOARD_METAS_{limpiar_nombre_archivo(programa_actual)}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="descarga_dashboard_metas_nacionales"
    )

# ======================================================
# appADMIN.py
# PARTE 4 DE 5 CORREGIDA
# INFORME PDF CON FECHA MANUAL, SIN HORA AUTOMÁTICA,
# GRÁFICOS INSTITUCIONALES E IMAGEN DEL MAPA
# ======================================================


def obtener_estilos_pdf():
    estilos = getSampleStyleSheet()

    estilos.add(ParagraphStyle(
        name="TituloPrincipalPDF",
        parent=estilos["Title"],
        fontName="Helvetica-Bold",
        fontSize=22,
        leading=26,
        alignment=TA_CENTER,
        textColor=colors.HexColor(COLOR_AZUL),
        spaceAfter=14
    ))

    estilos.add(ParagraphStyle(
        name="SubtituloPDF",
        parent=estilos["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=15,
        leading=18,
        alignment=TA_LEFT,
        textColor=colors.HexColor(COLOR_AZUL),
        spaceBefore=10,
        spaceAfter=8
    ))

    estilos.add(ParagraphStyle(
        name="TextoPDF",
        parent=estilos["BodyText"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        alignment=TA_JUSTIFY,
        textColor=colors.HexColor("#222222")
    ))

    estilos.add(ParagraphStyle(
        name="TextoTablaPDF",
        parent=estilos["BodyText"],
        fontName="Helvetica",
        fontSize=7,
        leading=9,
        alignment=TA_LEFT,
        textColor=colors.HexColor("#222222")
    ))

    estilos.add(ParagraphStyle(
        name="TextoTablaCentroPDF",
        parent=estilos["BodyText"],
        fontName="Helvetica",
        fontSize=7,
        leading=9,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#222222")
    ))

    return estilos


def parrafo_pdf(texto, estilo):
    texto = "" if texto is None else str(texto)
    texto = texto.replace("&", "&amp;")
    texto = texto.replace("<", "&lt;")
    texto = texto.replace(">", "&gt;")
    texto = texto.replace("\n", "<br/>")
    return Paragraph(texto, estilo)


def dibujar_encabezado_pie_pdf(canvas, doc):
    canvas.saveState()
    ancho, alto = landscape(letter)

    canvas.setFillColor(colors.HexColor(COLOR_AZUL))
    canvas.rect(0, alto - 1.05 * cm, ancho, 1.05 * cm, fill=1, stroke=0)

    canvas.setFillColor(colors.white)
    canvas.setFont("Helvetica-Bold", 8)
    canvas.drawString(
        1.2 * cm,
        alto - 0.65 * cm,
        "P.U.M.I. 2026 - Informe Nacional de Validación"
    )

    canvas.setFillColor(colors.HexColor(COLOR_DORADO))
    canvas.rect(0, 0.95 * cm, ancho, 0.08 * cm, fill=1, stroke=0)

    canvas.setFillColor(colors.HexColor("#444444"))
    canvas.setFont("Helvetica", 8)

    canvas.drawRightString(
        ancho - 1.2 * cm,
        0.55 * cm,
        f"Página {doc.page}"
    )

    canvas.restoreState()


def obtener_logo_pdf(ruta, ancho=4.0 * cm, alto=2.0 * cm):
    if not os.path.exists(ruta):
        return ""

    try:
        img = Image(ruta, width=ancho, height=alto)
        img.hAlign = "CENTER"
        return img
    except Exception:
        return ""


def convertir_figura_plotly_a_imagen(fig):
    try:
        imagen_bytes = fig.to_image(
            format="png",
            width=1200,
            height=650,
            scale=2
        )
        return BytesIO(imagen_bytes)
    except Exception:
        return None


def crear_figuras_pdf_admin(df):
    figuras = []
    df_metricas = limpiar_dataframe_metricas_admin(df)

    if df_metricas.empty:
        return figuras

    if "Estado Validación" in df_metricas.columns:
        conteo_estado = (
            df_metricas
            .groupby("Estado Validación")
            .size()
            .reset_index(name="Cantidad")
            .sort_values("Cantidad", ascending=False)
        )

        if not conteo_estado.empty:
            fig_estado = px.pie(
                conteo_estado,
                names="Estado Validación",
                values="Cantidad",
                title="Distribución por estado de validación",
                hole=0.35,
                color="Estado Validación",
                color_discrete_map={
                    "Pendiente": COLOR_AMARILLO,
                    "Aprobada": COLOR_VERDE,
                    "Rechazada": COLOR_ROJO,
                    "Con observaciones": COLOR_DORADO
                }
            )

            fig_estado.update_traces(
                textinfo="percent+label+value",
                marker=dict(line=dict(color="white", width=2))
            )

            fig_estado = aplicar_estilo_grafico_institucional(fig_estado)

            figuras.append({
                "titulo": "Gráfico 1. Distribución por estado de validación",
                "descripcion": "Muestra la proporción de actividades según su estado nacional de validación.",
                "figura": fig_estado
            })

    if "Programa" in df_metricas.columns:
        conteo_programa = (
            df_metricas
            .groupby("Programa")
            .size()
            .reset_index(name="Cantidad")
            .sort_values("Cantidad", ascending=False)
        )

        if not conteo_programa.empty:
            fig_programa = px.bar(
                conteo_programa,
                x="Programa",
                y="Cantidad",
                title="Cantidad de registros por programa",
                text="Cantidad",
                color_discrete_sequence=[COLOR_DORADO]
            )

            fig_programa.update_traces(
                marker_line_color=COLOR_AZUL,
                marker_line_width=1.5,
                textposition="outside",
                textfont=dict(color=COLOR_AZUL, size=14)
            )

            fig_programa.update_yaxes(rangemode="tozero", dtick=1)
            fig_programa = aplicar_estilo_grafico_institucional(fig_programa)

            figuras.append({
                "titulo": "Gráfico 2. Cantidad de registros por programa",
                "descripcion": "Permite identificar cuáles programas concentran mayor cantidad de actividades.",
                "figura": fig_programa
            })

    if "Dirección Regional" in df_metricas.columns:
        conteo_region = (
            df_metricas
            .groupby("Dirección Regional")
            .size()
            .reset_index(name="Cantidad")
            .sort_values("Cantidad", ascending=False)
        )

        if not conteo_region.empty:
            fig_region = px.bar(
                conteo_region,
                x="Dirección Regional",
                y="Cantidad",
                title="Cantidad de registros por Dirección Regional",
                text="Cantidad",
                color_discrete_sequence=[COLOR_AZUL]
            )

            fig_region.update_traces(
                marker_line_color=COLOR_DORADO,
                marker_line_width=1.5,
                textposition="outside",
                textfont=dict(color=COLOR_AZUL, size=14)
            )

            fig_region.update_yaxes(rangemode="tozero", dtick=1)
            fig_region.update_xaxes(tickangle=-30)
            fig_region = aplicar_estilo_grafico_institucional(fig_region)

            figuras.append({
                "titulo": "Gráfico 3. Cantidad de registros por Dirección Regional",
                "descripcion": "Muestra la distribución de registros según Dirección Regional.",
                "figura": fig_region
            })

    if "Delegación" in df_metricas.columns:
        conteo_delegacion = (
            df_metricas
            .groupby("Delegación")
            .size()
            .reset_index(name="Cantidad")
            .sort_values("Cantidad", ascending=False)
            .head(15)
        )

        if not conteo_delegacion.empty:
            fig_delegacion = px.bar(
                conteo_delegacion,
                x="Delegación",
                y="Cantidad",
                title="Top delegaciones por cantidad de registros",
                text="Cantidad",
                color_discrete_sequence=[COLOR_DORADO]
            )

            fig_delegacion.update_traces(
                marker_line_color=COLOR_AZUL,
                marker_line_width=1.5,
                textposition="outside",
                textfont=dict(color=COLOR_AZUL, size=14)
            )

            fig_delegacion.update_yaxes(rangemode="tozero", dtick=1)
            fig_delegacion.update_xaxes(tickangle=-30)
            fig_delegacion = aplicar_estilo_grafico_institucional(fig_delegacion)

            figuras.append({
                "titulo": "Gráfico 4. Top delegaciones por cantidad de registros",
                "descripcion": "Resume las delegaciones con mayor cantidad de registros.",
                "figura": fig_delegacion
            })

    return figuras


def agregar_graficos_automaticos_pdf(elementos, df, estilos):
    figuras = crear_figuras_pdf_admin(df)

    if not figuras:
        return

    elementos.append(PageBreak())
    elementos.append(parrafo_pdf("4. Gráficos automáticos del dashboard", estilos["SubtituloPDF"]))

    elementos.append(parrafo_pdf(
        "Los siguientes gráficos fueron generados automáticamente desde los datos filtrados en la app administrativa.",
        estilos["TextoPDF"]
    ))

    elementos.append(Spacer(1, 0.3 * cm))

    for item in figuras:
        imagen = convertir_figura_plotly_a_imagen(item["figura"])

        if imagen is None:
            continue

        bloque = [
            parrafo_pdf(item["titulo"], estilos["SubtituloPDF"]),
            parrafo_pdf(item["descripcion"], estilos["TextoPDF"]),
            Spacer(1, 0.2 * cm)
        ]

        img = Image(imagen, width=22 * cm, height=11 * cm)
        img.hAlign = "CENTER"
        bloque.append(img)
        bloque.append(Spacer(1, 0.4 * cm))

        elementos.append(KeepTogether(bloque))


def agregar_imagen_mapa_pdf(elementos, archivo_mapa, descripcion_mapa, estilos):
    if archivo_mapa is None:
        return

    try:
        elementos.append(PageBreak())
        elementos.append(parrafo_pdf("5. Imagen del mapa de referencia", estilos["SubtituloPDF"]))

        texto = (
            descripcion_mapa
            if descripcion_mapa and str(descripcion_mapa).strip()
            else "La siguiente imagen muestra la distribución territorial de las actividades filtradas en la app administrativa."
        )

        elementos.append(parrafo_pdf(texto, estilos["TextoPDF"]))
        elementos.append(Spacer(1, 0.3 * cm))

        imagen_bytes = BytesIO(archivo_mapa.read())
        img = Image(imagen_bytes, width=22 * cm, height=11.5 * cm)
        img.hAlign = "CENTER"

        elementos.append(img)
        elementos.append(Spacer(1, 0.4 * cm))

    except Exception:
        elementos.append(parrafo_pdf("No se pudo agregar la imagen del mapa.", estilos["TextoPDF"]))


def construir_portada_pdf(elementos, df, estilos, fecha_informe=""):
    logo_min = obtener_logo_pdf(LOGO_MINISTERIO, 5.0 * cm, 1.7 * cm)
    logo_pumi = obtener_logo_pdf(LOGO_PUMI, 4.0 * cm, 2.0 * cm)
    logo_fp = obtener_logo_pdf(LOGO_FUERZA_PUBLICA, 5.0 * cm, 2.0 * cm)

    tabla_logos = Table(
        [[logo_min, logo_pumi, logo_fp]],
        colWidths=[8 * cm, 8 * cm, 8 * cm]
    )

    tabla_logos.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor(COLOR_DORADO)),
        ("BACKGROUND", (0, 0), (-1, -1), colors.white),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))

    elementos.append(tabla_logos)
    elementos.append(Spacer(1, 0.7 * cm))

    elementos.append(parrafo_pdf(
        "INFORME NACIONAL DE VALIDACIÓN DE ACTIVIDADES",
        estilos["TituloPrincipalPDF"]
    ))

    elementos.append(parrafo_pdf(
        "Sistema P.U.M.I. 2026",
        estilos["TituloPrincipalPDF"]
    ))

    total = len(df)

    participantes = int(
        pd.to_numeric(
            df["Cantidad Participantes"],
            errors="coerce"
        ).fillna(0).sum()
    ) if not df.empty and "Cantidad Participantes" in df.columns else 0

    regiones = ", ".join(
        df["Dirección Regional"]
        .replace("", pd.NA)
        .dropna()
        .astype(str)
        .unique()
        .tolist()
    ) if "Dirección Regional" in df.columns and not df.empty else "Sin datos"

    delegaciones = ", ".join(
        df["Delegación"]
        .replace("", pd.NA)
        .dropna()
        .astype(str)
        .unique()
        .tolist()
    ) if "Delegación" in df.columns and not df.empty else "Sin datos"

    elementos.append(parrafo_pdf(
        "El presente informe consolida la información de actividades registradas en el sistema PUMI 2026, incorporando su estado nacional de validación, observaciones, métricas principales y elementos de análisis.",
        estilos["TextoPDF"]
    ))

    elementos.append(Spacer(1, 0.5 * cm))

    datos_portada = [
        ["Total de registros analizados", str(total)],
        ["Total de participantes", str(participantes)],
        ["Dirección Regional", regiones if regiones else "Varias"],
        ["Delegación", delegaciones if delegaciones else "Varias"],
        ["Fecha del informe", fecha_informe if fecha_informe else "No indicada"]
    ]

    tabla = Table(datos_portada, colWidths=[7 * cm, 17 * cm])

    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor(COLOR_AZUL)),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.white),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#BFBFBF")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (1, 0), (1, -1), [colors.white, colors.HexColor("#F4F7FB")]),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))

    elementos.append(tabla)
    elementos.append(PageBreak())


def construir_resumen_pdf(elementos, df, estilos):
    elementos.append(parrafo_pdf("1. Resumen ejecutivo", estilos["SubtituloPDF"]))

    elementos.append(parrafo_pdf(
        "Esta sección presenta una síntesis de los registros incluidos en el informe, considerando el total de actividades, participantes y estado de validación administrativa.",
        estilos["TextoPDF"]
    ))

    elementos.append(Spacer(1, 0.3 * cm))

    total = len(df)

    participantes = int(
        pd.to_numeric(
            df["Cantidad Participantes"],
            errors="coerce"
        ).fillna(0).sum()
    ) if "Cantidad Participantes" in df.columns and not df.empty else 0

    aprobadas = len(df[df["Estado Validación"] == "Aprobada"]) if not df.empty else 0
    pendientes = len(df[df["Estado Validación"] == "Pendiente"]) if not df.empty else 0
    rechazadas = len(df[df["Estado Validación"] == "Rechazada"]) if not df.empty else 0
    observadas = len(df[df["Estado Validación"] == "Con observaciones"]) if not df.empty else 0

    datos = [
        ["Indicador", "Resultado"],
        ["Total de actividades", total],
        ["Total de participantes", participantes],
        ["Actividades aprobadas", aprobadas],
        ["Actividades pendientes", pendientes],
        ["Actividades rechazadas", rechazadas],
        ["Actividades con observaciones", observadas]
    ]

    tabla = Table(datos, colWidths=[10 * cm, 5 * cm])

    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(COLOR_AZUL)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#BFBFBF")),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (1, 1), (1, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F7FB")]),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))

    elementos.append(tabla)
    elementos.append(Spacer(1, 0.4 * cm))


def construir_filtros_pdf(elementos, estilos):
    filtros = st.session_state.get("filtros_admin_aplicados", {})

    elementos.append(parrafo_pdf("2. Filtros aplicados", estilos["SubtituloPDF"]))

    if not filtros:
        elementos.append(parrafo_pdf("No se registraron filtros específicos.", estilos["TextoPDF"]))
        return

    datos = [["Filtro", "Valor aplicado"]]

    for clave, valor in filtros.items():
        if isinstance(valor, list):
            valor = ", ".join(valor) if valor else "Todos"
        elif not valor:
            valor = "Todos"

        datos.append([clave, valor])

    tabla = Table(datos, colWidths=[6 * cm, 18 * cm])

    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(COLOR_AZUL)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#BFBFBF")),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F7FB")]),
    ]))

    elementos.append(tabla)
    elementos.append(Spacer(1, 0.4 * cm))


def construir_tabla_agrupada_pdf(elementos, df, columna, titulo, estilos):
    if df.empty or columna not in df.columns:
        return

    elementos.append(parrafo_pdf(titulo, estilos["SubtituloPDF"]))

    resumen = (
        df.groupby(columna)
        .size()
        .reset_index(name="Cantidad")
        .sort_values("Cantidad", ascending=False)
    )

    datos = [["Categoría", "Cantidad"]]

    for _, row in resumen.iterrows():
        datos.append([
            parrafo_pdf(row[columna], estilos["TextoTablaPDF"]),
            parrafo_pdf(row["Cantidad"], estilos["TextoTablaCentroPDF"])
        ])

    tabla = Table(datos, colWidths=[18 * cm, 4 * cm], repeatRows=1)

    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(COLOR_AZUL)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#BFBFBF")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (1, 1), (1, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F7FB")]),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))

    elementos.append(tabla)
    elementos.append(Spacer(1, 0.35 * cm))


def construir_tabla_detalle_pdf(elementos, df, estilos):
    elementos.append(PageBreak())
    elementos.append(parrafo_pdf("6. Detalle de actividades validadas", estilos["SubtituloPDF"]))

    columnas = [
        "ID",
        "Fecha Actividad",
        "Dirección Regional",
        "Delegación",
        "Programa",
        "Actividad",
        "Cantidad Participantes",
        "Estado Verificación Regional",
        "Observaciones Verificación Regional",
        "Coordinador Regional",
        "Fecha Verificación Regional",
        "Estado Validación",
        "Observaciones Validación"
    ]

    columnas = [c for c in columnas if c in df.columns]
    datos = [[parrafo_pdf(c, estilos["TextoTablaCentroPDF"]) for c in columnas]]

    for _, row in df.iterrows():
        datos.append([
            parrafo_pdf(row.get(col, ""), estilos["TextoTablaPDF"])
            for col in columnas
        ])

    anchos = [
        1.2 * cm,
        2.1 * cm,
        3.5 * cm,
        3.5 * cm,
        2.2 * cm,
        5.8 * cm,
        2.0 * cm,
        2.6 * cm,
        4.8 * cm
    ][:len(columnas)]

    tabla = Table(datos, colWidths=anchos, repeatRows=1)

    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(COLOR_AZUL)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#BFBFBF")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F7FB")]),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))

    elementos.append(tabla)


def generar_pdf_validacion(
    df,
    fecha_informe="",
    incluir_tabla_detalle=True,
    incluir_graficos=True,
    imagen_mapa=None,
    descripcion_mapa=""
):
    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=1.1 * cm,
        leftMargin=1.1 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.3 * cm
    )

    estilos = obtener_estilos_pdf()
    elementos = []

    construir_portada_pdf(elementos, df, estilos, fecha_informe)
    construir_resumen_pdf(elementos, df, estilos)
    construir_filtros_pdf(elementos, estilos)

    elementos.append(PageBreak())
    elementos.append(parrafo_pdf("3. Resumen analítico", estilos["SubtituloPDF"]))

    elementos.append(parrafo_pdf(
        "En este apartado se presenta la distribución de los registros según verificación regional, estado de validación nacional, programa, Dirección Regional y delegación.",
        estilos["TextoPDF"]
    ))

    elementos.append(Spacer(1, 0.3 * cm))

    construir_tabla_agrupada_pdf(elementos, df, "Estado Verificación Regional", "3.1 Distribución por verificación regional", estilos)
    construir_tabla_agrupada_pdf(elementos, df, "Estado Validación", "3.2 Distribución por validación nacional", estilos)
    construir_tabla_agrupada_pdf(elementos, df, "Programa", "3.3 Distribución por programa", estilos)
    construir_tabla_agrupada_pdf(elementos, df, "Dirección Regional", "3.3 Distribución por Dirección Regional", estilos)
    construir_tabla_agrupada_pdf(elementos, df, "Delegación", "3.4 Distribución por delegación", estilos)

    if incluir_graficos:
        agregar_graficos_automaticos_pdf(elementos, df, estilos)

    if imagen_mapa is not None:
        agregar_imagen_mapa_pdf(
            elementos,
            imagen_mapa,
            descripcion_mapa,
            estilos
        )

    if incluir_tabla_detalle:
        construir_tabla_detalle_pdf(elementos, df, estilos)

    doc.build(
        elementos,
        onFirstPage=dibujar_encabezado_pie_pdf,
        onLaterPages=dibujar_encabezado_pie_pdf
    )

    buffer.seek(0)
    return buffer.getvalue()


def modulo_informe_pdf(df_filtrado):
    st.markdown("## Generar informe PDF de validación")

    if df_filtrado.empty:
        st.info("No hay datos filtrados para generar informe.")
        return

    st.markdown(
        """
        <div class="card-admin">
            <div class="texto-admin">
                Este módulo genera un informe PDF nacional con portada,
                resumen ejecutivo, filtros aplicados, cuadros analíticos,
                gráficos institucionales, imagen del mapa y detalle de actividades.
            </div>
        </div>
        """,
        unsafe_allow_html=True
    )

    st.markdown("### Datos del informe")

    fecha_informe = st.text_input(
        "Fecha del informe",
        value=date.today().strftime("%d/%m/%Y")
    )

    st.markdown("### Opciones del informe")

    col1, col2 = st.columns(2)

    with col1:
        incluir_detalle = st.checkbox(
            "Incluir tabla detallada de actividades",
            value=True
        )

    with col2:
        incluir_graficos = st.checkbox(
            "Incluir gráficos automáticos",
            value=True
        )

    st.markdown("### Imagen del mapa")

    imagen_mapa = st.file_uploader(
        "Cargar imagen del mapa filtrado para incluirla en el informe",
        type=["png", "jpg", "jpeg"],
        key="upload_mapa_pdf"
    )

    descripcion_mapa = st.text_area(
        "Descripción de la imagen del mapa",
        value=(
            "La imagen muestra la distribución territorial de las actividades "
            "filtradas en el panel nacional, según los criterios seleccionados "
            "para la generación del informe."
        )
    )

    if st.button("📄 Generar informe PDF"):
        pdf_bytes = generar_pdf_validacion(
            df=df_filtrado,
            fecha_informe=fecha_informe,
            incluir_tabla_detalle=incluir_detalle,
            incluir_graficos=incluir_graficos,
            imagen_mapa=imagen_mapa,
            descripcion_mapa=descripcion_mapa
        )

        nombre_pdf = generar_nombre_reporte(df_filtrado, tipo="PDF")

        st.success("Informe PDF generado correctamente.")

        st.download_button(
            "⬇️ Descargar informe PDF",
            data=pdf_bytes,
            file_name=nombre_pdf,
            mime="application/pdf",
            key="descargar_pdf_validacion"
        )
# ======================================================
# appADMIN.py
# PARTE 5 DE 5 CORREGIDA
# FLUJO PRINCIPAL DE LA APP ADMIN:
# SIDEBAR, CARGA DE EXCEL, MENÚ, DASHBOARD,
# VALIDACIÓN INDIVIDUAL, INFORME PDF CON FECHA MANUAL,
# IMAGEN DEL MAPA Y DESCARGA GLOBAL
# ======================================================


# ======================================================
# SIDEBAR
# ======================================================

mostrar_logo_sidebar()

st.sidebar.markdown("## Coordinadores Nacionales PUMI 2026")

st.sidebar.markdown(
    """
    <div style="
        background:#FFFFFF;
        padding:14px;
        border-radius:14px;
        border-left:6px solid #B88A2A;
        box-shadow:0px 4px 12px rgba(0,0,0,0.10);
        margin-bottom:15px;
        font-size:14px;
    ">
    Cargue el Excel verificado por la Dirección Regional para validar,
    analizar y generar informes nacionales.
    </div>
    """,
    unsafe_allow_html=True
)


# ======================================================
# CARGA DEL EXCEL PUMI
# Permite cargar uno o varios Excel regionales verificados.
# ======================================================

archivos_admin = st.sidebar.file_uploader(
    "Subir Excel regional verificado",
    type=["xlsx"],
    accept_multiple_files=True
)

if archivos_admin:
    nombres_archivos = " | ".join([archivo.name for archivo in archivos_admin])

    if st.session_state.archivo_admin_nombre != nombres_archivos:
        dataframes_cargados = []
        archivos_con_error = []

        for archivo_excel in archivos_admin:
            df_cargado = cargar_excel_pumi_admin(archivo_excel)

            if not df_cargado.empty:
                if "Archivo Origen" in df_cargado.columns:
                    df_cargado["Archivo Origen"] = df_cargado["Archivo Origen"].replace("", archivo_excel.name)
                else:
                    df_cargado["Archivo Origen"] = archivo_excel.name

                dataframes_cargados.append(df_cargado)
            else:
                archivos_con_error.append(archivo_excel.name)

        if dataframes_cargados:
            df_unificado = pd.concat(dataframes_cargados, ignore_index=True)
            df_unificado = preparar_dataframe_admin(df_unificado)

            st.session_state.df_admin = df_unificado
            st.session_state.archivo_admin_nombre = nombres_archivos
            st.session_state.admin_acceso_programa = False
            st.session_state.admin_programa_autenticado = ""
            st.session_state.admin_nombre_acceso = ""
            st.sidebar.success(f"{len(dataframes_cargados)} Excel cargado(s) correctamente.")

            if archivos_con_error:
                st.sidebar.warning("No se pudieron cargar estos archivos: " + ", ".join(archivos_con_error))
        else:
            st.sidebar.warning("Ningún Excel contiene registros válidos.")


# ======================================================
# ENCABEZADO GENERAL
# ======================================================

mostrar_encabezado_institucional()
mostrar_titulo_admin()

if not st.session_state.get("admin_acceso_programa", False):
    mostrar_login_coordinador_nacional()
    st.stop()

# ======================================================
# DATAFRAME BASE Y CONTROL DE ACCESO POR PROGRAMA
# ======================================================

df_admin_completo = st.session_state.df_admin.copy()

programa_autorizado = st.session_state.get("admin_programa_autenticado", "")
st.sidebar.markdown("---")
st.sidebar.markdown("### Acceso coordinador nacional")
st.sidebar.success(f"Acceso activo: {programa_autorizado}")

if st.sidebar.button("🔒 Cerrar sesión", key="cerrar_sesion_coordinador_nacional"):
    st.session_state.admin_acceso_programa = False
    st.session_state.admin_programa_autenticado = ""
    st.session_state.admin_nombre_acceso = ""
    st.session_state.df_admin = pd.DataFrame(columns=ENCABEZADOS_COMPLETOS)
    st.session_state.archivo_admin_nombre = ""
    st.rerun()

if df_admin_completo.empty:
    df_admin = df_admin_completo
    st.sidebar.info("Cargue uno o varios Excel regionales para ver la información del programa.")
else:
    df_admin = filtrar_dataframe_por_programa_admin(
        df_admin_completo,
        programa_autorizado
    )

    st.sidebar.info(
        f"Registros visibles para {programa_autorizado}: {len(df_admin)}"
    )

# ======================================================
# MENÚ PRINCIPAL
# ======================================================

opciones_menu_coord = [
    "Inicio",
    "Consulta y filtros",
    "Validación de actividades",
    "Editar o eliminar registros",
    "Dashboard",
    "Informe PDF"
]

if "menu_destino_coord" in st.session_state:
    destino = st.session_state.pop("menu_destino_coord")
    if destino in opciones_menu_coord:
        st.session_state["menu_coord"] = destino

if "menu_coord" not in st.session_state:
    st.session_state["menu_coord"] = "Inicio"

menu_admin = st.sidebar.radio(
    "Menú coordinadores nacionales",
    opciones_menu_coord,
    key="menu_coord"
)


# ======================================================
# INICIO
# ======================================================

if menu_admin == "Inicio":

    st.markdown(
        """
        <div class="card-admin">
            <div class="subtitulo-admin">
                Bienvenido a Coordinadores Nacionales PUMI 2026
            </div>
            <div class="texto-admin">
                Esta aplicación permite cargar el Excel verificado por la Dirección Regional,
                revisar los registros previamente verificados, aplicar validaciones administrativas de forma
                individual, consultar datos mediante filtros, visualizar gráficos y mapas,
                descargar un Excel validado con colores por estado y generar un informe
                PDF formal con gráficos institucionales e imagen del mapa.
            </div>
        </div>
        """,
        unsafe_allow_html=True
    )

    st.markdown("### Accesos rápidos")
    col_ac1, col_ac2, col_ac3 = st.columns(3)
    with col_ac1:
        if st.button("🔎 Consulta y filtros", use_container_width=True, key="inicio_ir_consulta_coord"):
            st.session_state.menu_destino_coord = "Consulta y filtros"
            st.rerun()
    with col_ac2:
        if st.button("✅ Validación de actividades", use_container_width=True, key="inicio_ir_validacion_coord"):
            st.session_state.menu_destino_coord = "Validación de actividades"
            st.rerun()
    with col_ac3:
        if st.button("📊 Dashboard", use_container_width=True, key="inicio_ir_dashboard_coord"):
            st.session_state.menu_destino_coord = "Dashboard"
            st.rerun()

    col_ac4, col_ac5, col_ac6 = st.columns(3)
    with col_ac4:
        if st.button("✏️ Editar o eliminar", use_container_width=True, key="inicio_ir_editar_coord"):
            st.session_state.menu_destino_coord = "Editar o eliminar registros"
            st.rerun()
    with col_ac5:
        if st.button("📄 Informe PDF", use_container_width=True, key="inicio_ir_pdf_coord"):
            st.session_state.menu_destino_coord = "Informe PDF"
            st.rerun()
    with col_ac6:
        if st.button("🏠 Inicio", use_container_width=True, key="inicio_ir_inicio_coord"):
            st.session_state.menu_destino_coord = "Inicio"
            st.rerun()

    st.markdown("---")

    if df_admin.empty:
        st.info("Suba uno o varios Excel regionales verificados desde el panel lateral para iniciar.")
    else:
        mostrar_metricas_admin(df_admin)

        total_verificadas = len(df_admin[df_admin["Estado Verificación Regional"].astype(str) == "Verificada para envío"])
        total_no_verificadas = len(df_admin) - total_verificadas
        st.info(f"Actividades listas para validación nacional: {total_verificadas}. Actividades no habilitadas por verificación regional: {total_no_verificadas}.")

        st.markdown("### Vista rápida de registros cargados")

        mostrar_tabla_resumen_validacion(df_admin)

        st.markdown("### Descargar Excel nacional")

        boton_descargar_excel_admin(
            df_admin,
            key="descarga_inicio_admin"
        )


# ======================================================
# CONSULTA Y FILTROS
# ======================================================

elif menu_admin == "Consulta y filtros":

    st.markdown("## Consulta general de registros")

    if df_admin.empty:
        st.info("Debe cargar primero uno o varios Excel regionales verificados.")
    else:
        df_filtrado = aplicar_filtros_admin(df_admin)

        st.markdown("---")
        mostrar_resumen_validacion(df_filtrado)

        st.markdown("### Registros filtrados")

        mostrar_tabla_resumen_validacion(df_filtrado)

        st.markdown("### Descargar Excel filtrado validado")

        boton_descargar_excel_admin(
            df_filtrado,
            key="descarga_consulta_filtrada"
        )


# ======================================================
# VALIDACIÓN DE ACTIVIDADES
# ======================================================

elif menu_admin == "Validación de actividades":

    st.markdown("## Módulo de validación administrativa")

    if df_admin.empty:
        st.info("Debe cargar primero uno o varios Excel regionales verificados.")
    else:
        df_filtrado = aplicar_filtros_admin(df_admin)

        st.markdown("---")

        modulo_validacion_actividades(df_filtrado)

        st.markdown("---")
        st.markdown("### Descargar Excel con validaciones")

        df_admin_actualizado_programa = filtrar_dataframe_por_programa_admin(
            st.session_state.df_admin,
            st.session_state.get("admin_programa_autenticado", "")
        )

        boton_descargar_excel_admin(
            df_admin_actualizado_programa,
            key="descarga_validacion_admin"
        )


# ======================================================
# EDITAR O ELIMINAR REGISTROS
# ======================================================

elif menu_admin == "Editar o eliminar registros":

    st.markdown("## Editar o eliminar registros")

    if df_admin.empty:
        st.info("Debe cargar primero uno o varios Excel regionales verificados.")
    else:
        df_filtrado = aplicar_filtros_admin(df_admin)

        st.markdown("---")

        modulo_editar_eliminar_registros(df_filtrado)

        st.markdown("---")
        st.markdown("### Descargar Excel actualizado")

        df_admin_actualizado_programa = filtrar_dataframe_por_programa_admin(
            st.session_state.df_admin,
            st.session_state.get("admin_programa_autenticado", "")
        )

        boton_descargar_excel_admin(
            df_admin_actualizado_programa,
            key="descarga_edicion_admin"
        )


# ======================================================
# DASHBOARD
# ======================================================

elif menu_admin == "Dashboard":

    st.markdown("## Dashboard nacional")

    programa_actual = st.session_state.get("admin_programa_autenticado", "")
    df_metas_todas = cargar_metas_regionales_nacionales()
    df_metas_programa = filtrar_metas_por_programa_autorizado(
        df_metas_todas,
        programa_actual
    )

    if not df_metas_programa.empty:
        st.info("Dashboard generado automáticamente desde los Excel de metas regionales cargados en el repositorio.")
        mostrar_dashboard_metas_nacionales(df_metas_programa, df_admin)

        if not df_admin.empty:
            with st.expander("Ver mapa de registros PUMI cargados para este programa", expanded=False):
                df_filtrado_registros = aplicar_filtros_admin(df_admin)
                mostrar_mapa_admin(
                    df_filtrado_registros,
                    key="mapa_dashboard_admin"
                )
    else:
        st.info("No se encontraron metas visibles para este programa en los Excel regionales. Si cargó un Excel verificado, se mostrará el dashboard de registros.")

        if df_admin.empty:
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Programa", programa_actual if programa_actual else "Sin acceso")
            c2.metric("Registros", 0)
            c3.metric("Regiones", 0)
            c4.metric("Delegaciones", 0)
        else:
            df_filtrado = aplicar_filtros_admin(df_admin)
            st.markdown("---")
            mostrar_dashboard_coordinador_nacional(df_filtrado)
            st.markdown("---")
            mostrar_mapa_admin(
                df_filtrado,
                key="mapa_dashboard_admin"
            )
            st.markdown("---")
            st.markdown("### Descargar Excel del dashboard filtrado")
            boton_descargar_excel_admin(
                df_filtrado,
                key="descarga_dashboard_filtrado"
            )


# ======================================================
# INFORME PDF
# ======================================================

elif menu_admin == "Informe PDF":

    st.markdown("## Informe PDF nacional")

    if df_admin.empty:
        st.info("Debe cargar primero uno o varios Excel regionales verificados.")
    else:
        df_filtrado = aplicar_filtros_admin(df_admin)

        st.markdown("---")

        mostrar_metricas_admin(df_filtrado)

        st.markdown("---")

        st.markdown("### Vista de apoyo para el informe")

        with st.expander("Ver gráficos y mapa antes de generar el PDF", expanded=False):
            mostrar_graficos_admin(df_filtrado)

            st.markdown("---")

            mostrar_mapa_admin(
                df_filtrado,
                key="mapa_previo_informe_pdf"
            )

        st.markdown("---")

        modulo_informe_pdf(df_filtrado)


# ======================================================
# DESCARGA GLOBAL Y LIMPIEZA EN SIDEBAR
# ======================================================

st.sidebar.markdown("---")
st.sidebar.markdown("### Opciones coordinadores nacionales")

if df_admin.empty:
    st.sidebar.info("No hay datos cargados.")
else:
    nombre_excel_admin = generar_nombre_reporte(
        df_admin,
        tipo="XLSX"
    )

    st.sidebar.download_button(
        "⬇️ Descargar Excel del programa",
        data=convertir_excel_admin(df_admin),
        file_name=nombre_excel_admin,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="descarga_sidebar_admin"
    )

    if st.sidebar.button("🧹 Limpiar datos cargados"):
        st.session_state.df_admin = pd.DataFrame(columns=ENCABEZADOS_COMPLETOS)
        st.session_state.archivo_admin_nombre = ""
        st.session_state.filtros_admin_aplicados = {}
        st.session_state.mapa_imagen_referencia = None
        st.session_state.grafico_imagen_referencia = None
        st.session_state.admin_acceso_programa = False
        st.session_state.admin_programa_autenticado = ""
        st.sidebar.success("Datos limpiados correctamente.")
        st.rerun()
